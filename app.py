# app.py
from __future__ import annotations

import os, csv, json
from datetime import datetime, date, timedelta
from io import BytesIO
from functools import wraps

import pandas as pd
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from flask import (
    Flask, render_template, request, redirect, url_for,
    flash, send_file, jsonify, session
)
from werkzeug.security import generate_password_hash, check_password_hash
from sqlalchemy import func, text
from sqlalchemy.sql import bindparam
from dotenv import load_dotenv

from utils import (
    now_thai, to_thai_be, to_be_date_str, TH_TZ, current_be_year,
    normalize_platform, sla_text, compute_due_date
)
from models import db, Shop, Product, Stock, Sales, OrderLine, User
from importers import import_products, import_stock, import_sales, import_orders
from allocation import compute_allocation

# โหลด environment variables จากไฟล์ .env (สำหรับ Local Development)
load_dotenv()

APP_NAME = os.environ.get("APP_NAME", "VNIX Order Management")


def get_google_credentials():
    """
    โหลด Google Service Account Credentials จาก Environment Variables หรือไฟล์

    สำหรับ Production (Railway):
    - ตั้งค่า Environment Variables ใน Railway Dashboard
    - ใช้ GOOGLE_CREDENTIALS_JSON (JSON string ทั้งหมด) หรือ
    - ใช้ตัวแปรแยก: GOOGLE_PROJECT_ID, GOOGLE_PRIVATE_KEY, GOOGLE_CLIENT_EMAIL, ฯลฯ

    สำหรับ Local Development:
    - วางไฟล์ credentials.json ในโฟลเดอร์โปรเจกต์
    """
    scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']

    # ลองอ่านจาก Environment Variable (JSON string ทั้งก้อน)
    google_creds_json = os.environ.get('GOOGLE_CREDENTIALS_JSON')
    if google_creds_json:
        try:
            creds_dict = json.loads(google_creds_json)
            return ServiceAccountCredentials.from_json_keyfile_dict(creds_dict, scope)
        except json.JSONDecodeError as e:
            raise RuntimeError(f"GOOGLE_CREDENTIALS_JSON ไม่ถูกต้อง: {e}")

    # ลองสร้างจาก Environment Variables แยก
    if os.environ.get('GOOGLE_PRIVATE_KEY'):
        creds_dict = {
            "type": "service_account",
            "project_id": os.environ.get('GOOGLE_PROJECT_ID'),
            "private_key_id": os.environ.get('GOOGLE_PRIVATE_KEY_ID'),
            "private_key": os.environ.get('GOOGLE_PRIVATE_KEY', '').replace('\\n', '\n'),
            "client_email": os.environ.get('GOOGLE_CLIENT_EMAIL'),
            "client_id": os.environ.get('GOOGLE_CLIENT_ID'),
            "auth_uri": os.environ.get('GOOGLE_AUTH_URI', 'https://accounts.google.com/o/oauth2/auth'),
            "token_uri": os.environ.get('GOOGLE_TOKEN_URI', 'https://oauth2.googleapis.com/token'),
            "auth_provider_x509_cert_url": os.environ.get('GOOGLE_AUTH_PROVIDER_CERT_URL', 'https://www.googleapis.com/oauth2/v1/certs'),
            "client_x509_cert_url": os.environ.get('GOOGLE_CLIENT_CERT_URL'),
            "universe_domain": os.environ.get('GOOGLE_UNIVERSE_DOMAIN', 'googleapis.com')
        }
        return ServiceAccountCredentials.from_json_keyfile_dict(creds_dict, scope)

    # ลองอ่านจากไฟล์ credentials.json (สำหรับ Local Development)
    creds_path = os.path.join(os.path.dirname(__file__), 'credentials.json')
    if os.path.exists(creds_path):
        return ServiceAccountCredentials.from_json_keyfile_name(creds_path, scope)

    # ไม่พบ credentials ทั้งหมด
    raise RuntimeError(
        "ไม่พบ Google Service Account Credentials\n\n"
        "สำหรับ Production: ตั้งค่า Environment Variables ใน Railway:\n"
        "- GOOGLE_CREDENTIALS_JSON (แนะนำ) หรือ\n"
        "- GOOGLE_PRIVATE_KEY, GOOGLE_CLIENT_EMAIL, ฯลฯ\n\n"
        "สำหรับ Local: วางไฟล์ credentials.json ในโฟลเดอร์โปรเจกต์"
    )


# -----------------------------
# สร้างแอป + บูตระบบเบื้องต้น
# -----------------------------
def create_app():
    app = Flask(__name__)
    app.secret_key = os.environ.get("SECRET_KEY", "vnix-secret")

    db_path = os.path.join(os.path.dirname(__file__), "data.db")
    app.config["SQLALCHEMY_DATABASE_URI"] = f"sqlite:///{db_path}"
    app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

    db.init_app(app)

    # =========[ NEW ]=========
    # Model: ออเดอร์ที่ถูกทำเป็น "ยกเลิก"
    class CancelledOrder(db.Model):
        __tablename__ = "cancelled_orders"
        id = db.Column(db.Integer, primary_key=True)
        order_id = db.Column(db.String(128), unique=True, index=True, nullable=False)
        imported_at = db.Column(db.DateTime, default=datetime.utcnow, index=True, nullable=False)
        imported_by_user_id = db.Column(db.Integer, db.ForeignKey("users.id"))
        note = db.Column(db.String(255))

    # =========[ NEW ]=========  Order "จ่ายงานแล้ว"
    class IssuedOrder(db.Model):
        __tablename__ = "issued_orders"
        id = db.Column(db.Integer, primary_key=True)
        order_id = db.Column(db.String(128), unique=True, index=True, nullable=False)
        issued_at = db.Column(db.DateTime, default=datetime.utcnow, index=True, nullable=False)
        issued_by_user_id = db.Column(db.Integer, db.ForeignKey("users.id"))
        source = db.Column(db.String(32))  # 'import' | 'print:picking' | 'print:warehouse' | 'manual'
        note = db.Column(db.String(255))
    # =========[ /NEW ]=========

    # =========[ NEW ]=========  Order ที่ถูกลบ (Soft Delete / Recycle Bin)
    class DeletedOrder(db.Model):
        __tablename__ = "deleted_orders"
        id = db.Column(db.Integer, primary_key=True)
        order_id = db.Column(db.String(128), unique=True, index=True, nullable=False)
        deleted_at = db.Column(db.DateTime, default=datetime.utcnow, index=True, nullable=False)
        deleted_by_user_id = db.Column(db.Integer, db.ForeignKey("users.id"))
        note = db.Column(db.String(255))
    # =========[ /NEW ]=========

    # =========[ NEW ]=========  ตารางเก็บประวัติการนำเข้า Orders (สำหรับ Dashboard)
    class ImportLog(db.Model):
        __tablename__ = "import_logs"
        id = db.Column(db.Integer, primary_key=True)
        import_date = db.Column(db.Date, index=True, nullable=False)  # วันที่ import
        platform = db.Column(db.String(50))
        shop_name = db.Column(db.String(128))  # ชื่อร้านที่นำเข้า
        filename = db.Column(db.String(255))
        
        # เก็บยอดเฉพาะเหตุการณ์ในไฟล์นั้นๆ
        added_count = db.Column(db.Integer, default=0)
        duplicates_count = db.Column(db.Integer, default=0)
        duplicates_same_day = db.Column(db.Integer, default=0)  # ซ้ำในวันเดียวกัน (ไม่แสดงในการ์ด)
        failed_count = db.Column(db.Integer, default=0)
        
        # เก็บรายชื่อ Error (JSON String)
        error_details = db.Column(db.Text, nullable=True)
        
        # เก็บ Batch Data (IDs ที่เพิ่ม/ซ้ำ/ไม่สำเร็จ) JSON String
        batch_data = db.Column(db.Text, nullable=True)
        
        created_at = db.Column(db.DateTime, default=datetime.utcnow)
    # =========[ /NEW ]=========

    # ---------- Helper: Table name (OrderLine) ----------
    def _ol_table_name() -> str:
        try:
            return OrderLine.__table__.name
        except Exception:
            return getattr(OrderLine, "__tablename__", "order_lines")

    # ---------- Auto-migrate: ensure print columns exist ----------
    def _ensure_orderline_print_columns():
        """Auto-migrate: เพิ่มคอลัมน์สำหรับติดตามสถานะการพิมพ์ Warehouse และ Picking"""
        tbl = _ol_table_name()
        with db.engine.connect() as con:
            cols = {row[1] for row in con.execute(text(f"PRAGMA table_info({tbl})")).fetchall()}

            def add(col, ddl):
                if col not in cols:
                    con.execute(text(f"ALTER TABLE {tbl} ADD COLUMN {col} {ddl}"))

            # สำหรับ "ใบงานคลัง (Warehouse Job Sheet)"
            add("printed_warehouse", "INTEGER DEFAULT 0")  # จำนวนครั้งที่พิมพ์
            add("printed_warehouse_at", "TEXT")  # timestamp ครั้งล่าสุด
            add("printed_warehouse_by", "TEXT")  # username ผู้พิมพ์

            # สำหรับ "Picking List"
            add("printed_picking", "INTEGER DEFAULT 0")  # จำนวนครั้งที่พิมพ์
            add("printed_picking_at", "TEXT")  # timestamp ครั้งล่าสุด
            add("printed_picking_by", "TEXT")  # username ผู้พิมพ์

            # สำหรับ "จ่ายงาน(รอบที่)"
            add("dispatch_round", "INTEGER")

            # สำหรับ "รายงานสินค้าน้อย" (แยกจากคลัง/Picking)
            add("printed_lowstock", "INTEGER DEFAULT 0")
            add("printed_lowstock_at", "TEXT")
            add("printed_lowstock_by", "TEXT")
            add("lowstock_round", "INTEGER")

            # สำหรับ "รายงานไม่มีสินค้า" (แยกจาก lowstock)
            add("printed_nostock", "INTEGER DEFAULT 0")
            add("printed_nostock_at", "TEXT")
            add("printed_nostock_by", "TEXT")
            add("nostock_round", "INTEGER")

            # สำหรับ "รายงานสินค้าไม่พอส่ง" (NOT_ENOUGH)
            add("printed_notenough", "INTEGER DEFAULT 0")
            add("printed_notenough_at", "TEXT")
            add("printed_notenough_by", "TEXT")
            add("notenough_round", "INTEGER")

            # สำหรับ "Barcode Scan Check" (Warehouse)
            add("scanned_at", "TEXT")
            add("scanned_by", "TEXT")

            con.commit()

    # ========== [NEW] Auto-migrate shops unique: (platform, name) ==========
    def _has_unique_index_on(conn, table: str, columns_exact: list[str]) -> tuple[bool, str | None]:
        idx_list = conn.execute(text(f"PRAGMA index_list({table})")).fetchall()
        for row in idx_list:
            idx_name = row[1]
            is_unique = int(row[2]) == 1
            if not is_unique:
                continue
            cols = [r[2] for r in conn.execute(text(f"PRAGMA index_info('{idx_name}')")).fetchall()]
            if cols == columns_exact:
                return True, idx_name
        return False, None

    def _migrate_shops_unique_to_platform_name():
        """ย้าย unique จาก name เดี่ยว → เป็น (platform, name)"""
        with db.engine.begin() as con:
            has_composite, _ = _has_unique_index_on(con, "shops", ["platform", "name"])
            if has_composite:
                return
            has_name_unique, idx_name = _has_unique_index_on(con, "shops", ["name"])
            if has_name_unique:
                is_auto = idx_name.startswith("sqlite_autoindex")
                if is_auto:
                    cols_info = con.execute(text("PRAGMA table_info(shops)")).fetchall()
                    col_names = [c[1] for c in cols_info]
                    has_created_at = "created_at" in col_names
                    con.execute(text("ALTER TABLE shops RENAME TO shops_old"))
                    create_sql = """
                    CREATE TABLE shops (
                        id INTEGER PRIMARY KEY,
                        platform TEXT,
                        name TEXT NOT NULL,
                        created_at TEXT
                    )
                    """ if has_created_at else """
                    CREATE TABLE shops (
                        id INTEGER PRIMARY KEY,
                        platform TEXT,
                        name TEXT NOT NULL
                    )
                    """
                    con.execute(text(create_sql))
                    copy_cols = "id, platform, name" + (", created_at" if has_created_at else "")
                    con.execute(text(f"INSERT INTO shops ({copy_cols}) SELECT {copy_cols} FROM shops_old"))
                    con.execute(text("DROP TABLE shops_old"))
                else:
                    con.execute(text(f"DROP INDEX IF EXISTS {idx_name}"))
            con.execute(text("CREATE UNIQUE INDEX IF NOT EXISTS uq_shops_platform_name ON shops(platform, name)"))
    # ========== [/NEW] ==========

    # =========[ NEW ]=========
    def _ensure_issue_table():
        try:
            IssuedOrder.__table__.create(bind=db.engine, checkfirst=True)
        except Exception as e:
            app.logger.warning(f"[issued_orders] ensure table failed: {e}")
    # =========[ /NEW ]=========

    # =========[ NEW ]=========  ตารางเก็บ Order ที่ถูกลบ
    def _ensure_deleted_table():
        try:
            DeletedOrder.__table__.create(bind=db.engine, checkfirst=True)
        except Exception as e:
            app.logger.warning(f"[deleted_orders] ensure table failed: {e}")
    # =========[ /NEW ]=========

    # =========[ NEW ]=========  ตารางเก็บประวัติการนำเข้า (Import Log)
    def _ensure_import_log_table():
        try:
            ImportLog.__table__.create(bind=db.engine, checkfirst=True)
        except Exception as e:
            app.logger.warning(f"[import_logs] ensure table failed: {e}")
    # =========[ /NEW ]=========

    # =========[ NEW ]=========  เพิ่มคอลัมน์ใหม่ให้ตาราง Shop และ ImportLog
    def _ensure_shop_url_and_log_batch_columns():
        """Auto-migrate: เพิ่มคอลัมน์ google_sheet_url ให้ Shop และ batch_data, shop_name, duplicates_same_day ให้ ImportLog"""
        with db.engine.connect() as con:
            # เพิ่ม google_sheet_url ให้ Shop
            cols_shop = {row[1] for row in con.execute(text("PRAGMA table_info(shops)")).fetchall()}
            if "google_sheet_url" not in cols_shop:
                con.execute(text("ALTER TABLE shops ADD COLUMN google_sheet_url TEXT"))
            
            # เพิ่มคอลัมน์ให้ ImportLog
            cols_log = {row[1] for row in con.execute(text("PRAGMA table_info(import_logs)")).fetchall()}
            if "batch_data" not in cols_log:
                con.execute(text("ALTER TABLE import_logs ADD COLUMN batch_data TEXT"))
            if "shop_name" not in cols_log:
                con.execute(text("ALTER TABLE import_logs ADD COLUMN shop_name TEXT"))
            if "duplicates_same_day" not in cols_log:
                con.execute(text("ALTER TABLE import_logs ADD COLUMN duplicates_same_day INTEGER DEFAULT 0"))
            con.commit()
    # =========[ /NEW ]=========

    with app.app_context():
        db.create_all()
        _ensure_orderline_print_columns()
        _migrate_shops_unique_to_platform_name()
        _ensure_issue_table()  # <<< NEW
        _ensure_deleted_table()  # <<< NEW สำหรับ Soft Delete
        _ensure_import_log_table()  # <<< NEW สำหรับ Import Dashboard
        _ensure_shop_url_and_log_batch_columns()  # <<< NEW สำหรับบันทึก URL และ Batch Data
        # bootstrap admin
        if User.query.count() == 0:
            admin = User(
                username="admin",
                password_hash=generate_password_hash("admin123"),
                role="admin",
                active=True
            )
            db.session.add(admin)
            db.session.commit()

    # -----------------
    # Jinja filters
    # -----------------
    @app.template_filter("thai_be")
    def thai_be_filter(dt):
        try:
            return to_thai_be(dt)
        except Exception:
            return ""

    @app.template_filter("be_date")
    def be_date_filter(d):
        try:
            return to_be_date_str(d)
        except Exception:
            return ""

    # -----------------
    # UI context
    # -----------------
    @app.context_processor
    def inject_globals():
        return {
            "APP_NAME": APP_NAME,
            "BE_YEAR": current_be_year(),
            "CURRENT_USER": current_user()
        }

    # ให้ template ตรวจ endpoint ได้ (กันพังค่า has_endpoint)
    @app.template_global()
    def has_endpoint(endpoint: str) -> bool:
        try:
            return endpoint in app.view_functions
        except Exception:
            return False

    # -----------------
    # Auth helpers
    # -----------------
    def current_user():
        uid = session.get("uid")
        if not uid:
            return None
        return db.session.get(User, uid)

    def login_required(fn):
        @wraps(fn)
        def wrapper(*args, **kwargs):
            if not current_user():
                return redirect(url_for("login", next=request.path))
            return fn(*args, **kwargs)
        return wrapper

    # -----------------
    # Utilities (app)
    # -----------------
    def parse_date_any(s: str | None):
        if not s:
            return None
        s = s.strip()
        try:
            if "-" in s:
                y, m, d = s.split("-")
                return date(int(y), int(m), int(d))
            else:
                d, m, y = s.split("/")
                y = int(y)
                if y > 2400:
                    y -= 543
                return date(y, int(m), int(d))
        except Exception:
            return None

    def _get_line_sku(line) -> str:
        if hasattr(line, "sku") and line.sku:
            return str(line.sku).strip()
        try:
            prod = getattr(line, "product", None)
            if prod and getattr(prod, "sku", None):
                return str(prod.sku).strip()
        except Exception:
            pass
        return ""

    def _calc_stock_qty_for_line(line: OrderLine) -> int:
        sku = _get_line_sku(line)
        if not sku:
            return 0
        prod = Product.query.filter_by(sku=sku).first()
        if prod and hasattr(prod, "stock_qty"):
            try:
                return int(prod.stock_qty or 0)
            except Exception:
                pass
        st = Stock.query.filter_by(sku=sku).first()
        try:
            return int(st.qty) if st and st.qty is not None else 0
        except Exception:
            return 0

    def _build_allqty_map(rows: list[dict]) -> dict[str, int]:
        total_by_sku: dict[str, int] = {}
        for r in rows:
            sku = (r.get("sku") or "").strip()
            if not sku:
                continue
            total_by_sku[sku] = total_by_sku.get(sku, 0) + int(r.get("qty", 0) or 0)
        return total_by_sku

    # [DEPRECATED] ฟังก์ชันนี้ไม่ใช้แล้ว - ใช้ compute_allocation() จาก allocation.py แทน
    # เก็บไว้สำหรับ reference เท่านั้น
    def _recompute_allocation_row(r: dict) -> dict:
        stock_qty = int(r.get("stock_qty", 0) or 0)
        allqty = int(r.get("allqty", r.get("qty", 0)) or 0)
        sales_status = (r.get("sales_status") or "").upper()
        packed_flag = bool(r.get("packed", False))
        accepted = bool(r.get("accepted", False))
        order_time = r.get("order_time")
        platform = r.get("platform") or (r.get("shop_platform") if r.get("shop_platform") else "")

        if sales_status == "PACKED" or packed_flag:
            allocation_status = "PACKED"
        elif accepted:
            allocation_status = "ACCEPTED"
        elif stock_qty <= 0:
            allocation_status = "SHORTAGE"
        elif allqty > stock_qty:
            allocation_status = "NOT_ENOUGH"
        elif stock_qty <= 3:
            allocation_status = "LOW_STOCK"
        else:
            allocation_status = "READY_ACCEPT"

        if allocation_status == "PACKED":
            sla = ""
        else:
            try:
                sla = sla_text(platform, order_time) if order_time else ""
            except Exception:
                sla = ""
        try:
            due_date = compute_due_date(platform, order_time) if order_time else None
        except Exception:
            due_date = None

        r["allocation_status"] = allocation_status
        r["sla"] = sla
        r["due_date"] = due_date
        return r

    def _check_mixed_status(order_id: str, all_rows: list[dict]) -> set:
        """
        ตรวจสอบว่า Order นี้มีสินค้าที่มีสถานะต่างกันปนอยู่หรือไม่
        คืนค่าเป็น set ของสถานะทั้งหมดที่พบใน Order นี้
        """
        statuses = set()
        for r in all_rows:
            if (r.get("order_id") or "").strip() == order_id:
                status = r.get("allocation_status")
                if status:
                    statuses.add(status)
        return statuses

    def _annotate_order_spans(rows: list[dict]) -> list[dict]:
        seen = set()
        for r in rows:
            oid = (r.get("order_id") or "").strip()
            if not oid:
                r["show_order_id"] = True
                r["order_id_display"] = ""
                continue
            if oid in seen:
                r["show_order_id"] = False
                r["order_id_display"] = ""
            else:
                r["show_order_id"] = True
                r["order_id_display"] = oid
                seen.add(oid)
        return rows

    def _group_rows_for_report(rows: list[dict]) -> list[dict]:
        def _key(r):
            return (
                (r.get("order_id") or ""),
                (r.get("platform") or ""),
                (r.get("shop") or ""),
                (r.get("logistic") or ""),
                (r.get("sku") or "")
            )
        rows = sorted(rows, key=_key)
        rows = _annotate_order_spans(rows)

        counts: dict[str, int] = {}
        for r in rows:
            oid = (r.get("order_id") or "").strip()
            counts[oid] = counts.get(oid, 0) + 1

        for r in rows:
            oid = (r.get("order_id") or "").strip()
            r["order_rowspan"] = counts.get(oid, 1) if r.get("show_order_id") else 0
            r["order_id_display"] = oid if r.get("show_order_id") else ""
        return rows

    def _group_rows_for_warehouse_report(rows: list[dict]) -> list[dict]:
        """Group rows by order_id to show only 1 row per order for warehouse report"""
        order_map = {}
        
        for r in rows:
            oid = (r.get("order_id") or "").strip()
            if not oid:
                continue
            
            if oid not in order_map:
                # First row for this order - keep it
                # ใช้ printed_warehouse_count หรือ printed_count ที่มาจาก DB (ไม่ใช่ printed_warehouse ที่เป็น 0 ตลอด)
                order_map[oid] = {
                    "order_id": oid,
                    "platform": r.get("platform", ""),
                    "shop": r.get("shop", ""),
                    "logistic": r.get("logistic", ""),
                    "accepted_by": r.get("accepted_by", ""),
                    "printed_count": r.get("printed_warehouse_count") or r.get("printed_count") or r.get("printed_warehouse") or 0,
                    "printed_warehouse": r.get("printed_warehouse_count") or r.get("printed_count") or r.get("printed_warehouse") or 0,
                    "printed_warehouse_at": r.get("printed_warehouse_at"),
                    "printed_warehouse_by": r.get("printed_warehouse_by"),
                    "dispatch_round": r.get("dispatch_round"),
                    "scanned_at": r.get("scanned_at"),
                }
        
        # Convert back to list and sort
        result = list(order_map.values())
        result.sort(key=lambda r: (r["platform"], r["shop"], r["order_id"]))
        return result

    # -----------------
    # สร้างเซ็ต Order พร้อมรับทั้งออเดอร์ / สินค้าน้อยทั้งออเดอร์
    # -----------------
    def _orders_ready_set(rows: list[dict]) -> set[str]:
        by_oid: dict[str, list[dict]] = {}
        for r in rows:
            oid = (r.get("order_id") or "").strip()
            if not oid:
                continue
            by_oid.setdefault(oid, []).append(r)

        ready = set()
        for oid, items in by_oid.items():
            if not items:
                continue
            all_ready = True
            for it in items:
                status = (it.get("allocation_status") or "").upper()
                accepted = bool(it.get("accepted", False))
                packed = (status == "PACKED") or bool(it.get("packed", False))
                is_issued = bool(it.get("is_issued", False))  # [NEW] เช็ค Order จ่ายงานแล้ว
                # [แก้ไข] ถ้าจ่ายงานแล้ว (is_issued) ถือว่าจบงาน ไม่ต้องนับเข้ากอง 1
                if not (status == "READY_ACCEPT" and not accepted and not packed and not is_issued):
                    all_ready = False
                    break
            if all_ready:
                ready.add(oid)
        return ready

    def _orders_lowstock_order_set(rows: list[dict]) -> set[str]:
        by_oid: dict[str, list[dict]] = {}
        for r in rows:
            oid = (r.get("order_id") or "").strip()
            if not oid:
                continue
            by_oid.setdefault(oid, []).append(r)

        result = set()
        for oid, items in by_oid.items():
            if not items:
                continue
            all_sendable = True
            has_low = False
            for it in items:
                status = (it.get("allocation_status") or "").upper()
                accepted = bool(it.get("accepted", False))
                packed = (status == "PACKED") or bool(it.get("packed", False))
                is_issued = bool(it.get("is_issued", False))  # [NEW] เช็ค Order จ่ายงานแล้ว
                # [แก้ไข] ถ้าจ่ายงานแล้ว (is_issued) ถือว่าจบงาน ไม่ต้องนับเข้ากอง 2
                if packed or accepted or is_issued:
                    all_sendable = False
                    break
                if status not in ("READY_ACCEPT", "LOW_STOCK"):
                    all_sendable = False
                    break
                if status == "LOW_STOCK":
                    has_low = True
            if all_sendable and has_low:
                result.add(oid)
        return result

    # ===================== NEW: Orders ที่ยังไม่มีการเปิดใบขาย =====================
    def _has_any_sales(r: dict) -> bool:
        """
        คืน True ถ้า row นี้ 'มีการเปิดใบขายแล้วบางส่วนหรือทั้งหมด'
        (ดูจากข้อความในฟิลด์สั่งขายต่างๆ)
        """
        text_pool = [
            str(r.get("sale_status") or ""),
            str(r.get("sale_text") or ""),
            str(r.get("sales_status") or ""),
            str(r.get("sales_note") or ""),
        ]
        joined = " ".join(text_pool).strip()
        if not joined:
            # ไม่มีข้อความอะไรเกี่ยวกับใบขายเลย ⇒ ถือว่ายังไม่มีการเปิดใบขาย
            return False

        low = joined.lower()
        # ถ้ามีคำว่า "ยังไม่มีการเปิดใบขาย" ชัดเจน ⇒ ยังไม่มีใบขาย
        if "ยังไม่มีการเปิดใบขาย" in low:
            return False

        # อย่างอื่น (มีคำว่า เปิดใบขาย, packed, ฯลฯ) ถือว่ามีใบขายแล้ว
        return True

    def _orders_no_sales_set(rows: list[dict]) -> set[str]:
        """
        คืนเซ็ต order_id ที่ 'ทุกบรรทัดของออเดอร์นั้นยังไม่มีการเปิดใบขาย'
        (ใช้ข้อความในคอลัมน์สั่งขายตัดสิน)
        
        หมายเหตุ: ไม่รวม Order ที่ยังไม่นำเข้า SBS (is_not_in_sbs = True)
        """
        by_oid: dict[str, list[dict]] = {}
        for r in rows:
            oid = (r.get("order_id") or "").strip()
            if not oid:
                continue
            by_oid.setdefault(oid, []).append(r)

        result: set[str] = set()
        for oid, items in by_oid.items():
            if not items:
                continue
            # [แก้ไข] ข้ามถ้าเป็น Order ยังไม่นำเข้า SBS
            if any(it.get("is_not_in_sbs") for it in items):
                continue
            # ถ้าทุกบรรทัดยังไม่มีใบขาย ⇒ นับเป็น Order ยังไม่มีการเปิดใบขาย
            if all(not _has_any_sales(it) for it in items):
                result.add(oid)
        return result
    # ===================== /NEW =====================

    # ===================== NEW: Orders ยังไม่นำเข้า SBS =====================
    def _orders_not_in_sbs_set(rows: list[dict]) -> set[str]:
        """
        คืนเซ็ต order_id ที่ 'ยังไม่มีข้อมูลใน SBS เลย' (Sales is None)
        ใช้ flag is_not_in_sbs ที่คำนวณมาจาก allocation.py
        """
        result: set[str] = set()
        for r in rows:
            if r.get("is_not_in_sbs"):
                oid = (r.get("order_id") or "").strip()
                if oid:
                    result.add(oid)
        return result
    # ===================== /NEW =====================

    # ===========================================================
    # Packed helpers — จาก "เปิดใบขายครบตามจำนวนแล้ว"
    # ===========================================================
    def _is_line_opened_full(r: dict) -> bool:
        text_pool = [
            str(r.get("sale_status") or ""),
            str(r.get("sale_text") or ""),
            str(r.get("sales_status") or ""),
            str(r.get("sales_note") or ""),
        ]
        norm = " ".join(s.strip() for s in text_pool if s).lower()
        flag = bool(r.get("sale_open_full") or r.get("opened_full") or r.get("is_opened_full"))
        return flag or ("เปิดใบขายครบตามจำนวนแล้ว" in norm) or ("opened_full" in norm)

    def _orders_packed_set(rows: list[dict]) -> set[str]:
        by_oid: dict[str, list[dict]] = {}
        for r in rows:
            oid = (r.get("order_id") or "").strip()
            if not oid:
                continue
            by_oid.setdefault(oid, []).append(r)
        packed: set[str] = set()
        for oid, items in by_oid.items():
            if items and all(_is_line_opened_full(it) for it in items):
                packed.add(oid)
        return packed

    # ---------------------------------------------------------
    # ฟังก์ชัน DB raw (ใช้เพราะคอลัมน์ถูกเพิ่มแบบ ALTER TABLE)
    # ---------------------------------------------------------
    def _detect_already_printed(oids: list[str], kind: str) -> set[str]:
        if not oids:
            return set()
        tbl = _ol_table_name()
        col = "printed_warehouse" if kind == "warehouse" else "printed_picking"
        # เปลี่ยนจาก =1 เป็น >=1 เพื่อกันกรณีพิมพ์หลายครั้ง
        sql = text(f"SELECT DISTINCT order_id FROM {tbl} WHERE order_id IN :oids AND {col} >= 1")
        sql = sql.bindparams(bindparam("oids", expanding=True))
        rows = db.session.execute(sql, {"oids": oids}).scalars().all()
        return set(r for r in rows if r)

    def _mark_printed(oids: list[str], kind: str, user_id: int | None, when_iso: str):
        """อัปเดตสถานะการพิมพ์ + timestamp + username"""
        if not oids:
            return
        
        # ดึง username จาก user_id
        username = None
        if user_id:
            user_obj = User.query.get(user_id)
            if user_obj:
                username = user_obj.username
        
        tbl = _ol_table_name()
        if kind == "warehouse":
            col_count = "printed_warehouse"
            col_at   = "printed_warehouse_at"
            col_by   = "printed_warehouse_by"
        else:
            col_count = "printed_picking"
            col_at   = "printed_picking_at"
            col_by   = "printed_picking_by"

        sql = text(
            f"""
            UPDATE {tbl}
               SET {col_count}=COALESCE({col_count},0)+1,
                   {col_at}=:ts,
                   {col_by}=:username
             WHERE order_id IN :oids
            """
        ).bindparams(bindparam("oids", expanding=True))
        db.session.execute(sql, {"username": username, "ts": when_iso, "oids": oids})
        db.session.commit()

    # --------------------------
    # Print count helpers (ใหม่)
    # --------------------------
    def _get_print_counts_local(oids: list[str], kind: str) -> dict[str, int]:
        """คืน dict: {order_id: count} อ่านจำนวนครั้งที่พิมพ์จากคอลัมน์ printed_warehouse หรือ printed_picking หรือ printed_lowstock"""
        if not oids:
            return {}
        tbl = _ol_table_name()
        if kind == "lowstock":
            col = "printed_lowstock"
        elif kind == "nostock":  # <<< เพิ่มสำหรับรายงานไม่มีสินค้า
            col = "printed_nostock"
        elif kind == "notenough":  # <<< เพิ่มสำหรับรายงานสินค้าไม่พอส่ง
            col = "printed_notenough"
        elif kind == "warehouse":
            col = "printed_warehouse"
        else:
            col = "printed_picking"
        sql = text(f"SELECT order_id, COALESCE(MAX({col}),0) AS c FROM {tbl} WHERE order_id IN :oids GROUP BY order_id")
        sql = sql.bindparams(bindparam("oids", expanding=True))
        rows_sql = db.session.execute(sql, {"oids": oids}).all()
        return {str(r[0]): int(r[1] or 0) for r in rows_sql if r and r[0]}

    def _mark_lowstock_printed(oids: list[str], username: str | None, when_iso: str):
        """อัปเดตการพิมพ์สำหรับรายงานสินค้าน้อย"""
        if not oids:
            return
        tbl = _ol_table_name()
        sql = text(f"""
            UPDATE {tbl}
               SET printed_lowstock=COALESCE(printed_lowstock,0)+1,
                   printed_lowstock_at=:ts,
                   printed_lowstock_by=:byu
             WHERE order_id IN :oids
        """).bindparams(bindparam("oids", expanding=True))
        db.session.execute(sql, {"ts": when_iso, "byu": username, "oids": oids})
        db.session.commit()

    def _mark_nostock_printed(oids: list[str], username: str | None, when_iso: str):
        """อัปเดตการพิมพ์สำหรับรายงานไม่มีสินค้า"""
        if not oids:
            return
        tbl = _ol_table_name()
        sql = text(f"""
            UPDATE {tbl}
               SET printed_nostock=COALESCE(printed_nostock,0)+1,
                   printed_nostock_at=:ts,
                   printed_nostock_by=:byu
             WHERE order_id IN :oids
        """).bindparams(bindparam("oids", expanding=True))
        db.session.execute(sql, {"ts": when_iso, "byu": username, "oids": oids})
        db.session.commit()

    def _mark_notenough_printed(oids: list[str], username: str | None, when_iso: str):
        """อัปเดตการพิมพ์สำหรับรายงานสินค้าไม่พอส่ง"""
        if not oids:
            return
        tbl = _ol_table_name()
        sql = text(f"""
            UPDATE {tbl}
               SET printed_notenough=COALESCE(printed_notenough,0)+1,
                   printed_notenough_at=:ts,
                   printed_notenough_by=:byu
             WHERE order_id IN :oids
        """).bindparams(bindparam("oids", expanding=True))
        db.session.execute(sql, {"ts": when_iso, "byu": username, "oids": oids})
        db.session.commit()

    def _inject_scan_status(rows: list[dict]):
        """ดึงข้อมูลว่าออเดอร์ไหนสแกนแล้วบ้าง"""
        oids = sorted({(r.get("order_id") or "").strip() for r in rows if r.get("order_id")})
        if not oids:
            return
        
        tbl = _ol_table_name()
        sql = text(f"SELECT order_id, MAX(scanned_at) FROM {tbl} WHERE order_id IN :oids GROUP BY order_id")
        sql = sql.bindparams(bindparam("oids", expanding=True))
        res = db.session.execute(sql, {"oids": oids}).fetchall()
        scan_map = {r[0]: r[1] for r in res if r[0]}
        
        for r in rows:
            oid = (r.get("order_id") or "").strip()
            r["scanned_at"] = scan_map.get(oid)

    def _inject_print_counts_to_rows(rows: list[dict], kind: str):
        """ฝัง printed_*_count และ printed_*_at ลงในแต่ละแถว (ใช้กับ Warehouse report)"""
        oids = sorted({(r.get("order_id") or "").strip() for r in rows if r.get("order_id")})
        counts = _get_print_counts_local(oids, kind)
        
        # Also get the timestamp of last print
        if not oids:
            return
        
        tbl = _ol_table_name()
        col_at = "printed_warehouse_at" if kind == "warehouse" else "printed_picking_at"
        sql = text(f"SELECT order_id, MAX({col_at}) AS last_printed_at FROM {tbl} WHERE order_id IN :oids GROUP BY order_id")
        sql = sql.bindparams(bindparam("oids", expanding=True))
        rows_sql = db.session.execute(sql, {"oids": oids}).all()
        timestamps = {}
        
        # Convert ISO string to datetime object
        for r_sql in rows_sql:
            if r_sql and r_sql[0] and r_sql[1]:
                try:
                    # Parse ISO datetime string
                    dt = datetime.fromisoformat(r_sql[1])
                    if dt.tzinfo is None:
                        dt = TH_TZ.localize(dt)
                    timestamps[str(r_sql[0])] = dt
                except Exception:
                    pass
        
        for r in rows:
            oid = (r.get("order_id") or "").strip()
            c = int(counts.get(oid, 0))
            r["printed_count"] = c
            if kind == "warehouse":
                r["printed_warehouse_count"] = c
                r["printed_warehouse"] = c  # <-- เพิ่มบรรทัดนี้เพื่อให้เทมเพลตอ่ยใช้ได้
                r["printed_warehouse_at"] = timestamps.get(oid)
            else:
                r["printed_picking_count"] = c
                r["printed_picking"] = c  # <-- และบรรทัดน้
                r["printed_picking_at"] = timestamps.get(oid)

    # =========[ NEW ]=========
    # ส่วนเสริมเพื่อ "Order ยกเลิก"
    try:
        from openpyxl import load_workbook, Workbook
        _OPENPYXL_OK = True
    except Exception:
        _OPENPYXL_OK = False

    def _ensure_cancel_table():
        try:
            CancelledOrder.__table__.create(bind=db.engine, checkfirst=True)
        except Exception as e:
            app.logger.warning(f"[cancelled_orders] ensure table failed: {e}")

    def _cancelled_oids_set() -> set[str]:
        """คืนค่า set ของ order_id ที่ถูกยกเลิก (สำหรับ backward compatibility)"""
        rows = db.session.query(CancelledOrder.order_id).all()
        return {r[0] for r in rows if r and r[0]}

    def _cancelled_oids_map() -> dict[str, dict]:
        """คืนค่า dict ของ {order_id: {'note': note, 'at': timestamp}}"""
        rows = db.session.query(
            CancelledOrder.order_id, 
            CancelledOrder.note, 
            CancelledOrder.imported_at
        ).all()
        # เก็บทั้ง Note และ เวลา
        return {r[0]: {'note': (r[1] or ""), 'at': r[2]} for r in rows if r and r[0]}

    def _filter_out_cancelled_rows(rows: list[dict]) -> list[dict]:
        canc = _cancelled_oids_set()
        if not canc:
            return rows
        res = []
        for r in rows:
            oid = (r.get("order_id") or "").strip()
            if oid and oid in canc:
                continue
            res.append(r)
        return res

    # ===== HELPER: Issued (จ่ายงานแล้ว) =====
    def _issued_oids_set() -> set[str]:
        rows = db.session.query(IssuedOrder.order_id).all()
        return {r[0] for r in rows if r and r[0]}

    def _filter_out_issued_rows(rows: list[dict]) -> list[dict]:
        issued = _issued_oids_set()
        if not issued:
            return rows
        res = []
        for r in rows:
            oid = (r.get("order_id") or "").strip()
            if oid and oid in issued:
                continue
            res.append(r)
        return res

    # ===== HELPER: Deleted Orders (ถูกลบ / ถังขยะ) =====
    def _deleted_oids_set() -> set[str]:
        """ดึง order_id ทั้งหมดที่ถูกลบ (Soft Delete)"""
        rows = db.session.query(DeletedOrder.order_id).all()
        return {r[0] for r in rows if r and r[0]}

    def _filter_out_deleted_rows(rows: list[dict]) -> list[dict]:
        """กรอง order ที่ถูกลบออกจากรายการ"""
        deleted = _deleted_oids_set()
        if not deleted:
            return rows
        res = []
        for r in rows:
            oid = (r.get("order_id") or "").strip()
            if oid and oid in deleted:
                continue
            res.append(r)
        return res

    # ===== HELPER: Low Stock Printed (พิมพ์รายงานสินค้าน้อยแล้ว) =====
    def _lowstock_printed_oids_set() -> set[str]:
        """ดึง order_id ที่เคยพิมพ์รายงานสินค้าน้อยแล้ว"""
        tbl = _ol_table_name()
        rows = db.session.execute(text(f"""
            SELECT DISTINCT order_id
            FROM {tbl}
            WHERE printed_lowstock > 0
        """)).fetchall()
        return {r[0] for r in rows if r and r[0]}

    def _filter_out_lowstock_printed_rows(rows: list[dict]) -> list[dict]:
        """กรองออเดอร์ที่พิมพ์รายงานสินค้าน้อยออกแล้ว (ข้อ 2)"""
        printed = _lowstock_printed_oids_set()
        if not printed:
            return rows
        res = []
        for r in rows:
            oid = (r.get("order_id") or "").strip()
            if oid and oid in printed:
                continue
            res.append(r)
        return res

    def _mark_issued(oids: list[str], user_id: int | None, source: str = "manual", when_dt=None):
        """ทำเครื่องหมาย 'จ่ายงานแล้ว' โดยไม่แก้ทับข้อมูลเก่า (ยึดเวลาเดิม)"""
        if not oids:
            return 0
        # ใช้เวลาที่ส่งมา (เช่น ตอน import) ถ้าไม่ส่งมาก็ใช้เวลาปัจจุบันโซนไทย
        when_dt = when_dt or now_thai()
        try:
            # เก็บแบบ naive เพื่อให้ SQLite รับได้
            if getattr(when_dt, "tzinfo", None) is not None:
                when_dt = when_dt.replace(tzinfo=None)
        except Exception:
            pass

        existing = {
            r[0] for r in db.session.query(IssuedOrder.order_id)
            .filter(IssuedOrder.order_id.in_(oids)).distinct().all()
        }
        inserted = 0
        for oid in oids:
            oid = (oid or "").strip()
            if not oid or oid in existing:
                # มีข้อมูลเก่าแล้ว (เช่นมาจากการพิมพ์) ก็ไม่แก้ทับ ⇒ ยึดเวลาเก่าไว้
                continue
            db.session.add(IssuedOrder(order_id=oid, issued_at=when_dt, issued_by_user_id=user_id, source=source))
            inserted += 1
        db.session.commit()
        return inserted

    def _unissue(oids: list[str]) -> int:
        if not oids:
            return 0
        n = db.session.query(IssuedOrder).filter(IssuedOrder.order_id.in_(oids)).delete(synchronize_session=False)
        db.session.commit()
        return n

    # ให้ import "จ่ายงานแล้ว" ตั้งค่า counter ขั้นต่ำเป็น 1
    def _ensure_min_print_count(oids: list[str], min_count: int = 1, user_id: int | None = None, when_iso: str | None = None):
        """บังคับให้ printed_picking_count >= min_count (เฉพาะ Picking เท่านั้น)"""
        if not oids:
            return
        tbl = _ol_table_name()
        when_iso = when_iso or now_thai().isoformat()

        # เซ็ตเฉพาะ Picking (ไม่แตะ Warehouse)
        sql = text(f"""
            UPDATE {tbl}
               SET printed_picking=1,
                   printed_picking_count = CASE WHEN COALESCE(printed_picking_count,0) < :mc THEN :mc ELSE printed_picking_count END,
                   printed_picking_by_user_id = COALESCE(printed_picking_by_user_id, :uid),
                   printed_picking_at = COALESCE(printed_picking_at, :ts)
             WHERE order_id IN :oids
        """).bindparams(bindparam("oids", expanding=True))
        db.session.execute(sql, {"mc": min_count, "uid": user_id, "ts": when_iso, "oids": oids})

        db.session.commit()

    def _ensure_shops_from_df(df, platform: str, default_shop_name: str = None):
        """สร้างหรือใช้ Shop ที่มีอยู่แล้ว ก่อนที่จะ import orders (กัน UNIQUE constraint พัง)"""
        from utils import normalize_platform
        platform = normalize_platform(platform)
        
        # รวบรวม shop names ที่มีใน df (ลองดูหลายคอลัมน์ที่อาจมีชื่อร้าน)
        shop_names = set()
        for col in df.columns:
            col_lower = str(col).lower()
            if "shop" in col_lower or "ร้าน" in col_lower:
                for val in df[col].dropna().unique():
                    name = str(val).strip()
                    if name:
                        shop_names.add(name)
        
        # ถ้าไม่เจอใน df ให้ใช้ default_shop_name
        if not shop_names and default_shop_name:
            shop_names.add(default_shop_name.strip())
        
        # สร้าง/ใช้ shop ที่มีอยู่แล้ว
        for name in shop_names:
            existing = Shop.query.filter_by(platform=platform, name=name).first()
            if not existing:
                new_shop = Shop(platform=platform, name=name)
                db.session.add(new_shop)
        db.session.commit()

    def _parse_order_ids_from_upload(file_storage) -> list[str]:
        filename = (file_storage.filename or "").lower()
        data = file_storage.read()
        file_storage.stream.seek(0)

        order_ids: list[str] = []

        # Excel
        if filename.endswith(".xlsx") or filename.endswith(".xls"):
            if not _OPENPYXL_OK:
                raise RuntimeError("ไม่พบไลบรารี openpyxl สำหรับอ่านไฟล์ Excel, ติดตั้งด้วย: pip install openpyxl")
            wb = load_workbook(filename=BytesIO(data), read_only=True, data_only=True)
            ws = wb.active
            for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if not row:
                    continue
                val = row[0]
                if i == 1 and isinstance(val, str) and val.strip().lower() in {"order_id", "order_no", "เลขออเดอร์"}:
                    continue
                if val is None:
                    continue
                s = str(val).strip()
                if s:
                    order_ids.append(s)
            return order_ids

        # CSV
        if filename.endswith(".csv"):
            text_data = data.decode("utf-8-sig", errors="ignore")
            reader = csv.reader(text_data.splitlines())
            for i, row in enumerate(reader, start=1):
                if not row:
                    continue
                val = row[0]
                if i == 1 and isinstance(val, str) and val.strip().lower() in {"order_id", "order_no", "เลขออเดอร์"}:
                    continue
                s = str(val).strip()
                if s:
                    order_ids.append(s)
            return order_ids

        raise RuntimeError("รองรับเฉพาะไฟล์ .xlsx .xls หรือ .csv เท่านั้น")
    # =========[ /NEW ]=========

    # -------------
    # Routes: Auth & Users
    # -------------

    # --------- Admin: Shops (เดิม) ---------
    @app.route("/admin/shops")
    @login_required
    def admin_shops():
        cu = current_user()
        if not cu or cu.role not in {"admin", "staff"}:
            flash("ต้องเป็นผู้ดูแลระบบหรือพนักงานเท่านั้น", "danger")
            return redirect(url_for("dashboard"))
        shops = Shop.query.order_by(Shop.platform.asc(), Shop.name.asc()).all()
        counts = {s.id: db.session.query(func.count(OrderLine.id)).filter_by(shop_id=s.id).scalar() for s in shops}
        return render_template("admin_shops.html", shops=shops, counts=counts)

    @app.route("/admin/shops/<int:shop_id>/delete", methods=["POST"])
    @login_required
    def delete_shop(shop_id):
        cu = current_user()
        if not cu or cu.role != "admin":
            flash("เฉพาะแอดมินเท่านั้นที่ลบได้", "danger")
            return redirect(url_for("admin_shops"))
        s = Shop.query.get(shop_id)
        if not s:
            flash("ไม่พบร้านนี้", "warning")
            return redirect(url_for("admin_shops"))
        cnt = db.session.query(func.count(OrderLine.id)).filter_by(shop_id=s.id).scalar()
        if cnt and cnt > 0:
            flash("ไม่สามารถลบได้: มีออเดอร์ผูกกับร้านนี้อยู่", "danger")
            return redirect(url_for("admin_shops"))
        db.session.delete(s)
        db.session.commit()
        flash(f"ลบร้าน '{s.name}' แล้ว", "success")
        return redirect(url_for("admin_shops"))
    # --------------------------------------

    # -----------------------
    # API: จัดการ Link Google Sheet ของร้าน
    # -----------------------
    @app.route("/api/shop/url", methods=["POST"])
    @login_required
    def api_shop_url():
        data = request.get_json() or {}
        shop_name = (data.get("shop_name") or "").strip()
        platform = normalize_platform(data.get("platform") or "")
        url = (data.get("url") or "").strip()
        action = data.get("action")  # 'save' or 'delete'
        
        # ถ้าไม่ได้ระบุชื่อร้าน ให้ใช้ชื่อ Platform เป็นชื่อร้านแทน (เพื่อเก็บ URL กลาง)
        if not shop_name and platform:
            shop_name = platform
        
        if not shop_name:
            return jsonify({"success": False, "msg": "กรุณาเลือกแพลตฟอร์ม"})
        
        # ค้นหาร้านตาม platform + name
        shop = Shop.query.filter_by(platform=platform, name=shop_name).first()
        if not shop:
            # ถ้าไม่มี platform ตรง ลองหาแค่ชื่อร้าน
            shop = Shop.query.filter_by(name=shop_name).first()
        
        if not shop:
            # สร้างร้านใหม่ถ้ายังไม่มี
            shop = Shop(platform=platform or "อื่นๆ", name=shop_name)
            db.session.add(shop)
            db.session.commit()  # Commit เพื่อให้ได้ ID มาใช้
        
        if action == "save":
            # [แก้ไข] ใช้ SQL Update ตรงๆ
            db.session.execute(
                text("UPDATE shops SET google_sheet_url = :u WHERE id = :id"),
                {"u": url, "id": shop.id}
            )
            db.session.commit()
            return jsonify({"success": True, "msg": "บันทึกลิงก์เรียบร้อย"})
        elif action == "delete":
            # [แก้ไข] ใช้ SQL Update ตรงๆ (Set เป็น NULL)
            db.session.execute(
                text("UPDATE shops SET google_sheet_url = NULL WHERE id = :id"),
                {"id": shop.id}
            )
            db.session.commit()
            return jsonify({"success": True, "msg": "ลบลิงก์เรียบร้อย"})
            
        return jsonify({"success": False, "msg": "ไม่ระบุ action"})
    # --------------------------------------

    @app.route("/login", methods=["GET", "POST"])
    def login():
        if request.method == "POST":
            username = request.form.get("username", "").strip()
            password = request.form.get("password", "")
            u = User.query.filter_by(username=username, active=True).first()
            if u and check_password_hash(u.password_hash, password):
                session["uid"] = u.id
                flash("เข้าสู่ระบบสำเร็จ", "success")
                return redirect(request.args.get("next") or url_for("dashboard"))
            flash("ชื่อผู้ใช้หรือรหัสผ่านไม่ถูกต้อง", "danger")
        return render_template("login.html")

    @app.route("/logout")
    def logout():
        session.clear()
        flash("ออกจากระบบแล้ว", "info")
        return redirect(url_for("login"))

    @app.route("/admin/users", methods=["GET", "POST"])
    @login_required
    def admin_users():
        cu = current_user()
        if cu.role != "admin":
            flash("ต้องเป็นผู้ดูแลระบบเท่านั้น", "danger")
            return redirect(url_for("dashboard"))
        if request.method == "POST":
            action = request.form.get("action")
            if action == "create":
                username = request.form.get("username").strip()
                password = request.form.get("password")
                role = request.form.get("role", "user")
                if not username or not password:
                    flash("กรุณากรอกชื่อผู้ใช้/รหัสผ่าน", "danger")
                elif User.query.filter_by(username=username).first():
                    flash("มีชื่อผู้ใช้นี้อยู่แล้ว", "warning")
                else:
                    u = User(
                        username=username,
                        password_hash=generate_password_hash(password),
                        role=role,
                        active=True
                    )
                    db.session.add(u)
                    db.session.commit()
                    flash(f"สร้างผู้ใช้ {username} แล้ว", "success")
            elif action == "delete":
                uid = int(request.form.get("uid"))
                if uid == cu.id:
                    flash("ลบตัวเองไม่ได้", "warning")
                else:
                    User.query.filter_by(id=uid).delete()
                    db.session.commit()
                    flash("ลบผู้ใช้แล้ว", "success")
        users = User.query.order_by(User.created_at.desc()).all() if hasattr(User, "created_at") else User.query.all()
        return render_template("users.html", users=users)

    # -------------
    # Dashboard
    # -------------
    @app.route("/")
    @login_required
    def dashboard():
        platform = normalize_platform(request.args.get("platform"))
        shop_id = request.args.get("shop_id")
        
        # [แก้ไข] เปลี่ยนจาก import_date เดี่ยว เป็น Range
        import_from_str = request.args.get("import_from")
        import_to_str = request.args.get("import_to")
        
        date_from = request.args.get("date_from")
        date_to = request.args.get("date_to")
        status = request.args.get("status")
        q = (request.args.get("q") or "").strip()  # รับค่าคำค้นหา Global Search
        all_time = request.args.get("all_time")  # Flag สำหรับดูทั้งหมด
        mode = request.args.get("mode")  # [NEW] โหมด Order ปัจจุบัน (today)

        shops = Shop.query.order_by(Shop.name.asc()).all()

        # แปลงวันที่
        def _p(s): return parse_date_any(s)
        
        imp_from = _p(import_from_str)
        imp_to = _p(import_to_str)
        d_from = datetime.combine(_p(date_from), datetime.min.time(), tzinfo=TH_TZ) if date_from else None
        d_to = datetime.combine(_p(date_to) + timedelta(days=1), datetime.min.time(), tzinfo=TH_TZ) if date_to else None

        # ตรวจสอบว่ามี Filter วันที่หรือไม่
        has_date_filter = bool(imp_from or imp_to or d_from or d_to)
        
        # ตรวจสอบโหมด All Time (Lock)
        is_all_time = bool(all_time)

        # กำหนด Strategy การดึงข้อมูล
        use_default_view = False
        rows = []

        # Base Filters (Platform/Shop)
        base_filters = {
            "platform": platform if platform else None,
            "shop_id": int(shop_id) if shop_id else None,
        }

        if is_all_time:
            # CASE 1: All Time -> ดึงข้อมูลทั้งหมด ไม่สนวันที่
            # active_only=False เพื่อให้ดึง Packed/Cancelled ด้วย
            filters = base_filters.copy()
            filters["active_only"] = False 
            filters["all_time"] = True
            rows, _ = compute_allocation(db.session, filters)

        elif mode == 'today':
            # [NEW] CASE 1.5: Order ปัจจุบัน (วันนี้)
            # กรองเฉพาะ Import Date = วันนี้ + Order ที่ยกเลิกวันนี้
            today = now_thai().date()
            
            # 1. ดึง Order ที่นำเข้าวันนี้
            filters = base_filters.copy()
            filters["active_only"] = False
            filters["import_from"] = today
            filters["import_to"] = today
            rows_import, _ = compute_allocation(db.session, filters)
            
            # 2. ดึง Order ที่ "ยกเลิกวันนี้" (บวก 7 ชม. เพื่อให้ตรงกับเวลาไทย)
            cancel_today_oids = [
                r[0] for r in db.session.query(CancelledOrder.order_id)
                .filter(func.date(CancelledOrder.imported_at, '+7 hours') == today).all()
            ]
            
            rows_cancel = []
            if cancel_today_oids:
                # ดึงข้อมูลของ Order ที่ cancel วันนี้ (ใช้ all_time แล้ว filter เอาเฉพาะ ID)
                f_cancel = base_filters.copy()
                f_cancel["all_time"] = True
                f_cancel["active_only"] = False
                temp_rows, _ = compute_allocation(db.session, f_cancel)
                rows_cancel = [r for r in temp_rows if r.get("order_id") in cancel_today_oids]
            
            # 3. รวมรายการ (ตัดตัวซ้ำด้วย id)
            seen_ids = set()
            rows = []
            for r in (rows_import + rows_cancel):
                rid = r.get("id")
                if rid not in seen_ids:
                    rows.append(r)
                    seen_ids.add(rid)

        elif has_date_filter:
            # CASE 2: มีการเลือกช่วงเวลา (Import Date หรือ Order Date) -> ดึงตามช่วงเวลานั้น
            filters = base_filters.copy()
            filters["active_only"] = False
            filters["import_from"] = imp_from
            filters["import_to"] = imp_to
            filters["date_from"] = d_from
            filters["date_to"] = d_to
            rows, _ = compute_allocation(db.session, filters)
            
        else:
            # CASE 3: Default View (ไม่มี Filter วันที่ และไม่ใช่ All Time)
            use_default_view = True
            
            # 3.1 ดึง Order ค้างทั้งหมด (Active Orders - All Time)
            f_active = base_filters.copy()
            f_active["active_only"] = True
            rows_active, _ = compute_allocation(db.session, f_active)
            
            # 3.2 ดึง Order จบแล้ว (Packed/Cancelled) ของ "วันนี้" เท่านั้น
            today = now_thai().date()
            f_inactive = base_filters.copy()
            f_inactive["active_only"] = False
            f_inactive["import_from"] = today  # เฉพาะวันนี้
            f_inactive["import_to"] = today
            
            rows_today_all, _ = compute_allocation(db.session, f_inactive)
            
            # คัดเฉพาะ Packed/Cancelled จากของวันนี้
            existing_ids = set(r["id"] for r in rows_active)
            rows = list(rows_active)
            
            for r in rows_today_all:
                if r["id"] not in existing_ids:
                    # ถ้าไม่อยู่ใน Active แสดงว่าเป็น Packed หรือ Cancelled
                    if r.get("is_packed") or r.get("is_cancelled"):
                         rows.append(r)

        # --- Post-Processing Rows ---
        # ดึงเซ็ต/แมป Order ยกเลิก/จ่ายแล้ว/แพ็คแล้ว
        cancelled_map = _cancelled_oids_map()  # dict: order_id -> note
        packed_oids = _orders_packed_set(rows)
        orders_not_in_sbs = _orders_not_in_sbs_set(rows)
        orders_no_sales = _orders_no_sales_set(rows)
        
        # [NEW] ดึง Order ที่ถูกลบ (Soft Delete) และกรองออก
        deleted_oids = _deleted_oids_set()
        rows = [r for r in rows if (r.get("order_id") or "").strip() not in deleted_oids]
        # Process Row Attributes
        totals = _build_allqty_map(rows)
        for r in rows:
            oid = (r.get("order_id") or "").strip()
            
            # เติม stock
            if "stock_qty" not in r:
                sku = (r.get("sku") or "").strip()
                stock_qty = 0
                if sku:
                    prod = Product.query.filter_by(sku=sku).first()
                    if prod and hasattr(prod, "stock_qty"):
                        try:
                            stock_qty = int(prod.stock_qty or 0)
                        except Exception:
                            stock_qty = 0
                    else:
                        st = Stock.query.filter_by(sku=sku).first()
                        stock_qty = int(st.qty) if st and st.qty is not None else 0
                r["stock_qty"] = stock_qty

            r["allqty"] = int(totals.get((r.get("sku") or "").strip(), r.get("qty", 0)) or 0)
            r["accepted"] = bool(r.get("accepted", False))
            r["sales_status"] = r.get("sales_status", None)
            r["logistic"] = r.get("logistic") or r.get("logistic_type") or "-"
            
            # Set Flags
            r["is_cancelled"] = False
            r["is_not_in_sbs"] = False
            r["packed"] = False
            r["cancel_reason"] = ""  # [NEW] เหตุผลการยกเลิก
            r["cancel_at"] = ""      # [NEW] เวลาที่ยกเลิก (สำหรับแสดงใน HTML)
            r["cancel_str"] = ""     # [NEW] ข้อความรวมสำหรับ Excel
            
            # [NEW] เช็คว่า Order นี้เคยแพ็คแล้วหรือยัง (ก่อนถูกยกเลิก)
            r["was_packed"] = (oid in packed_oids)

            if oid in cancelled_map:
                r["allocation_status"] = "CANCELLED"
                r["is_cancelled"] = True
                
                # [NEW] แกะข้อมูล Note และ Time จาก dict ซ้อน
                c_info = cancelled_map[oid]
                note_txt = c_info.get('note', '')
                time_obj = c_info.get('at')
                
                # จัด Format เวลา (แปลงเป็น พ.ศ.)
                time_str = ""
                if time_obj:
                    try:
                        # แปลงเป็น พ.ศ. ถ้าปียังเป็น ค.ศ.
                        if time_obj.year < 2400:
                            time_obj_be = time_obj.replace(year=time_obj.year + 543)
                        else:
                            time_obj_be = time_obj
                        time_str = time_obj_be.strftime("%d/%m/%Y %H:%M")
                    except Exception:
                        pass
                
                r["cancel_reason"] = note_txt
                r["cancel_at"] = time_str  # ส่งไปโชว์ใน HTML
                r["cancel_str"] = f"{note_txt} [เมื่อ: {time_str}]" if time_str else note_txt  # ส่งไป Excel
                
                r["actions_disabled"] = True
            elif oid in packed_oids:
                r["allocation_status"] = "PACKED"
                r["packed"] = True
                r["actions_disabled"] = True
            else:
                r["actions_disabled"] = False
                if oid in orders_not_in_sbs:
                    r["is_not_in_sbs"] = True

        # --- STEP 3: คำนวณ KPI จากข้อมูลทั้งหมดใน Scope (ก่อนถูกซ่อนจากตาราง) ---
        # [แก้ไขจุดที่ 2] ใช้ rows (ซึ่งคือข้อมูลทั้งหมดใน Scope นี้) คำนวณ KPI เลย
        # เพื่อให้ปุ่ม Packed/Cancelled แสดงยอดได้ถูกต้อง แม้ตารางจะไม่ได้โชว์
        
        scope_rows = list(rows)  # สำรองข้อมูลไว้คำนวณ KPI
        
        # Helper lists for KPI counts from scope
        kpi_orders_ready = _orders_ready_set(scope_rows)
        kpi_orders_low = _orders_lowstock_order_set(scope_rows)
        kpi_orders_nosales = _orders_no_sales_set(scope_rows)
        kpi_orders_not_in_sbs = _orders_not_in_sbs_set(scope_rows)
        
        # [แก้ไข] ลบ Order ที่ยกเลิกออกจาก KPI "ยังไม่มีใบขาย" และ "ยังไม่เข้า SBS"
        # เพื่อไม่ให้ยอดเด้งทั้งที่ยกเลิกไปแล้ว
        cancelled_all_ids = set(cancelled_map.keys())
        kpi_orders_nosales = kpi_orders_nosales - cancelled_all_ids
        kpi_orders_not_in_sbs = kpi_orders_not_in_sbs - cancelled_all_ids
        
        # [NEW] คำนวณ Set ของ Order ที่เป็น "ไม่มีสินค้า" หรือ "สินค้าไม่พอส่ง"
        # ใช้ Set เพื่อให้เลข Order ไม่ซ้ำกัน
        kpi_orders_problem = set()
        for r in scope_rows:
            # [แก้ไข] เพิ่มเงื่อนไข: ต้องยังไม่จ่ายงาน (is_issued) ด้วย ถึงจะนับเข้ากอง 3
            if not r.get("packed") and not r.get("is_cancelled") and not r.get("is_issued"):
                status_alloc = (r.get("allocation_status") or "").strip().upper()
                if status_alloc in ("SHORTAGE", "NOT_ENOUGH"):
                    oid = (r.get("order_id") or "").strip()
                    if oid:
                        kpi_orders_problem.add(oid)
        
        # Packed Sets จาก Scope
        kpi_packed_oids = set(r.get("order_id") for r in scope_rows if r.get("packed"))
        
        # [NEW] แยก KPI Order ยกเลิก เป็น 2 กลุ่ม
        # 1. ยกเลิกก่อนแพ็ค (ไม่เคยแพ็ค)
        # 2. ยกเลิกหลังแพ็ค (เคยแพ็คแล้ว)
        kpi_cancel_nopack = set()  # ยกเลิก (ก่อนแพ็ค)
        kpi_cancel_packed = set()  # ยกเลิก (หลังแพ็ค)
        
        for r in scope_rows:
            if r.get("is_cancelled"):
                oid = (r.get("order_id") or "").strip()
                if oid:
                    if r.get("was_packed"):
                        kpi_cancel_packed.add(oid)
                    else:
                        kpi_cancel_nopack.add(oid)

        # --- STEP 4: กรองข้อมูลเพื่อแสดงผลในตาราง (Filtering View) ---
        status_norm = (status or "").strip().upper()
        
        # กรณีที่ 1: มีการค้นหา (Global Search)
        if q:
            q_lower = q.lower()
            filtered_rows = []
            for r in rows:
                search_text = (
                    str(r.get("order_id") or "") + " " +
                    str(r.get("sku") or "") + " " +
                    str(r.get("brand") or "") + " " +
                    str(r.get("model") or "") + " " +
                    str(r.get("shop") or "") + " " +
                    str(r.get("sales_status") or "")
                ).lower()
                if q_lower in search_text:
                    filtered_rows.append(r)
            rows = filtered_rows
            
            # [เพิ่มเติม] ถ้ามีการค้นหา ให้ KPI นับตามผลการค้นหาด้วย
            scope_rows = rows
            
            # Recalculate sets for filtered scope (กรณี search)
            kpi_orders_ready = _orders_ready_set(scope_rows)
            kpi_orders_low = _orders_lowstock_order_set(scope_rows)
            kpi_orders_nosales = _orders_no_sales_set(scope_rows)
            kpi_orders_not_in_sbs = _orders_not_in_sbs_set(scope_rows)
            
            # [แก้ไข] ลบ Order ที่ยกเลิกออกจาก KPI "ยังไม่มีใบขาย" และ "ยังไม่เข้า SBS" (กรณี search)
            cancelled_all_ids = set(cancelled_map.keys())
            kpi_orders_nosales = kpi_orders_nosales - cancelled_all_ids
            kpi_orders_not_in_sbs = kpi_orders_not_in_sbs - cancelled_all_ids
            
            kpi_packed_oids = set(r.get("order_id") for r in scope_rows if r.get("packed"))
            
            # Recalculate kpi_cancel_nopack / kpi_cancel_packed for search
            kpi_cancel_nopack = set()
            kpi_cancel_packed = set()
            for r in scope_rows:
                if r.get("is_cancelled"):
                    oid = (r.get("order_id") or "").strip()
                    if oid:
                        if r.get("was_packed"):
                            kpi_cancel_packed.add(oid)
                        else:
                            kpi_cancel_nopack.add(oid)
            
            # Recalculate kpi_orders_problem for search
            kpi_orders_problem = set()
            for r in scope_rows:
                # [แก้ไข] เพิ่มเงื่อนไข: ต้องยังไม่จ่ายงาน (is_issued) ถึงจะนับเป็นงานค้างกอง 3
                if not r.get("packed") and not r.get("is_cancelled") and not r.get("is_issued"):
                    status_alloc = (r.get("allocation_status") or "").strip().upper()
                    if status_alloc in ("SHORTAGE", "NOT_ENOUGH"):
                        oid = (r.get("order_id") or "").strip()
                        if oid:
                            kpi_orders_problem.add(oid)

        # กรณีที่ 2: ไม่ได้ค้นหา -> ใช้ Logic การกรองตามสถานะ
        else:
            if status_norm == "TOTAL":
                # [NEW] ถ้าเลือก "รวม Order" ให้แสดงทุกอย่างใน Scope (ไม่ซ่อน Packed/Cancelled)
                pass
            elif status_norm == "ORDER_CANCELLED":
                # [แก้ไข] กรองเฉพาะยกเลิกที่ยังไม่เคยแพ็ค (ก่อนแพ็ค)
                rows = [r for r in rows if r.get("is_cancelled") and not r.get("was_packed")]
            elif status_norm == "ORDER_CANCELLED_PACKED":
                # [NEW] กรองเฉพาะยกเลิกหลังแพ็ค (เคยแพ็คแล้ว)
                rows = [r for r in rows if r.get("is_cancelled") and r.get("was_packed")]
            elif status_norm == "ORDER_NOT_IN_SBS":
                rows = [r for r in rows if r.get("is_not_in_sbs")]
            elif status_norm == "ORDER_PROBLEM":
                rows = [r for r in rows if (r.get("order_id") or "").strip() in kpi_orders_problem]
            elif status_norm == "PACKED":
                rows = [r for r in rows if r.get("packed")]
            elif status_norm == "ORDER_READY":
                rows = [r for r in rows if (r.get("order_id") or "").strip() in kpi_orders_ready]
            elif status_norm in {"ORDER_LOW_STOCK", "ORDER_LOW"}:
                rows = [r for r in rows if (r.get("order_id") or "").strip() in kpi_orders_low]
            elif status_norm == "ORDER_NO_SALES":
                rows = [r for r in rows if (r.get("order_id") or "").strip() in kpi_orders_nosales]
            elif status_norm:
                # กรองสถานะรายบรรทัด
                rows = [r for r in rows if (r.get("allocation_status") or "").strip().upper() == status_norm]
            else:
                # Default Table View: ซ่อน Packed/Cancelled ออกจากตารางหลัก
                # แต่ข้อมูลใน scope_rows ยังอยู่ครบ ทำให้ KPI ไม่เป็น 0
                if not status:
                     rows = [r for r in rows if not r.get("packed") and not r.get("is_cancelled")]

        # --- STEP 5: สร้าง Dict KPI ---
        kpis = {
            "total_items": len(scope_rows),
            "total_qty": sum(int(r.get("qty", 0) or 0) for r in scope_rows),
            
            # [แก้ไข] ปรับสูตรนับ "รวม Order":
            # - ถ้ามีการเลือกสถานะ (status ไม่ว่าง) -> นับจาก rows (ตามที่กรอง)
            # - ถ้าไม่มีการเลือกสถานะ (หน้า All Time/ปกติ) -> นับจาก scope_rows (รวม Packed/Cancel)
            "orders_total": len(set(
                r.get("order_id") for r in (rows if status else scope_rows) 
                if r.get("order_id")
            )),
            
            # Active (Pending)
            "orders_unique": len(set(
                r.get("order_id") for r in scope_rows 
                if r.get("order_id") and not r.get("packed") and not r.get("is_cancelled")
            )),
            
            # นับจาก Scope (ไม่ว่าจะซ่อนในตารางหรือไม่ ก็จะโชว์ตัวเลข)
            "ready": sum(1 for r in scope_rows if r.get("allocation_status") == "READY_ACCEPT" and not r.get("packed") and not r.get("is_cancelled")),
            "accepted": sum(1 for r in scope_rows if r.get("allocation_status") == "ACCEPTED"),
            "low": sum(1 for r in scope_rows if r.get("allocation_status") == "LOW_STOCK" and not r.get("packed") and not r.get("is_cancelled")),
            "nostock": sum(1 for r in scope_rows if r.get("allocation_status") == "SHORTAGE" and not r.get("packed") and not r.get("is_cancelled")),
            "notenough": sum(1 for r in scope_rows if r.get("allocation_status") == "NOT_ENOUGH" and not r.get("packed") and not r.get("is_cancelled")),
            
            "packed": len(kpi_packed_oids),
            
            # [แก้ไข] แยกเป็น 2 ยอด: ยกเลิกก่อนแพ็ค / ยกเลิกหลังแพ็ค
            "orders_cancelled": len(kpi_cancel_nopack),
            "orders_cancelled_packed": len(kpi_cancel_packed),

            "orders_ready": len(kpi_orders_ready),
            "orders_low": len(kpi_orders_low),
            "orders_nosales": len(kpi_orders_nosales),
            "orders_not_in_sbs": len(kpi_orders_not_in_sbs),
            "orders_problem": len(kpi_orders_problem),
            
            # [NEW] จำนวน Order ที่ถูกลบ (Soft Delete)
            "orders_deleted": len(deleted_oids),
        }

        # Sort
        def _sort_key(r):
            return ((r.get("order_id") or ""), (r.get("platform") or ""), (r.get("shop") or ""), (r.get("sku") or ""))
        rows = sorted(rows, key=_sort_key)

        # --- [แก้ไขจุดที่ 1] คำนวณยอด "Order จ่ายแล้ว" (Issued) ให้ขยับตามฟิลเตอร์ ---
        iq = db.session.query(IssuedOrder.order_id)\
               .join(OrderLine, OrderLine.order_id == IssuedOrder.order_id)\
               .join(Shop, Shop.id == OrderLine.shop_id)

        # 1. กรอง Platform / Shop
        if platform:
            iq = iq.filter(Shop.platform == platform)
        if shop_id:
            iq = iq.filter(Shop.id == int(shop_id))

        # 2. [เพิ่ม] กรองตามคำค้นหา (Global Search)
        if q:
            iq = iq.filter(IssuedOrder.order_id.contains(q))

        # 3. กรองวันที่
        if is_all_time:
            # All Time -> ไม่กรองวันที่
            pass
        elif mode == 'today':
            # [NEW] ถ้าโหมดวันนี้ -> กรองเฉพาะออเดอร์ที่นำเข้าวันนี้
            iq = iq.filter(OrderLine.import_date == now_thai().date())
        elif has_date_filter:
            if imp_from: iq = iq.filter(OrderLine.import_date >= imp_from)
            if imp_to:   iq = iq.filter(OrderLine.import_date <= imp_to)
            if d_from:   iq = iq.filter(OrderLine.order_time >= d_from)
            if d_to:     iq = iq.filter(OrderLine.order_time < d_to)
        else:
            # [แก้ไข] Default View -> ไม่กรองวันที่ เพื่อให้นับยอดจ่ายแล้วทั้งหมด
            # ให้ตรงกับความเข้าใจว่าเป็น Order ที่เคยค้าง (ตาม Scope ร้านค้า)
            pass

        # ใช้ distinct เพราะ 1 Order มีหลาย Line
        issued_count = iq.distinct().count()

        return render_template(
            "dashboard.html",
            rows=rows,
            shops=shops,
            platform_sel=platform,
            shop_sel=shop_id,
            import_from_sel=import_from_str,  # ส่งกลับไปแสดงผล
            import_to_sel=import_to_str,      # ส่งกลับไปแสดงผล
            status_sel=status,
            date_from_sel=date_from,
            date_to_sel=date_to,
            kpis=kpis,
            packed_oids=packed_oids,
            issued_count=issued_count,
            all_time=all_time,
            use_default_view=use_default_view,
            q=q,
            mode=mode,  # [NEW] ส่งค่า mode ไปหน้าเว็บ
            ready_oids=kpi_orders_ready,  # [NEW] ส่งรายชื่อ Order ที่พร้อม 100% ไปใช้หน้าเว็บ
        )

    # =========[ NEW ]=========  กดรับ Order ในหน้า Dashboard
    @app.post("/dashboard/accept_order")
    @login_required
    def dashboard_accept_order():
        cu = current_user()
        if not cu:
            flash("กรุณาเข้าสู่ระบบก่อน", "danger")
            return redirect(url_for("login"))

        order_id = request.form.get("order_id")
        sku = request.form.get("sku")
        platform = request.form.get("platform")
        shop_id = request.form.get("shop_id")

        if not order_id or not sku:
            flash("ข้อมูลไม่ครบถ้วน", "danger")
            return redirect(url_for("dashboard"))

        # อัปเดท OrderLine ให้เป็น accepted
        try:
            ol = OrderLine.query.filter_by(order_id=order_id, sku=sku).first()
            if ol:
                ol.accepted = True
                ol.accepted_at = now_thai()
                ol.accepted_by_user_id = cu.id
                ol.accepted_by_username = cu.username
                db.session.commit()
                flash(f"รับออเดอร์ {order_id} (SKU: {sku}) สำเร็จ", "success")
            else:
                flash("ไม่พบรายการที่ต้องการรับ", "warning")
        except Exception as e:
            db.session.rollback()
            app.logger.exception("Accept order failed")
            flash(f"เกิดข้อผิดพลาด: {e}", "danger")

        # redirect กลับไปหน้าเดิมพร้อมฟิลเตอร์
        return redirect(url_for("dashboard", platform=platform, shop_id=shop_id))
    # =========[ /NEW ]=========

    # -----------------------
    # Import endpoints
    # -----------------------
    @app.route("/import/orders", methods=["GET", "POST"])
    @login_required
    def import_orders_view():
        # 1. จัดการวันที่แบบ Range (จาก...ถึง...)
        date_from_str = request.args.get("date_from")
        date_to_str = request.args.get("date_to")
        # Fallback: รองรับ filter_date เดิมสำหรับ backwards compatibility
        filter_date_str = request.args.get("filter_date")
        
        today_str = now_thai().date().isoformat()
        
        # ถ้ามี filter_date เดิม ให้ใช้เป็น date_from และ date_to เดียวกัน
        if filter_date_str and not date_from_str:
            date_from_str = filter_date_str
            date_to_str = filter_date_str
        
        if not date_from_str:
            date_from_str = today_str
        if not date_to_str:
            date_to_str = today_str
            
        try:
            view_date_from = datetime.strptime(date_from_str, "%Y-%m-%d").date()
        except Exception:
            view_date_from = now_thai().date()
        try:
            view_date_to = datetime.strptime(date_to_str, "%Y-%m-%d").date()
        except Exception:
            view_date_to = now_thai().date()

        # สร้างวันที่แบบไทยสำหรับแสดงผล (วัน/เดือน/ปี พ.ศ.)
        def _th_date(d):
            return f"{d.day:02d}/{d.month:02d}/{d.year + 543}"
        
        view_date_range_thai = _th_date(view_date_from)
        if view_date_from != view_date_to:
            view_date_range_thai += f" - {_th_date(view_date_to)}"

        # 2. Handle POST (การนำเข้า)
        if request.method == "POST":
            platform = request.form.get("platform")
            shop_name = request.form.get("shop_name")
            f = request.files.get("file")
            
            # วันที่นำเข้า = วันปัจจุบันเสมอ
            current_import_date = now_thai().date()

            if not platform or not f:
                flash("กรุณาเลือกแพลตฟอร์ม และเลือกไฟล์", "danger")
                return redirect(url_for("import_orders_view"))
            try:
                df = pd.read_excel(f)
                # >>> สร้าง/ใช้ร้านเดิมก่อนเสมอ (กัน UNIQUE พัง)
                _ensure_shops_from_df(df, platform=platform, default_shop_name=shop_name)
                
                # เรียก Importer ใหม่
                stats = import_orders(
                    df, platform=platform, shop_name=shop_name, import_date=current_import_date
                )
                
                # เก็บ Batch Data (IDs) ลง Log
                batch_data = json.dumps({
                    "added_ids": stats.get("added_ids", []),
                    "duplicate_ids": stats.get("duplicate_ids", []),
                    "duplicate_old_ids": stats.get("duplicate_old_ids", []),
                    "duplicate_today_ids": stats.get("duplicate_today_ids", []),
                    "failed_ids": stats.get("failed_ids", [])
                }, ensure_ascii=False)
                
                # บันทึก Log ลง DB
                log_entry = ImportLog(
                    import_date=current_import_date,
                    platform=platform,
                    filename=f.filename or "uploaded_file.xlsx",
                    added_count=stats["added"],
                    duplicates_count=stats["duplicates"],
                    failed_count=stats["failed"],
                    error_details=json.dumps(stats["errors"], ensure_ascii=False) if stats["errors"] else "[]"
                )
                # เพิ่ม batch_data และ shop_name ถ้าคอลัมน์มีอยู่
                if hasattr(log_entry, 'batch_data'):
                    log_entry.batch_data = batch_data
                if hasattr(log_entry, 'shop_name'):
                    log_entry.shop_name = shop_name or ""
                if hasattr(log_entry, 'duplicates_same_day'):
                    log_entry.duplicates_same_day = stats.get("duplicates_today", 0)
                db.session.add(log_entry)
                db.session.commit()

                # สร้างข้อความแจ้งเตือนแยกประเภทซ้ำ
                dup_old = stats.get('duplicates_old', 0)
                dup_today = stats.get('duplicates_today', 0)
                dup_msg = f"ซ้ำข้ามวัน {dup_old}"
                if dup_today > 0:
                    dup_msg += f" (ซ้ำวันนี้ {dup_today} - ไม่นับ)"
                
                flash(
                    f"นำเข้า: เพิ่ม {stats['added']} | {dup_msg} | ไม่สำเร็จ {stats['failed']}", 
                    "success" if stats['failed'] == 0 else "warning"
                )
                # Redirect กลับมาหน้า Dashboard ของวันที่นำเข้า
                return redirect(url_for('import_orders_view', date_from=current_import_date.isoformat(), date_to=current_import_date.isoformat()))

            except Exception as e:
                db.session.rollback()
                flash(f"เกิดข้อผิดพลาดในการนำเข้าออเดอร์: {e}", "danger")
                return redirect(url_for("import_orders_view"))

        # 3. คำนวณ Dashboard (นับ Unique Order IDs) - รองรับ Date Range
        
        # A. ยอดสำเร็จ (Success): นับ Order ID ไม่ซ้ำจาก OrderLine จริงๆ ตามช่วงวันที่เลือก
        success_count = db.session.query(func.count(func.distinct(OrderLine.order_id)))\
            .filter(OrderLine.import_date >= view_date_from)\
            .filter(OrderLine.import_date <= view_date_to).scalar() or 0

        # B. ดึง Logs ของช่วงวันนั้นเพื่อนับ Duplicate และ Failed (ตาม Order ID ไม่ซ้ำ)
        logs = ImportLog.query.filter(
            ImportLog.import_date >= view_date_from,
            ImportLog.import_date <= view_date_to
        ).order_by(ImportLog.created_at.desc()).all()
        
        # [แก้ไขใหม่] ใช้ Set เก็บ ID เพื่อตัดตัวซ้ำ (Unique Count)
        # ต่อให้นำเข้าไฟล์เดิม 10 รอบ ID เดิมก็จะถูกนับแค่ 1 ครั้ง
        log_dup_old_ids: set[str] = set()  # ซ้ำข้ามวัน (แสดงในการ์ด)
        log_dup_today_ids: set[str] = set()  # ซ้ำในวัน (ไม่แสดงในการ์ด)
        log_failed_ids: set[str] = set()  # Failed IDs (Unique) - เก็บ Order ID ที่ Failed
        anon_error_set: set[str] = set()  # เก็บข้อความ Error ที่ไม่มี ID (เช่น "แถว 12...") (ไม่ซ้ำ)
        grouped_errors: list[dict] = []  # เก็บ Error แยกตาม Log
        
        for l in logs:
            # ดึง Batch Data (IDs)
            batch_data_str = getattr(l, 'batch_data', None)
            batch_data = {}
            if batch_data_str:
                try:
                    batch_data = json.loads(batch_data_str)
                    log_dup_old_ids.update(batch_data.get("duplicate_old_ids", []))
                    log_dup_today_ids.update(batch_data.get("duplicate_today_ids", []))
                    
                    # Failed: เก็บ ID ที่ระบุได้ลง Set (ตัดซ้ำอัตโนมัติ)
                    current_failed_ids = batch_data.get("failed_ids", [])
                    for fid in current_failed_ids:
                        if fid:
                            log_failed_ids.add(str(fid).strip())
                except Exception:
                    pass
                
            # ดึง Error Details และจัดกลุ่ม
            if l.error_details and l.error_details != "[]":
                try:
                    errs = json.loads(l.error_details)
                    if errs:
                        # [สำคัญ] วนลูปเช็ค Error แต่ละบรรทัดเพื่อตัดซ้ำ
                        for err_msg in errs:
                            err_msg = str(err_msg).strip()
                            # ถ้าเป็น Error ที่มีคำว่า "Order " แสดงว่ามี ID แล้ว -> ข้าม (เพราะถูกนับใน log_failed_ids แล้ว)
                            # แต่ถ้าเป็น "แถว 12..." หรือ Error อื่นๆ -> เก็บลง Set
                            if not err_msg.startswith("Order "):
                                anon_error_set.add(err_msg)
                        
                        # แปลงเวลา Log เป็นไทย (UTC+7)
                        ts = l.created_at
                        time_str = "-"
                        if ts:
                            try:
                                ts_thai = ts + timedelta(hours=7)
                                year_be = ts_thai.year + 543
                                time_str = f"{ts_thai.day:02d}/{ts_thai.month:02d}/{year_be} ({ts_thai.hour:02d}:{ts_thai.minute:02d})"
                            except Exception:
                                pass
                        
                        grouped_errors.append({
                            "platform": l.platform or "-",
                            "shop_name": getattr(l, 'shop_name', '') or l.filename or "-",
                            "filename": l.filename or "-",
                            "time": time_str,
                            "errors": errs
                        })
                except Exception:
                    pass
        
        # C. คำนวณยอดสรุป (นับจาก Set ที่ตัดซ้ำแล้ว - ไม่ใช้ Fallback บวกสะสมอีกต่อไป)
        # ยอด Failed = (จำนวน Order ID ที่ไม่ซ้ำ) + (จำนวนข้อความ Error แถวที่ไม่ซ้ำ)
        real_fail_count = len(log_failed_ids) + len(anon_error_set)
        dup_old_count = len(log_dup_old_ids)  # ซ้ำข้ามวัน (แสดงในการ์ด) - Unique เท่านั้น
        dup_today_count = len(log_dup_today_ids)  # ซ้ำในวัน (ไม่แสดงในการ์ด)
        
        # Total = Success + Failed (ไม่รวม Duplicate เพราะคือออเดอร์เดิม)
        total_count = success_count + real_fail_count
        
        # D. ดึงข้อมูลร้านและ URL
        shops = Shop.query.order_by(Shop.name.asc()).all()
        
        # [แก้ไข] ใช้ SQL ดึง URL โดยตรง (แก้ปัญหารีเฟรชแล้วหาย)
        shop_urls = {}
        try:
            # ดึง name และ google_sheet_url จากตาราง shops ตรงๆ
            rows_url = db.session.execute(text("SELECT name, google_sheet_url FROM shops")).fetchall()
            for r_name, r_url in rows_url:
                shop_urls[r_name] = r_url or ""
        except Exception as e:
            app.logger.warning(f"Fetch shop urls failed: {e}")
        
        return render_template(
            "import_orders.html", 
            shops=shops,
            shop_urls=shop_urls,
            date_from=view_date_from.isoformat(),
            date_to=view_date_to.isoformat(),
            view_date_range_thai=view_date_range_thai,
            dash={
                "total": total_count,
                "success": success_count,
                "duplicate": dup_old_count,        # แสดงเฉพาะซ้ำข้ามวัน
                "duplicate_today": dup_today_count, # ซ้ำในวัน (ไม่แสดงในการ์ด แต่เก็บไว้อ้างอิง)
                "failed": real_fail_count,
                "grouped_errors": grouped_errors   # ส่งแบบกลุ่มไป
            }
        )

    # =========[ NEW ]=========
    # Import Orders จาก Google Sheet
    @app.route("/import/orders/gsheet", methods=["POST"])
    @login_required
    def import_orders_gsheet():
        platform = request.form.get("platform")
        shop_name = request.form.get("shop_name")
        sheet_url = request.form.get("sheet_url")

        if not platform or not sheet_url:
            flash("กรุณาระบุแพลตฟอร์มและลิงก์ Google Sheet", "danger")
            return redirect(url_for("import_orders_view"))

        # [NEW] อัปเดต URL ล่าสุดให้ร้านอัตโนมัติเมื่อกดดึงข้อมูล
        # Logic ใหม่: ถ้าไม่ได้ระบุชื่อร้าน ให้บันทึกเข้า Platform Name (เป็น URL กลาง)
        platform_std = normalize_platform(platform)
        
        # ชื่อที่จะใช้บันทึก URL (ถ้ามีชื่อร้านใช้ชื่อร้าน ถ้าไม่มีใช้ชื่อ Platform)
        target_save_name = shop_name.strip() if shop_name and shop_name.strip() else platform_std
        
        if sheet_url:
            # ค้นหาร้าน หรือ สร้างใหม่ถ้าไม่เจอ (เพื่อเก็บ URL)
            s = Shop.query.filter_by(platform=platform_std, name=target_save_name).first()
            if not s:
                # ถ้าไม่เจอ ลองหาจากชื่ออย่างเดียว (กรณีชื่อ Platform)
                s = Shop.query.filter_by(name=target_save_name).first()
            
            if not s:
                s = Shop(platform=platform_std, name=target_save_name)
                db.session.add(s)
                db.session.commit()  # Commit เพื่อให้ได้ ID มาใช้
            
            # บันทึก URL (ใช้ SQL ตรงๆ)
            if sheet_url:
                db.session.execute(
                    text("UPDATE shops SET google_sheet_url = :u WHERE id = :id"),
                    {"u": sheet_url, "id": s.id}
                )
                db.session.commit()

        # กำหนดชื่อ Tab ตามแพลตฟอร์ม
        target_tab_name = ""
        if platform == "Shopee":
            target_tab_name = "Import_Shopee"
        elif platform == "Lazada":
            target_tab_name = "Import_Lazada"
        elif platform == "TikTok":
            target_tab_name = "Import_Tiktok"
        else:
            target_tab_name = "Import_Order_other"

        try:
            # 1. เชื่อมต่อ Google API
            creds = get_google_credentials()
            client = gspread.authorize(creds)

            # 2. เปิด Google Sheet
            sheet = client.open_by_url(sheet_url)
            
            # 3. เลือก Tab ตามชื่อ
            try:
                worksheet = sheet.worksheet(target_tab_name)
            except gspread.WorksheetNotFound:
                flash(f"❌ ไม่พบ Tab ชื่อ '{target_tab_name}' ใน Google Sheet นี้", "danger")
                return redirect(url_for("import_orders_view"))
            
            # 4. ดึงข้อมูล
            data = worksheet.get_all_records()
            if not data:
                flash(f"Tab '{target_tab_name}' ไม่มีข้อมูล", "warning")
                return redirect(url_for("import_orders_view"))

            # 5. แปลงเป็น DataFrame และนำเข้า
            df = pd.DataFrame(data)
            
            # สร้าง/เช็คชื่อร้าน
            _ensure_shops_from_df(df, platform=platform, default_shop_name=shop_name)
            
            # วันที่นำเข้า = วันปัจจุบันเสมอ
            current_import_date = now_thai().date()
            
            # เรียก Importer ใหม่
            stats = import_orders(
                df, platform=platform, shop_name=shop_name, import_date=current_import_date
            )
            
            # เก็บ Batch Data (IDs) ลง Log
            batch_data = json.dumps({
                "added_ids": stats.get("added_ids", []),
                "duplicate_ids": stats.get("duplicate_ids", []),
                "duplicate_old_ids": stats.get("duplicate_old_ids", []),
                "duplicate_today_ids": stats.get("duplicate_today_ids", []),
                "failed_ids": stats.get("failed_ids", [])
            }, ensure_ascii=False)
            
            # บันทึก Log ลง DB
            log_entry = ImportLog(
                import_date=current_import_date,
                platform=platform,
                filename=f"Google Sheet ({target_tab_name})",
                added_count=stats["added"],
                duplicates_count=stats["duplicates"],
                failed_count=stats["failed"],
                error_details=json.dumps(stats["errors"], ensure_ascii=False) if stats["errors"] else "[]"
            )
            # เพิ่ม batch_data, shop_name และ duplicates_same_day ถ้าคอลัมน์มีอยู่
            if hasattr(log_entry, 'batch_data'):
                log_entry.batch_data = batch_data
            if hasattr(log_entry, 'shop_name'):
                log_entry.shop_name = shop_name or ""
            if hasattr(log_entry, 'duplicates_same_day'):
                log_entry.duplicates_same_day = stats.get("duplicates_today", 0)
            db.session.add(log_entry)
            db.session.commit()
            
            # สร้างข้อความแจ้งเตือนแยกประเภทซ้ำ
            dup_old = stats.get('duplicates_old', 0)
            dup_today = stats.get('duplicates_today', 0)
            dup_msg = f"ซ้ำข้ามวัน {dup_old}"
            if dup_today > 0:
                dup_msg += f" (ซ้ำวันนี้ {dup_today} - ไม่นับ)"
            
            flash(
                f"✅ ดึงข้อมูลจาก {target_tab_name}: เพิ่ม {stats['added']} | {dup_msg} | ไม่สำเร็จ {stats['failed']}", 
                "success" if stats['failed'] == 0 else "warning"
            )
            return redirect(url_for('import_orders_view', date_from=current_import_date.isoformat(), date_to=current_import_date.isoformat()))

        except Exception as e:
            db.session.rollback()
            if "PERMISSION_DENIED" in str(e):
                flash("❌ บอทเข้าถึงไฟล์ไม่ได้! กรุณาตรวจสอบสิทธิ์การแชร์ (Share) ของ Google Sheet", "danger")
            else:
                app.logger.exception("Google Sheet Import Error")
                flash(f"เกิดข้อผิดพลาด: {str(e)}", "danger")
            return redirect(url_for("import_orders_view"))

    # =========[ NEW ]=========
    # ล้างประวัติ Import Log (พร้อมออปชั่นลบข้อมูลออเดอร์จริง)
    @app.route("/import/orders/clear_log", methods=["POST"])
    @login_required
    def clear_import_log():
        cu = current_user()
        if not cu or cu.role not in {"admin", "staff"}:
            flash("ไม่มีสิทธิ์ลบประวัติ", "danger")
            return redirect(url_for("import_orders_view"))
            
        mode = request.form.get("mode")  # 'range' or 'all'
        delete_data = request.form.get("delete_data")  # 'yes' ถ้าติ๊ก checkbox
        
        try:
            order_deleted_count = 0
            log_deleted_count = 0
            
            if mode == 'all':
                # 1. ถ้าเลือกติ๊กลบข้อมูล -> ลบออเดอร์ทั้งหมด
                if delete_data == 'yes':
                    # ลบ OrderLines ทั้งหมด
                    order_deleted_count = OrderLine.query.delete()
                    # ลบ DeletedOrder ถังขยะด้วย
                    try:
                        db.session.query(DeletedOrder).delete()
                    except Exception:
                        pass
                
                # 2. ลบ Log ทั้งหมด
                log_deleted_count = ImportLog.query.delete()
                
                if order_deleted_count > 0:
                    msg = f"ล้างเกลี้ยง! (Log {log_deleted_count} รายการ, ออเดอร์ {order_deleted_count} รายการ)"
                else:
                    msg = f"ล้างประวัติการนำเข้าทั้งหมดเรียบร้อย ({log_deleted_count} รายการ)"

            else:
                # ลบตามช่วงวันที่
                d_from_str = request.form.get("date_from")
                d_to_str = request.form.get("date_to")
                
                if not d_from_str or not d_to_str:
                    flash("ระบุวันที่ไม่ถูกต้อง", "warning")
                    return redirect(url_for("import_orders_view"))
                    
                d_from = datetime.strptime(d_from_str, "%Y-%m-%d").date()
                d_to = datetime.strptime(d_to_str, "%Y-%m-%d").date()
                
                # 1. ถ้าเลือกติ๊กลบข้อมูล -> ลบออเดอร์ในช่วงวันที่นำเข้านั้น
                if delete_data == 'yes':
                    order_deleted_count = OrderLine.query.filter(
                        OrderLine.import_date >= d_from,
                        OrderLine.import_date <= d_to
                    ).delete(synchronize_session=False)
                
                # 2. ลบ Log ในช่วงวันที่
                log_deleted_count = ImportLog.query.filter(
                    ImportLog.import_date >= d_from,
                    ImportLog.import_date <= d_to
                ).delete(synchronize_session=False)
                
                if order_deleted_count > 0:
                    msg = f"ล้างข้อมูลช่วง {to_be_date_str(d_from)} - {to_be_date_str(d_to)} เรียบร้อย (Log {log_deleted_count}, ออเดอร์ {order_deleted_count})"
                else:
                    msg = f"ล้างประวัติช่วง {to_be_date_str(d_from)} - {to_be_date_str(d_to)} เรียบร้อย ({log_deleted_count} รายการ)"
                
            db.session.commit()
            flash(msg, "success")
            
        except Exception as e:
            db.session.rollback()
            flash(f"เกิดข้อผิดพลาด: {e}", "danger")
            
        return redirect(url_for("import_orders_view"))
    # =========[ /NEW ]=========

    # =========[ NEW ]=========
    # Import Orders ยกเลิก + Template
    @app.route("/import/cancel/template")
    @login_required
    def import_cancel_template():
        fmt = (request.args.get("format") or "xlsx").lower()
        sample = ["ORDER-001", "ORDER-002", "ORDER-ABC-003"]

        if fmt == "xlsx" and _OPENPYXL_OK:
            wb = Workbook()
            ws = wb.active
            ws.title = "cancelled_orders"
            ws["A1"] = "order_id"
            for i, no in enumerate(sample, start=2):
                ws[f"A{i}"] = no
            bio = BytesIO()
            wb.save(bio)
            bio.seek(0)
            return send_file(
                bio,
                as_attachment=True,
                download_name="template_import_orders_cancel.xlsx",
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        # Fallback CSV
        csv_io = BytesIO()
        csv_io.write(("order_id\n" + "\n".join(sample)).encode("utf-8-sig"))
        csv_io.seek(0)
        return send_file(
            csv_io,
            as_attachment=True,
            download_name="template_import_orders_cancel.csv",
            mimetype="text/csv",
        )

    # =========================================================
    #  Helper: Unique Daily Stats สำหรับ Cancel Orders
    # =========================================================
    def _get_cancel_daily_stats(view_date):
        """
        คำนวณสถิติ Dashboard ตามโจทย์:
        1. Success = นับจำนวน Order ที่ถูกสร้างใน DB 'วันนี้' จริงๆ (New Success)
        2. Duplicate = (จำนวน Unique ID ที่ตรวจสอบวันนี้) - (Success)
           * วิธีนี้ช่วยให้ถ้านำเข้าไฟล์เดิมซ้ำ 3 รอบ ยอด Duplicate จะคงที่ ไม่เบิ้ลตามจำนวนรอบ
        """
        # 1. ยอด Success: นับจาก DB โดยตรง (บวก 7 ชม. เพื่อให้ตรงกับเวลาไทย)
        success_count = CancelledOrder.query.filter(
            func.date(CancelledOrder.imported_at, '+7 hours') == view_date
        ).count()

        # 2. ยอด Unique Input: ดึง ID ทั้งหมดที่เคยนำเข้าวันนี้จาก Log มาทำ Unique Set
        logs = ImportLog.query.filter(
            ImportLog.import_date == view_date,
            ImportLog.platform == 'CANCEL_SYSTEM'
        ).all()
        
        all_attempted_ids = set()
        failed_total = 0
        
        for log in logs:
            failed_total += (log.failed_count or 0)
            # ดึง ID ที่บันทึกไว้ใน batch_data มารวมกัน
            if log.batch_data:
                try:
                    data = json.loads(log.batch_data)
                    if "ids" in data:
                        all_attempted_ids.update(data["ids"])
                except:
                    pass
                    
        # 3. ยอด Duplicate: (ที่เช็คทั้งหมด - ที่เข้า DB ได้)
        # หมายถึงรายการที่มีใน DB อยู่แล้ว (ไม่ว่าจะเก่าหรือใหม่)
        duplicate_count = max(0, len(all_attempted_ids) - success_count)

        return {
            "success": success_count,
            "duplicate": duplicate_count,
            "failed": failed_total
        }

    def _process_cancel_import(order_ids: list, source_name: str, user_id: int):
        """ประมวลผลการนำเข้า order ยกเลิก"""
        # 1. คลีนข้อมูลและตัดตัวซ้ำในไฟล์ (Internal Deduplicate)
        unique_input_ids = set()
        for oid in order_ids:
            s = str(oid).strip()
            if s: unique_input_ids.add(s)
        
        if not unique_input_ids:
            return 0, 0

        # 2. หา ID ที่มีอยู่แล้วใน DB (เช็คซ้ำทั้งหมด ไม่สนวันที่)
        existing_query = db.session.query(CancelledOrder.order_id).filter(
            CancelledOrder.order_id.in_(unique_input_ids)
        ).all()
        existing_ids = {r[0] for r in existing_query}

        # 3. แยกรายการ ใหม่ vs ซ้ำ
        new_ids = unique_input_ids - existing_ids
        
        # 4. บันทึกรายการใหม่ลง DB
        if new_ids:
            timestamp = datetime.utcnow()
            new_entries = []
            for oid in new_ids:
                new_entries.append(CancelledOrder(
                    order_id=oid,
                    imported_at=timestamp,
                    imported_by_user_id=user_id,
                    note=f"Import via {source_name}"
                ))
            db.session.bulk_save_objects(new_entries)
            db.session.commit()

        # 5. บันทึก Log พร้อม batch_data (สำคัญสำหรับคำนวณหน้าเว็บ)
        log = ImportLog(
            import_date=now_thai().date(),
            platform="CANCEL_SYSTEM",
            shop_name="-",
            filename=source_name,
            added_count=len(new_ids),
            duplicates_count=len(existing_ids),
            failed_count=0,
            # เก็บ ID ทั้งหมดเพื่อไปทำ Union ที่ Dashboard
            batch_data=json.dumps({"ids": list(unique_input_ids)})
        )
        db.session.add(log)
        db.session.commit()

        return len(new_ids), len(existing_ids)

    @app.route("/import/cancel", methods=["GET"])
    @login_required
    def import_cancel_view():
        _ensure_cancel_table()
        
        # ตรวจสอบให้ ImportLog table มีอยู่
        try:
            ImportLog.__table__.create(bind=db.engine, checkfirst=True)
        except Exception:
            pass

        # [แก้ไข] รับค่าเป็นช่วงวันที่
        date_from_str = request.args.get("date_from")
        date_to_str = request.args.get("date_to")
        
        # Default: วันนี้
        if not date_from_str:
            date_from_str = now_thai().date().isoformat()
        if not date_to_str:
            date_to_str = now_thai().date().isoformat()

        d_from = parse_date_any(date_from_str)
        d_to = parse_date_any(date_to_str)

        # 1. ดึง Log ตามช่วงเวลา
        logs = ImportLog.query.filter(
            ImportLog.import_date >= d_from,
            ImportLog.import_date <= d_to,
            ImportLog.platform == 'CANCEL_SYSTEM'
        ).order_by(ImportLog.created_at.desc()).all()

        # 2. คำนวณ Success (นับจาก DB จริงที่สร้างในช่วงเวลานั้น + ปรับ Timezone ไทย UTC+7)
        success_count = CancelledOrder.query.filter(
            func.date(CancelledOrder.imported_at, '+7 hours') >= d_from,
            func.date(CancelledOrder.imported_at, '+7 hours') <= d_to
        ).count()

        # 3. [แก้ตรงนี้] คำนวณ Duplicate แบบไม่นับเบิ้ล (Unique ID)
        all_attempted_ids = set()  # ใช้ Set เพื่อตัดตัวซ้ำอัตโนมัติ
        failed_sum = 0

        for log in logs:
            failed_sum += (log.failed_count or 0)
            # ดึงรายชื่อ ID ที่เคยยิงเข้ามาทั้งหมดจาก batch_data
            if log.batch_data:
                try:
                    data = json.loads(log.batch_data)
                    # ในฟังก์ชัน _process_cancel_import เราบันทึก key "ids" เอาไว้
                    if "ids" in data:
                        all_attempted_ids.update(data["ids"])
                except Exception:
                    pass
        
        # สูตร: จำนวน ID ทั้งหมดที่ไม่ซ้ำที่ยิงเข้ามา - จำนวนที่สำเร็จจริง = จำนวนที่ซ้ำ
        unique_duplicate_count = max(0, len(all_attempted_ids) - success_count)

        stats = {
            "success": success_count,
            "duplicate": unique_duplicate_count,  # ใช้ค่าใหม่ที่คำนวณแบบ Unique
            "failed": failed_sum
        }

        # ดึง URL ที่บันทึกไว้ (Config) จาก Shop
        saved_url = ""
        try:
            config_row = db.session.execute(
                text("SELECT google_sheet_url FROM shops WHERE platform = 'CANCEL_SYSTEM' AND name = 'GoogleSheet' LIMIT 1")
            ).fetchone()
            if config_row and config_row[0]:
                saved_url = config_row[0]
        except Exception:
            pass
            
        # เตรียมวันที่ภาษาไทยสำหรับหัวข้อ Modal
        date_from_thai = to_be_date_str(d_from) if d_from else ""
        date_to_thai = to_be_date_str(d_to) if d_to else ""

        return render_template(
            "import_cancel.html",
            date_from=date_from_str,
            date_to=date_to_str,
            date_from_thai=date_from_thai,
            date_to_thai=date_to_thai,
            stats=stats,
            logs=logs,
            saved_url=saved_url
        )

    @app.route("/import/cancel/action", methods=["POST"])
    @login_required
    def import_cancel_action():
        _ensure_cancel_table()
        
        # ตรวจสอบให้ ImportLog table มีอยู่
        try:
            ImportLog.__table__.create(bind=db.engine, checkfirst=True)
        except Exception:
            pass

        cu = current_user()
        if not cu or cu.role not in {"admin", "staff"}:
            flash("ต้องเป็นผู้ดูแลระบบหรือพนักงานเท่านั้น", "danger")
            return redirect(url_for("dashboard"))

        mode = request.form.get("mode")  # 'file' or 'gsheet'
        
        try:
            order_ids = []
            source_name = "File"
            
            # กรณี 1: Google Sheet
            if mode == "gsheet":
                url = request.form.get("sheet_url", "").strip()
                if not url:
                    flash("กรุณาระบุ URL", "warning")
                    return redirect(url_for("import_cancel_view"))
                    
                # Connect Google Sheet
                creds = get_google_credentials()
                client = gspread.authorize(creds)
                sh = client.open_by_url(url)
                source_name = f"GSheet: {sh.title}"
                
                # จับ Tab ชื่อ Import_Cancel
                try:
                    ws = sh.worksheet("Import_Cancel")
                except gspread.WorksheetNotFound:
                    flash("ไม่พบ Tab ชื่อ 'Import_Cancel' ใน Google Sheet นี้", "danger")
                    return redirect(url_for("import_cancel_view"))
                    
                # อ่านข้อมูล (หาคอลัมน์ Order ID หรือ อ่านคอลัมน์แรก)
                rows = ws.get_all_values()
                if rows:
                    header = [str(h).lower().strip() for h in rows[0]]
                    col_idx = 0  # Default คอลัมน์แรก
                    for idx, h in enumerate(header):
                        if h in ["order_id", "order id", "order_no", "เลขคำสั่งซื้อ", "เลขออเดอร์"]:
                            col_idx = idx
                            break
                    
                    # เก็บ ID (ข้าม Header)
                    for r in rows[1:]:
                        if len(r) > col_idx:
                            val = str(r[col_idx]).strip()
                            if val: order_ids.append(val)

            # กรณี 2: Upload File
            elif mode == "file":
                f = request.files.get("file")
                if f and f.filename:
                    order_ids = _parse_order_ids_from_upload(f)
                    source_name = f.filename
                else:
                    flash("โปรดเลือกไฟล์ Excel/CSV", "warning")
                    return redirect(url_for("import_cancel_view"))

            # ประมวลผล
            if order_ids:
                added, dups = _process_cancel_import(order_ids, source_name, cu.id)
                flash(f"✅ นำเข้าสำเร็จ: เพิ่มใหม่ {added}, ซ้ำ {dups} รายการ", "success")
            else:
                flash("ไม่พบข้อมูล Order ID", "warning")

        except Exception as e:
            db.session.rollback()
            app.logger.exception("Import cancelled orders failed")
            flash(f"เกิดข้อผิดพลาด: {e}", "danger")

        return redirect(url_for("import_cancel_view"))

    # =========[ NEW ]=========
    # ล้างประวัติ Import Cancel Log
    @app.route("/import/cancel/clear_log", methods=["POST"])
    @login_required
    def clear_cancel_log():
        cu = current_user()
        if not cu or cu.role not in {"admin", "staff"}:
            flash("ไม่มีสิทธิ์ลบประวัติ", "danger")
            return redirect(url_for("import_cancel_view"))
            
        mode = request.form.get("mode")  # 'range' or 'all'
        delete_data = request.form.get("delete_data")  # 'yes' ถ้าติ๊ก checkbox
        
        try:
            data_deleted_count = 0
            log_deleted_count = 0
            
            if mode == 'all':
                # 1. ถ้าเลือกติ๊กลบข้อมูล -> ลบข้อมูลใน CancelledOrder ทั้งหมด
                if delete_data == 'yes':
                    data_deleted_count = db.session.query(CancelledOrder).delete()
                
                # 2. ลบ Log ทั้งหมดที่เป็นของ CANCEL_SYSTEM
                log_deleted_count = ImportLog.query.filter_by(platform='CANCEL_SYSTEM').delete()
                
                if data_deleted_count > 0:
                    msg = f"ล้างเกลี้ยง! (Log {log_deleted_count} รายการ, ข้อมูล {data_deleted_count} รายการ)"
                else:
                    msg = f"ล้างประวัติทั้งหมดเรียบร้อย ({log_deleted_count} รายการ)"

            else:
                # ลบตามช่วงวันที่
                d_from_str = request.form.get("date_from")
                d_to_str = request.form.get("date_to")
                
                if not d_from_str or not d_to_str:
                    flash("ระบุวันที่ไม่ถูกต้อง", "warning")
                    return redirect(url_for("import_cancel_view"))
                    
                d_from = datetime.strptime(d_from_str, "%Y-%m-%d").date()
                d_to = datetime.strptime(d_to_str, "%Y-%m-%d").date()
                
                # 1. ถ้าเลือกติ๊กลบข้อมูล -> ลบข้อมูลในช่วงวันที่
                if delete_data == 'yes':
                    # สร้าง timestamp ครอบคลุมทั้งวัน
                    dt_start = datetime.combine(d_from, datetime.min.time())
                    dt_end = datetime.combine(d_to, datetime.max.time())
                    
                    data_deleted_count = CancelledOrder.query.filter(
                        CancelledOrder.imported_at >= dt_start,
                        CancelledOrder.imported_at <= dt_end
                    ).delete(synchronize_session=False)
                
                # 2. ลบ Log ในช่วงวันที่ (เฉพาะ CANCEL_SYSTEM)
                log_deleted_count = ImportLog.query.filter(
                    ImportLog.platform == 'CANCEL_SYSTEM',
                    ImportLog.import_date >= d_from,
                    ImportLog.import_date <= d_to
                ).delete(synchronize_session=False)
                
                if data_deleted_count > 0:
                    msg = f"ล้างข้อมูลช่วง {to_be_date_str(d_from)} - {to_be_date_str(d_to)} เรียบร้อย (Log {log_deleted_count}, ข้อมูล {data_deleted_count})"
                else:
                    msg = f"ล้างประวัติช่วง {to_be_date_str(d_from)} - {to_be_date_str(d_to)} เรียบร้อย ({log_deleted_count} รายการ)"
                
            db.session.commit()
            flash(msg, "success")
            
        except Exception as e:
            db.session.rollback()
            flash(f"เกิดข้อผิดพลาด: {e}", "danger")
            
        return redirect(url_for("import_cancel_view"))
    # =========[ /NEW ]=========

    # =========[ NEW ]=========  Import Orders (จ่ายงานแล้ว)
    @app.route("/import/issued/template")
    @login_required
    def import_issued_template():
        # ใช้ logic เดียวกับ template ของ cancel (คืนไฟล์คอลัมน์ order_id)
        sample = ["ORDER-001", "ORDER-002", "ORDER-003"]
        try:
            from openpyxl import Workbook
            wb = Workbook(); ws = wb.active; ws.title = "issued_orders"; ws["A1"] = "order_id"
            for i, no in enumerate(sample, start=2): ws[f"A{i}"] = no
            bio = BytesIO(); wb.save(bio); bio.seek(0)
            return send_file(bio, as_attachment=True, download_name="template_import_orders_issued.xlsx",
                             mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        except Exception:
            csv_io = BytesIO()
            csv_io.write(("order_id\n" + "\n".join(sample)).encode("utf-8-sig"))
            csv_io.seek(0)
            return send_file(csv_io, as_attachment=True, download_name="template_import_orders_issued.csv", mimetype="text/csv")

    @app.route("/import/issued", methods=["GET", "POST"])
    @login_required
    def import_issued_orders():
        cu = current_user()
        if not cu or cu.role not in {"admin", "staff"}:
            flash("ต้องเป็นผู้ดูแลระบบหรือพนักงานเท่านั้น", "danger")
            return redirect(url_for("dashboard"))

        result = None
        if request.method == "POST":
            f = request.files.get("file")
            if not f or (f.filename or "").strip() == "":
                flash("โปรดเลือกไฟล์ Excel/CSV ก่อน", "warning")
                return redirect(url_for("import_issued_orders"))
            try:
                order_ids_raw = _parse_order_ids_from_upload(f)
                order_ids = [s.strip() for s in order_ids_raw if s and s.strip()]
                order_ids = list(dict.fromkeys(order_ids))  # unique + preserve order

                # มีอยู่จริงในระบบ?
                exists_set = {
                    r[0] for r in db.session.query(OrderLine.order_id)
                    .filter(OrderLine.order_id.in_(order_ids)).distinct().all()
                }
                not_found = [s for s in order_ids if s not in exists_set]

                # mark เป็น "จ่ายงานแล้ว" พร้อมบันทึกเวลา import
                imported_at = now_thai()
                inserted = _mark_issued(list(exists_set), user_id=cu.id, source="import", when_dt=imported_at)

                # ตาม requirement: ถ้ายังไม่เคยนับพิมพ์ ให้ตั้งเป็น 1
                if exists_set:
                    _ensure_min_print_count(list(exists_set), min_count=1, user_id=cu.id, when_iso=now_thai().isoformat())

                result = {
                    "total_in_file": len(order_ids),
                    "matched_in_system": len(exists_set),
                    "inserted_issued": inserted,
                    "not_found": not_found[:50],
                }
                flash(f"ทำเครื่องหมาย 'จ่ายงานแล้ว' {inserted} ออเดอร์", "success")

            except Exception as e:
                db.session.rollback()
                app.logger.exception("Import issued orders failed")
                flash(f"เกิดข้อผิดพลาด: {e}", "danger")
                result = None

        return render_template("import_issued.html", result=result)
    # =========[ /NEW ]=========

    # =========[ DEPRECATED: ยกเลิก - รวมไว้ใน Dashboard หลักแล้ว ]=========
    # @app.route("/dashboard/cancelled")
    # @login_required
    # def dashboard_cancelled():
    # def dashboard_cancelled():
    #     # สร้างตารางถ้ายังไม่มี (จากแพตช์ Import Orders ยกเลิก)
    #     ... (code commented out)
    #     return render_template(
    #         "dashboard_cancelled.html",
    #         rows=rows,
    #         q=q,
    #         platforms=platforms,
    #         shops=shops,
    #         platform_sel=platform_sel,
    #         shop_sel=shop_sel,
    #     )
    # =========[ /DEPRECATED ]=========

    # =========[ NEW ]=========  Dashboard: Order จ่ายแล้ว
    @app.route("/dashboard/issued")
    @login_required
    def dashboard_issued():
        if not current_user():
            return redirect(url_for("login"))

        q = (request.args.get("q") or "").strip()
        platform_sel = normalize_platform(request.args.get("platform"))
        shop_sel = request.args.get("shop_id")
        shop_sel = int(shop_sel) if shop_sel and str(shop_sel).isdigit() else None

        # Date range filter
        date_from_str = request.args.get("date_from") or ""
        date_to_str = request.args.get("date_to") or ""
        date_from_dt = None
        date_to_dt = None
        if date_from_str:
            try:
                date_from_dt = datetime.strptime(date_from_str, "%Y-%m-%d").replace(tzinfo=TH_TZ)
            except:
                pass
        if date_to_str:
            try:
                date_to_dt = datetime.strptime(date_to_str, "%Y-%m-%d").replace(hour=23, minute=59, second=59, tzinfo=TH_TZ)
            except:
                pass

        # สำหรับ dropdown เลือกแพลตฟอร์ม/ร้าน
        platforms = [p for (p,) in db.session.query(Shop.platform).filter(Shop.platform.isnot(None)).distinct().order_by(Shop.platform.asc()).all()]
        shop_query = Shop.query
        if platform_sel:
            shop_query = shop_query.filter(Shop.platform == platform_sel)
        shops = shop_query.order_by(Shop.name.asc()).all()

        # subquery map order_id -> (platform, shop_name, shop_id)
        sub = (
            db.session.query(
                OrderLine.order_id.label("oid"),
                func.min(OrderLine.shop_id).label("shop_id"),
                func.min(Shop.platform).label("platform"),
                func.min(Shop.name).label("shop_name"),
                func.min(OrderLine.logistic_type).label("logistic"),
            )
            .outerjoin(Shop, Shop.id == OrderLine.shop_id)
            .group_by(OrderLine.order_id)
            .subquery()
        )

        qry = (
            db.session.query(
                IssuedOrder.order_id,
                IssuedOrder.issued_at,
                sub.c.platform,
                sub.c.shop_name,
                sub.c.shop_id,
                sub.c.logistic,
            )
            .outerjoin(sub, sub.c.oid == IssuedOrder.order_id)
        )

        if q:
            qry = qry.filter(IssuedOrder.order_id.contains(q))
        if platform_sel:
            qry = qry.filter(sub.c.platform == platform_sel)
        if shop_sel:
            qry = qry.filter(sub.c.shop_id == shop_sel)
        if date_from_dt:
            qry = qry.filter(IssuedOrder.issued_at >= date_from_dt)
        if date_to_dt:
            qry = qry.filter(IssuedOrder.issued_at <= date_to_dt)

        rows = qry.order_by(IssuedOrder.issued_at.desc()).all()

        return render_template(
            "dashboard_issued.html",
            rows=rows, q=q, platforms=platforms, shops=shops,
            platform_sel=platform_sel, shop_sel=shop_sel,
            date_from_sel=date_from_str, date_to_sel=date_to_str
        )

    @app.post("/issued/unissue")
    @login_required
    def issued_unissue():
        cu = current_user()
        if not cu or cu.role not in {"admin", "staff"}:
            flash("ต้องเป็นผู้ดูแลระบบหรือพนักงานเท่านั้น", "danger")
            return redirect(url_for("dashboard_issued"))

        ids = request.form.getlist("order_ids[]")
        if not ids:
            oid = request.form.get("order_id")
            if oid:
                ids = [oid]
        n = _unissue(ids or [])
        if n > 0:
            flash(f"ยกเลิกจ่ายงานแล้ว {n} ออเดอร์", "success")
        else:
            flash("ไม่พบออเดอร์ที่จะยกเลิกจ่ายงาน", "warning")
        return redirect(url_for("dashboard_issued"))
    # =========[ /NEW ]=========

    # =========[ NEW ]=========  Dashboard: Order ที่ถูกลบ (Recycle Bin)
    @app.route("/dashboard/deleted")
    @login_required
    def dashboard_deleted():
        if not current_user():
            return redirect(url_for("login"))

        q = (request.args.get("q") or "").strip()
        platform_sel = normalize_platform(request.args.get("platform"))
        shop_sel = request.args.get("shop_id")
        shop_sel = int(shop_sel) if shop_sel and str(shop_sel).isdigit() else None

        # Date range filter
        date_from_str = request.args.get("date_from") or ""
        date_to_str = request.args.get("date_to") or ""
        date_from_dt = None
        date_to_dt = None
        if date_from_str:
            try:
                date_from_dt = datetime.strptime(date_from_str, "%Y-%m-%d").replace(tzinfo=TH_TZ)
            except:
                pass
        if date_to_str:
            try:
                date_to_dt = datetime.strptime(date_to_str, "%Y-%m-%d").replace(hour=23, minute=59, second=59, tzinfo=TH_TZ)
            except:
                pass

        # สำหรับ dropdown เลือกแพลตฟอร์ม/ร้าน
        platforms = [p for (p,) in db.session.query(Shop.platform).filter(Shop.platform.isnot(None)).distinct().order_by(Shop.platform.asc()).all()]
        shop_query = Shop.query
        if platform_sel:
            shop_query = shop_query.filter(Shop.platform == platform_sel)
        shops = shop_query.order_by(Shop.name.asc()).all()

        # subquery map order_id -> (platform, shop_name, shop_id, logistic)
        sub = (
            db.session.query(
                OrderLine.order_id.label("oid"),
                func.min(OrderLine.shop_id).label("shop_id"),
                func.min(Shop.platform).label("platform"),
                func.min(Shop.name).label("shop_name"),
                func.min(OrderLine.logistic_type).label("logistic"),
            )
            .outerjoin(Shop, Shop.id == OrderLine.shop_id)
            .group_by(OrderLine.order_id)
            .subquery()
        )

        qry = (
            db.session.query(
                DeletedOrder.order_id,
                DeletedOrder.deleted_at,
                sub.c.platform,
                sub.c.shop_name,
                sub.c.shop_id,
                sub.c.logistic,
                User.username.label("deleted_by")
            )
            .outerjoin(sub, sub.c.oid == DeletedOrder.order_id)
            .outerjoin(User, User.id == DeletedOrder.deleted_by_user_id)
        )

        if q:
            qry = qry.filter(DeletedOrder.order_id.contains(q))
        if platform_sel:
            qry = qry.filter(sub.c.platform == platform_sel)
        if shop_sel:
            qry = qry.filter(sub.c.shop_id == shop_sel)
        if date_from_dt:
            qry = qry.filter(DeletedOrder.deleted_at >= date_from_dt)
        if date_to_dt:
            qry = qry.filter(DeletedOrder.deleted_at <= date_to_dt)

        rows = qry.order_by(DeletedOrder.deleted_at.desc()).all()

        return render_template(
            "dashboard_deleted.html",
            rows=rows, q=q, platforms=platforms, shops=shops,
            platform_sel=platform_sel, shop_sel=shop_sel,
            date_from_sel=date_from_str, date_to_sel=date_to_str
        )

    @app.post("/deleted/restore")
    @login_required
    def deleted_restore():
        cu = current_user()
        if not cu or cu.role not in {"admin", "staff"}:
            flash("ต้องเป็นผู้ดูแลระบบหรือพนักงานเท่านั้น", "danger")
            return redirect(url_for("dashboard_deleted"))

        ids = request.form.getlist("order_ids[]")
        if not ids:
            oid = request.form.get("order_id")
            if oid:
                ids = [oid]
        
        if not ids:
            flash("ไม่พบรายการที่จะกู้คืน", "warning")
            return redirect(url_for("dashboard_deleted"))

        # ลบออกจาก DeletedOrder = กู้คืนกลับหน้าหลัก
        n = db.session.query(DeletedOrder).filter(DeletedOrder.order_id.in_(ids)).delete(synchronize_session=False)
        db.session.commit()
        
        if n > 0:
            flash(f"✅ กู้คืน {n} ออเดอร์ เรียบร้อยแล้ว", "success")
        else:
            flash("ไม่พบออเดอร์ที่จะกู้คืน", "warning")
        return redirect(url_for("dashboard_deleted"))
    # =========[ /NEW ]=========

    @app.route("/import/products", methods=["GET", "POST"])
    @login_required
    def import_products_view():
        # --- ส่วนที่ 1: ดึง URL ที่บันทึกไว้ ---
        saved_url = ""
        CONFIG_SHOP_NAME = "GoogleSheet_Products"

        try:
            config_row = db.session.execute(
                text("SELECT google_sheet_url FROM shops WHERE platform = 'PRODUCTS_SYSTEM' AND name = :name LIMIT 1"),
                {"name": CONFIG_SHOP_NAME}
            ).fetchone()
            if config_row and config_row[0]:
                saved_url = config_row[0]
        except Exception:
            db.session.rollback()

        # --- ส่วนที่ 2: จัดการนำเข้า (POST) ---
        if request.method == "POST":
            mode = request.form.get("mode")
            df = None
            source_name = "Unknown"

            try:
                # >>>> Case 1: Google Sheet
                if mode == "gsheet":
                    sheet_url = request.form.get("sheet_url")
                    if not sheet_url:
                        flash("กรุณาระบุ Google Sheet URL", "danger")
                        return redirect(url_for("import_products_view"))

                    creds = get_google_credentials()
                    client = gspread.authorize(creds)

                    try:
                        sh = client.open_by_url(sheet_url)
                        worksheet = sh.worksheet("Import_product_master")
                    except gspread.WorksheetNotFound:
                        flash("ไม่พบ Tab ชื่อ 'Import_product_master'", "danger")
                        return redirect(url_for("import_products_view"))

                    data = worksheet.get_all_records()
                    if not data:
                        flash("ไม่มีข้อมูลใน Tab", "warning")
                        return redirect(url_for("import_products_view"))

                    df = pd.DataFrame(data)
                    source_name = "Google Sheet"

                    # Auto-save URL
                    try:
                        s = Shop.query.filter_by(platform='PRODUCTS_SYSTEM', name=CONFIG_SHOP_NAME).first()
                        if not s:
                            s = Shop(platform='PRODUCTS_SYSTEM', name=CONFIG_SHOP_NAME)
                            db.session.add(s)
                            db.session.commit()
                        db.session.execute(
                            text("UPDATE shops SET google_sheet_url = :u WHERE id = :id"),
                            {"u": sheet_url, "id": s.id}
                        )
                        db.session.commit()
                    except Exception as e_save:
                        db.session.rollback()
                        app.logger.error(f"Auto-save URL failed: {e_save}")

                # >>>> Case 2: File Upload
                else:
                    f = request.files.get("file")
                    if not f:
                        flash("กรุณาเลือกไฟล์", "danger")
                        return redirect(url_for("import_products_view"))
                    df = pd.read_excel(f)
                    source_name = f.filename

                # >>>> Process Import
                if df is not None:
                    # ลบแถวว่างท้ายไฟล์ทิ้ง
                    df.dropna(how='all', inplace=True)

                    cnt = import_products(df)

                    flash(f"✅ นำเข้าสินค้าสำเร็จ {cnt} รายการ (จาก {source_name})", "success")
                    return redirect(url_for("import_products_view"))

            except Exception as e:
                db.session.rollback()
                app.logger.exception("Import Products Error")
                flash(f"เกิดข้อผิดพลาด: {e}", "danger")
                return redirect(url_for("import_products_view"))

        # --- ส่วนที่ 3: นับจำนวน SKU ทั้งหมด ---
        total_skus = 0
        try:
            total_skus = Product.query.count()
        except Exception:
            pass

        return render_template("import_products.html", saved_url=saved_url, total_skus=total_skus)

    @app.route("/import/stock", methods=["GET", "POST"])
    @login_required
    def import_stock_view():
        # --- ส่วนที่ 1: ดึง URL ที่บันทึกไว้มาแสดง (GET) ---
        saved_url = ""
        try:
            # ใช้ platform='STOCK_SYSTEM' และ name='SabuySoft' เพื่อเก็บ URL
            config_row = db.session.execute(
                text("SELECT google_sheet_url FROM shops WHERE platform = 'STOCK_SYSTEM' AND name = 'SabuySoft' LIMIT 1")
            ).fetchone()
            if config_row and config_row[0]:
                saved_url = config_row[0]
        except Exception:
            pass

        # --- ส่วนที่ 2: จัดการการนำเข้า (POST) ---
        if request.method == "POST":
            mode = request.form.get("mode")  # 'file' หรือ 'gsheet'
            
            try:
                df = None
                
                # ==== กรณีนำเข้าผ่าน Google Sheet ====
                if mode == "gsheet":
                    sheet_url = request.form.get("sheet_url")
                    if not sheet_url:
                        flash("กรุณาระบุ Google Sheet URL", "danger")
                        return redirect(url_for("import_stock_view"))
                    
                    # 1. เชื่อมต่อ Google API
                    creds = get_google_credentials()
                    client = gspread.authorize(creds)
                    
                    # 2. เปิด Sheet และ Tab
                    try:
                        sh = client.open_by_url(sheet_url)
                        worksheet = sh.worksheet("Import_sabuysoft_stock")  # ตามชื่อที่ระบุ
                    except gspread.WorksheetNotFound:
                        flash("ไม่พบ Tab ชื่อ 'Import_sabuysoft_stock'", "danger")
                        return redirect(url_for("import_stock_view"))
                    except Exception as e:
                        flash(f"เข้าถึง Google Sheet ไม่ได้: {e}", "danger")
                        return redirect(url_for("import_stock_view"))
                        
                    # 3. ดึงข้อมูลแปลงเป็น DataFrame
                    data = worksheet.get_all_records()
                    if not data:
                        flash("ไม่พบข้อมูลใน Tab Import_sabuysoft_stock", "warning")
                        return redirect(url_for("import_stock_view"))
                    
                    df = pd.DataFrame(data)
                    
                    # (Optional) บันทึก URL ล่าสุดอัตโนมัติถ้าทำรายการสำเร็จ
                    try:
                        # หา Shop หรือสร้างใหม่
                        s = Shop.query.filter_by(platform='STOCK_SYSTEM', name='SabuySoft').first()
                        if not s:
                            s = Shop(platform='STOCK_SYSTEM', name='SabuySoft')
                            db.session.add(s)
                            db.session.commit()
                        
                        # Update URL
                        db.session.execute(
                            text("UPDATE shops SET google_sheet_url = :u WHERE id = :id"),
                            {"u": sheet_url, "id": s.id}
                        )
                        db.session.commit()
                    except Exception:
                        pass  # ถ้าบันทึก URL อัตโนมัติไม่ได้ ก็ไม่เป็นไร ให้ Import ต่อไป

                # ==== กรณีนำเข้าผ่านไฟล์ Excel ====
                else:
                    f = request.files.get("file")
                    if not f:
                        flash("กรุณาเลือกไฟล์สต็อก", "danger")
                        return redirect(url_for("import_stock_view"))
                    df = pd.read_excel(f)

                # ==== ส่ง DataFrame ไปเข้าฟังก์ชัน import_stock เดิม ====
                if df is not None:
                    cnt = import_stock(df)
                    source_text = "Google Sheet" if mode == "gsheet" else "ไฟล์"
                    flash(f"✅ นำเข้าสต็อกสำเร็จ {cnt} รายการ (จาก {source_text})", "success")
                    return redirect(url_for("import_stock_view"))

            except Exception as e:
                db.session.rollback()
                app.logger.exception("Import Stock Error")
                flash(f"เกิดข้อผิดพลาดในการนำเข้าสต็อก: {e}", "danger")
                return redirect(url_for("import_stock_view"))

        return render_template("import_stock.html", saved_url=saved_url)

    @app.route("/import/sales", methods=["GET", "POST"])
    @login_required
    def import_sales_view():
        # --- ส่วนที่ 1: ดึง URL ที่บันทึกไว้ ---
        saved_url = ""
        CONFIG_SHOP_NAME = "GoogleSheet_Sales" 
        
        try:
            config_row = db.session.execute(
                text("SELECT google_sheet_url FROM shops WHERE platform = 'SALES_SYSTEM' AND name = :name LIMIT 1"),
                {"name": CONFIG_SHOP_NAME}
            ).fetchone()
            if config_row and config_row[0]:
                saved_url = config_row[0]
        except Exception:
            db.session.rollback()

        # ตัวแปร Filter วันที่
        date_from_str = request.args.get("date_from")
        date_to_str = request.args.get("date_to")
        if not date_from_str: date_from_str = now_thai().date().isoformat()
        if not date_to_str: date_to_str = now_thai().date().isoformat()
        d_from = parse_date_any(date_from_str)
        d_to = parse_date_any(date_to_str)

        # --- ส่วนที่ 2: จัดการนำเข้า (POST) ---
        if request.method == "POST":
            mode = request.form.get("mode")
            df = None
            source_name = "Unknown"
            
            try:
                # >>>> Case 1: Google Sheet
                if mode == "gsheet":
                    sheet_url = request.form.get("sheet_url")
                    if not sheet_url:
                        flash("กรุณาระบุ URL", "danger")
                        return redirect(url_for("import_sales_view"))
                    
                    creds = get_google_credentials()
                    client = gspread.authorize(creds)
                    
                    try:
                        sh = client.open_by_url(sheet_url)
                        worksheet = sh.worksheet("Import_sabuysoft_sales_10d")
                    except gspread.WorksheetNotFound:
                        flash("ไม่พบ Tab ชื่อ 'Import_sabuysoft_sales_10d'", "danger")
                        return redirect(url_for("import_sales_view"))
                    
                    data = worksheet.get_all_records()
                    if not data:
                        flash("ไม่มีข้อมูลใน Tab", "warning")
                        return redirect(url_for("import_sales_view"))
                    
                    df = pd.DataFrame(data)
                    source_name = "Google Sheet"
                    
                    # Auto-save URL
                    try:
                        s = Shop.query.filter_by(platform='SALES_SYSTEM', name=CONFIG_SHOP_NAME).first()
                        if not s:
                            s = Shop(platform='SALES_SYSTEM', name=CONFIG_SHOP_NAME)
                            db.session.add(s)
                            db.session.commit()
                        db.session.execute(
                            text("UPDATE shops SET google_sheet_url = :u WHERE id = :id"),
                            {"u": sheet_url, "id": s.id}
                        )
                        db.session.commit()
                    except Exception as e_save:
                        db.session.rollback()
                        app.logger.error(f"Auto-save URL failed: {e_save}")

                # >>>> Case 2: File Upload
                else:
                    f = request.files.get("file")
                    if not f:
                        flash("กรุณาเลือกไฟล์", "danger")
                        return redirect(url_for("import_sales_view"))
                    df = pd.read_excel(f)
                    source_name = f.filename

                # >>>> Process Import
                if df is not None:
                    # [แก้จุด A] ลบแถวว่างท้ายไฟล์ทิ้ง (Clean Empty Rows)
                    df.dropna(how='all', inplace=True)
                    
                    # นับจำนวนบรรทัดที่มีข้อมูล (Total)
                    total_rows = len(df)

                    # [แก้จุด B] เรียก Importer และรับ List ของ ID กลับมา
                    processed_ids = import_sales(df)
                    cnt = len(processed_ids)
                    
                    # คำนวณยอดไม่สำเร็จ (Failed) = ทั้งหมด - สำเร็จ
                    failed_cnt = max(0, total_rows - cnt)
                    
                    # [แก้จุด C] บันทึก Log พร้อมรายชื่อ ID (JSON) เพื่อเอาไว้นับ Unique
                    log = ImportLog(
                        import_date=now_thai().date(),
                        platform="SALES_SYSTEM",
                        shop_name="-",
                        filename=source_name,
                        added_count=cnt,
                        duplicates_count=0,
                        failed_count=failed_cnt,
                        batch_data=json.dumps({"ids": processed_ids}) # เก็บ ID ไว้ตัดซ้ำ
                    )
                    db.session.add(log)
                    db.session.commit()

                    msg = f"✅ อัปเดตข้อมูลสั่งขายสำเร็จ {cnt} รายการ"
                    if failed_cnt > 0:
                        msg += f" (ไม่สำเร็จ {failed_cnt} รายการ)"
                    
                    flash(msg, "success" if failed_cnt == 0 else "warning")
                    return redirect(url_for("import_sales_view"))

            except Exception as e:
                db.session.rollback()
                app.logger.exception("Import Sales Error")
                flash(f"เกิดข้อผิดพลาด: {e}", "danger")
                return redirect(url_for("import_sales_view"))

        # --- ส่วนที่ 3: คำนวณ Dashboard Stats (นับ Unique) ---
        total_unique_success = 0
        total_failed = 0
        
        try:
            logs = ImportLog.query.filter(
                ImportLog.import_date >= d_from,
                ImportLog.import_date <= d_to,
                ImportLog.platform == 'SALES_SYSTEM'
            ).all()
            
            # [แก้จุด D] ใช้ Set เพื่อตัด Order ID ที่ซ้ำกันออก (สำหรับการ์ด Success)
            unique_ids_set = set()
            
            for l in logs:
                total_failed += (l.failed_count or 0)
                # แกะ batch_data เพื่อเอา ID มารวมใน Set
                if l.batch_data:
                    try:
                        data = json.loads(l.batch_data)
                        if "ids" in data:
                            unique_ids_set.update(data["ids"])
                    except:
                        pass
                else:
                    # Fallback สำหรับ Log เก่าที่ไม่มี batch_data
                    pass 

            # ถ้ายอด Set มีค่า (Log ใหม่) ให้ใช้ยอดนั้น
            if len(unique_ids_set) > 0:
                total_unique_success = len(unique_ids_set)
            else:
                # ถ้าเป็น Log เก่า (ก่อนแก้โค้ด) ให้ใช้การบวกยอดเอา
                total_unique_success = sum(l.added_count for l in logs)

        except:
            total_unique_success = 0
            total_failed = 0
        
        stats = {"success": total_unique_success, "failed": total_failed}
        date_from_thai = to_be_date_str(d_from) if d_from else ""
        date_to_thai = to_be_date_str(d_to) if d_to else ""

        return render_template(
            "import_sales.html",
            stats=stats,
            saved_url=saved_url,
            date_from=date_from_str,
            date_to=date_to_str,
            date_from_thai=date_from_thai,
            date_to_thai=date_to_thai
        )

    # =========[ NEW ]========= ล้างประวัติ Sales
    @app.route("/import/sales/clear_log", methods=["POST"])
    @login_required
    def clear_sales_log():
        cu = current_user()
        if not cu or cu.role not in {"admin", "staff"}:
            flash("ไม่มีสิทธิ์ลบประวัติ", "danger")
            return redirect(url_for("import_sales_view"))
            
        mode = request.form.get("mode") # 'range' or 'all'
        delete_data = request.form.get("delete_data") # 'yes' check
        
        try:
            log_deleted = 0
            data_deleted = 0
            
            if mode == 'all':
                log_deleted = ImportLog.query.filter_by(platform='SALES_SYSTEM').delete()
                if delete_data == 'yes':
                    data_deleted = db.session.query(Sales).delete()
            else:
                d_from = datetime.strptime(request.form.get("date_from"), "%Y-%m-%d").date()
                d_to = datetime.strptime(request.form.get("date_to"), "%Y-%m-%d").date()
                
                log_deleted = ImportLog.query.filter(
                    ImportLog.platform == 'SALES_SYSTEM',
                    ImportLog.import_date >= d_from,
                    ImportLog.import_date <= d_to
                ).delete(synchronize_session=False)
                
                # Sales ไม่มี import_date ใน Model มาตรฐาน (ปกติมีแค่ created_at)
                # ถ้าต้องการลบ data แบบช่วงเวลา อาจต้องใช้ created_at แต่เพื่อความปลอดภัย
                # ถ้าเป็น range จะลบแค่ log เว้นแต่ Sales มี import_date
                if delete_data == 'yes' and hasattr(Sales, 'created_at'):
                     dt_start = datetime.combine(d_from, datetime.min.time())
                     dt_end = datetime.combine(d_to, datetime.max.time())
                     data_deleted = Sales.query.filter(
                         Sales.created_at >= dt_start,
                         Sales.created_at <= dt_end
                     ).delete(synchronize_session=False)

            db.session.commit()
            msg = f"ล้างประวัติเรียบร้อย (Log {log_deleted})"
            if data_deleted: msg += f" (Data {data_deleted})"
            flash(msg, "success")
            
        except Exception as e:
            db.session.rollback()
            flash(f"Error: {e}", "danger")
            
        return redirect(url_for("import_sales_view"))
    # =========[ /NEW ]=========

    # -----------------------
    # Accept / Cancel / Bulk
    # -----------------------
    @app.route("/accept/<int:order_line_id>", methods=["POST"])
    @login_required
    def accept_order(order_line_id):
        ol = OrderLine.query.get_or_404(order_line_id)
        # ห้ามกดรับถ้าเลข Order ถูกทำเป็นยกเลิก
        if db.session.query(CancelledOrder.id).filter_by(order_id=ol.order_id).first():
            flash(f"Order {ol.order_id} ถูกทำเป็น 'ยกเลิก' แล้ว — ไม่สามารถกดรับได้", "warning")
            return redirect(url_for("dashboard", **request.args))

        cu = current_user()
        sales_status = (getattr(ol, "sales_status", "") or "").upper()
        if sales_status == "PACKED" or bool(getattr(ol, "packed", False)):
            flash("รายการนี้ถูกแพ็คแล้ว (PACKED) — ไม่สามารถกดรับได้", "warning")
            return redirect(url_for("dashboard", **request.args))

        stock_qty = _calc_stock_qty_for_line(ol)
        if stock_qty <= 0:
            flash("สต็อกหมด — ไม่สามารถกดรับได้", "warning")
            return redirect(url_for("dashboard", **request.args))

        sku = _get_line_sku(ol)
        if not sku:
            flash("ไม่พบ SKU ของรายการนี้ — ไม่สามารถกดรับได้", "warning")
            return redirect(url_for("dashboard", **request.args))

        accepted_qty = db.session.query(func.coalesce(func.sum(OrderLine.qty), 0))\
            .filter(OrderLine.id != ol.id)\
            .filter(OrderLine.accepted.is_(True))\
            .filter(getattr(OrderLine, "sku") == sku).scalar() or 0

        proposed_total = int(accepted_qty) + int(ol.qty or 0)
        if proposed_total > int(stock_qty):
            flash("สินค้าไม่พอส่ง — ยอดที่รับจะเกินสต็อกรวมของ SKU นี้", "warning")
            return redirect(url_for("dashboard", **request.args))

        ol.accepted = True
        ol.accepted_at = now_thai()
        ol.accepted_by_user_id = cu.id if cu else None
        ol.accepted_by_username = cu.username if cu else None
        db.session.commit()
        flash(f"ทำเครื่องหมายกดรับ Order {ol.order_id} • SKU {sku} แล้ว", "success")
        return redirect(url_for("dashboard", **request.args))

    @app.route("/cancel_accept/<int:order_line_id>", methods=["POST"])
    @login_required
    def cancel_accept(order_line_id):
        ol = OrderLine.query.get_or_404(order_line_id)
        ol.accepted = False
        ol.accepted_at = None
        ol.accepted_by_user_id = None
        ol.accepted_by_username = None
        db.session.commit()
        flash(f"ยกเลิกการกดรับ Order {ol.order_id} • SKU {getattr(ol, 'sku', '')}", "warning")
        return redirect(url_for("dashboard", **request.args))

    # =========[ NEW ]========= ฟังก์ชั่นยกเลิก Order ถาวร (พร้อมเหตุผล)
    @app.post("/cancel_order_permanent")
    @login_required
    def cancel_order_permanent():
        """ยกเลิก Order ถาวร พร้อมบันทึกเหตุผล - ใช้ได้ทุกเวลา ทั้งก่อน/หลังจ่ายงาน"""
        cu = current_user()
        order_id = (request.form.get("order_id") or "").strip()
        reason = (request.form.get("reason") or "").strip()

        if not order_id:
            flash("ไม่พบเลข Order", "danger")
            return redirect(url_for("dashboard", **request.args))
        
        if not reason:
            flash("กรุณาระบุเหตุผลการยกเลิก", "warning")
            return redirect(url_for("dashboard", **request.args))

        # ตรวจสอบว่ามีอยู่แล้วหรือไม่ (อัปเดตเหตุผลใหม่ได้)
        existing = CancelledOrder.query.filter_by(order_id=order_id).first()
        if existing:
            existing.note = reason
            existing.imported_by_user_id = cu.id if cu else None
            existing.imported_at = datetime.utcnow()
            flash(f"อัปเดตข้อมูลการยกเลิก Order {order_id} แล้ว (เหตุผล: {reason})", "info")
        else:
            new_cancel = CancelledOrder(
                order_id=order_id, 
                note=reason, 
                imported_by_user_id=cu.id if cu else None,
                imported_at=datetime.utcnow()
            )
            db.session.add(new_cancel)
            flash(f"ยกเลิก Order {order_id} สำเร็จ (เหตุผล: {reason})", "success")

        db.session.commit()
        return redirect(url_for("dashboard", **request.args))
    # =========[ /NEW ]=========

    @app.route("/bulk_accept", methods=["POST"])
    @login_required
    def bulk_accept():
        cu = current_user()
        order_line_ids = request.form.getlist("order_line_ids[]")
        if not order_line_ids:
            flash("กรุณาเลือกรายการที่ต้องการกดรับ", "warning")
            return redirect(url_for("dashboard", **request.args))
        success_count = 0
        error_messages = []
        for ol_id in order_line_ids:
            try:
                ol = OrderLine.query.get(int(ol_id))
                if not ol:
                    continue
                # [NEW] block ถ้าจ่ายงานแล้ว
                if db.session.query(IssuedOrder.id).filter_by(order_id=ol.order_id).first():
                    error_messages.append(f"Order {ol.order_id} จ่ายงานแล้ว")
                    continue
                # block ถ้ายกเลิก
                if db.session.query(CancelledOrder.id).filter_by(order_id=ol.order_id).first():
                    error_messages.append(f"Order {ol.order_id} ถูกยกเลิก")
                    continue
                sales_status = (getattr(ol, "sales_status", "") or "").upper()
                if sales_status == "PACKED" or bool(getattr(ol, "packed", False)):
                    error_messages.append(f"Order {ol.order_id} ถูกแพ็คแล้ว")
                    continue
                stock_qty = _calc_stock_qty_for_line(ol)
                if stock_qty <= 0:
                    error_messages.append(f"Order {ol.order_id} สต็อกหมด")
                    continue
                # [NEW] ป้องกัน Low Stock (สินค้าน้อย <= 3 ชิ้น) ห้ามกดรับแบบกลุ่ม
                if stock_qty <= 3:
                    error_messages.append(f"Order {ol.order_id} สินค้าน้อย (Low Stock) - กรุณาตรวจสอบและกดรับรายออเดอร์")
                    continue
                sku = _get_line_sku(ol)
                if not sku:
                    error_messages.append(f"Order {ol.order_id} ไม่พบ SKU")
                    continue
                accepted_qty = db.session.query(func.coalesce(func.sum(OrderLine.qty), 0))\
                    .filter(OrderLine.id != ol.id)\
                    .filter(OrderLine.accepted.is_(True))\
                    .filter(getattr(OrderLine, "sku") == sku).scalar() or 0
                proposed_total = int(accepted_qty) + int(ol.qty or 0)
                if proposed_total > int(stock_qty):
                    error_messages.append(f"Order {ol.order_id} สินค้าไม่พอส่ง")
                    continue
                ol.accepted = True
                ol.accepted_at = now_thai()
                ol.accepted_by_user_id = cu.id if cu else None
                ol.accepted_by_username = cu.username if cu else None
                success_count += 1
            except Exception as e:
                error_messages.append(f"Order ID {ol_id}: {str(e)}")
                continue
        db.session.commit()
        if success_count > 0:
            flash(f"✅ กดรับสำเร็จ {success_count} รายการ", "success")
        if error_messages:
            for msg in error_messages[:5]:
                flash(f"⚠️ {msg}", "warning")
            if len(error_messages) > 5:
                flash(f"... และอีก {len(error_messages) - 5} รายการที่ไม่สามารถกดรับได้", "warning")
        return redirect(url_for("dashboard", **request.args))

    @app.route("/bulk_cancel", methods=["POST"])
    @login_required
    def bulk_cancel():
        order_line_ids = request.form.getlist("order_line_ids[]")
        if not order_line_ids:
            flash("กรุณาเลือกรายการที่ต้องการยกเลิก", "warning")
            return redirect(url_for("dashboard", **request.args))
        success_count = 0
        for ol_id in order_line_ids:
            try:
                ol = OrderLine.query.get(int(ol_id))
                if ol:
                    ol.accepted = False
                    ol.accepted_at = None
                    ol.accepted_by_user_id = None
                    ol.accepted_by_username = None
                    success_count += 1
            except Exception:
                continue
        db.session.commit()
        if success_count > 0:
            flash(f"✅ ยกเลิกสำเร็จ {success_count} รายการ", "success")
        return redirect(url_for("dashboard", **request.args))

    # ================== NEW: Bulk Delete Orders (เปลี่ยนเป็น Soft Delete) ==================
    @app.route("/bulk_delete_orders", methods=["POST"])
    @login_required
    def bulk_delete_orders():
        cu = current_user()
        if not cu or cu.role not in {"admin", "staff"}:
            flash("เฉพาะแอดมินหรือพนักงานเท่านั้นที่ลบได้", "danger")
            return redirect(url_for("dashboard", **request.args))

        ids = request.form.getlist("order_line_ids[]")
        if not ids:
            flash("กรุณาเลือกรายการที่ต้องการลบ", "warning")
            return redirect(url_for("dashboard", **request.args))

        # แปลง id -> set ของ order_id
        id_ints = [int(i) for i in ids if str(i).isdigit()]
        lines = OrderLine.query.filter(OrderLine.id.in_(id_ints)).all()
        oids = { (l.order_id or "").strip() for l in lines if l and l.order_id }
        if not oids:
            flash("ไม่พบเลข Order สำหรับลบ", "warning")
            return redirect(url_for("dashboard", **request.args))

        # [NEW] ย้ายไปถังขยะ (Soft Delete) แทนการลบจริง
        existing_deleted = _deleted_oids_set()
        inserted = 0
        
        for oid in oids:
            oid = (oid or "").strip()
            if not oid or oid in existing_deleted:
                # มีข้อมูลอยู่ในถังขยะแล้ว ข้ามไป
                continue
            db.session.add(DeletedOrder(
                order_id=oid,
                deleted_at=now_thai(),
                deleted_by_user_id=cu.id if cu else None
            ))
            inserted += 1
        
        db.session.commit()
        
        if inserted > 0:
            flash(f"🗑️ ย้าย {inserted} ออเดอร์ ไปที่ 'Order ที่ถูกลบ' เรียบร้อยแล้ว", "success")
        else:
            flash("ออเดอร์ที่เลือกถูกย้ายไปถังขยะแล้วก่อนหน้านี้", "info")
            
        return redirect(url_for("dashboard", **request.args))
    # ================== /NEW ==================

    # ================== NEW: Update Dispatch Round ==================
    @app.route("/update_dispatch_round", methods=["POST"])
    @login_required
    def update_dispatch_round():
        """Update dispatch_round for selected orders"""
        cu = current_user()
        if not cu:
            return jsonify({"success": False, "error": "Unauthorized"}), 401
        
        try:
            data = request.get_json()
            order_ids = data.get("order_ids", [])
            dispatch_round = data.get("dispatch_round")
            
            if not order_ids:
                return jsonify({"success": False, "error": "ไม่มีออเดอร์ที่เลือก"}), 400
            
            if dispatch_round is None or dispatch_round == "":
                return jsonify({"success": False, "error": "กรุณาระบุรอบการจ่ายงาน"}), 400
            
            # Convert to integer
            try:
                dispatch_round = int(dispatch_round)
            except (ValueError, TypeError):
                return jsonify({"success": False, "error": "รอบการจ่ายงานต้องเป็นตัวเลข"}), 400
            
            # Update all OrderLine records matching the order_ids
            updated = db.session.query(OrderLine).filter(
                OrderLine.order_id.in_(order_ids)
            ).update(
                {"dispatch_round": dispatch_round},
                synchronize_session=False
            )
            
            db.session.commit()
            
            return jsonify({
                "success": True,
                "message": f"อัปเดตรอบการจ่ายงานเป็น {dispatch_round} สำเร็จ {updated} รายการ",
                "updated": updated
            })
            
        except Exception as e:
            db.session.rollback()
            return jsonify({"success": False, "error": str(e)}), 500
    # ================== /NEW ==================

    # ================== NEW: Barcode Scan API ==================
    @app.route("/api/scan_order", methods=["POST"])
    @login_required
    def api_scan_order():
        """บันทึกการสแกนบาร์โค้ดลง Database"""
        cu = current_user()
        if not cu:
            return jsonify({"success": False, "error": "Unauthorized"}), 401
        
        try:
            data = request.get_json() or {}
            order_id = data.get("order_id")
            if not order_id:
                return jsonify({"success": False, "error": "Missing order_id"}), 400
            
            # อัปเดตเวลาที่สแกนลงในฐานข้อมูล
            tbl = _ol_table_name()
            sql = text(f"UPDATE {tbl} SET scanned_at=:now, scanned_by=:u WHERE order_id=:oid")
            db.session.execute(sql, {
                "now": now_thai().isoformat(),
                "u": cu.username,
                "oid": order_id
            })
            db.session.commit()
            
            return jsonify({"success": True})
        except Exception as e:
            db.session.rollback()
            return jsonify({"success": False, "error": str(e)}), 500

    @app.route("/api/reset_scans", methods=["POST"])
    @login_required
    def api_reset_scans():
        cu = current_user()
        if not cu:
            return jsonify({"success": False, "error": "Unauthorized"}), 401

        try:
            data = request.get_json() or {}
            order_ids = data.get("order_ids", [])
            if not order_ids:
                return jsonify({"success": False, "error": "Missing order_ids"}), 400

            tbl = _ol_table_name()
            reset_count = 0

            for order_id in order_ids:
                sql = text(f"UPDATE {tbl} SET scanned_at=NULL, scanned_by=NULL WHERE order_id=:oid")
                db.session.execute(sql, {"oid": order_id})
                reset_count += 1

            db.session.commit()

            return jsonify({"success": True, "message": f"Reset {reset_count} scans"})
        except Exception as e:
            db.session.rollback()
            return jsonify({"success": False, "error": str(e)}), 500
    # ================== /NEW ==================

    # ================== NEW: Check Order Status API (สำหรับสแกนแยกงาน) ==================
    @app.route("/api/check_order_status", methods=["POST"])
    @login_required
    def api_check_order_status():
        """เช็คสถานะ Order อย่างละเอียด - รองรับหลายสถานะพร้อมกัน (Multi-Status)"""
        cu = current_user()
        if not cu:
            return jsonify({"found": False, "message": "Unauthorized"}), 401
        
        try:
            data = request.get_json() or {}
            oid = (data.get("order_id") or "").strip()
            
            if not oid:
                return jsonify({"found": False, "message": "ไม่ระบุเลข Order"})

            # --- เริ่มเก็บสถานะ (ใช้ List) ---
            found_statuses = []
            
            # 1. เช็คสถานะหลัก (Cancelled / Issued)
            if db.session.query(CancelledOrder).filter_by(order_id=oid).first():
                found_statuses.append("CANCELLED")
            
            if db.session.query(IssuedOrder).filter_by(order_id=oid).first():
                found_statuses.append("ISSUED")

            # 2. ดึงรายการสินค้าเพื่อเช็คสถานะอื่นๆ
            lines = OrderLine.query.filter_by(order_id=oid).all()
            if not lines:
                return jsonify({"found": False, "message": f"❌ ไม่พบ Order {oid} ในระบบ"})

            # 3. เช็ค Sales Status (SBS / Packed)
            sale = Sales.query.filter_by(order_id=oid).first()
            if not sale:
                found_statuses.append("NOT_IN_SBS")
            else:
                s_status = (sale.status or "").upper()
                if "PACKED" in s_status or "แพ็คแล้ว" in s_status or "ครบตามจำนวน" in s_status:
                    found_statuses.append("PACKED")

            # 4. เช็ค Stock รายสินค้า
            stock_statuses = []
            for line in lines:
                sku = (line.sku or "").strip()
                qty = int(line.qty or 0)
                stock_qty = 0
                
                if sku:
                    prod = Product.query.filter_by(sku=sku).first()
                    if prod and hasattr(prod, "stock_qty"):
                        try: stock_qty = int(prod.stock_qty or 0)
                        except: stock_qty = 0
                    else:
                        st = Stock.query.filter_by(sku=sku).first()
                        if st and st.qty is not None:
                            stock_qty = int(st.qty)
                
                # Logic คำนวณสถานะ Stock
                if stock_qty <= 0:
                    stock_statuses.append("SHORTAGE")
                elif stock_qty < qty:
                    stock_statuses.append("NOT_ENOUGH")
                elif stock_qty - qty <= 3:
                    stock_statuses.append("LOW_STOCK")
                else:
                    stock_statuses.append("READY")

            # สรุปสถานะ Stock (เอาที่แย่ที่สุดอันเดียวพอ)
            if "SHORTAGE" in stock_statuses:
                found_statuses.append("SHORTAGE")
            elif "NOT_ENOUGH" in stock_statuses:
                found_statuses.append("NOT_ENOUGH")
            elif "LOW_STOCK" in stock_statuses:
                found_statuses.append("LOW_STOCK")
            else:
                found_statuses.append("READY")

            # --- กำหนดสีตามความรุนแรง ---
            color = "success"
            if "CANCELLED" in found_statuses or "SHORTAGE" in found_statuses or "NOT_ENOUGH" in found_statuses:
                color = "danger"
            elif "NOT_IN_SBS" in found_statuses or "LOW_STOCK" in found_statuses:
                color = "warning"
            elif "PACKED" in found_statuses:
                color = "dark"
            elif "ISSUED" in found_statuses:
                color = "info"

            # สร้างข้อความรวม (Fallback)
            msg = f"สถานะ: {', '.join(found_statuses)}"

            return jsonify({
                "found": True, 
                "statuses": found_statuses,  # ส่งกลับเป็น List
                "status": found_statuses[0] if found_statuses else "UNKNOWN",  # รองรับโค้ดเก่า
                "message": msg, 
                "color": color
            })
            
        except Exception as e:
            return jsonify({"found": False, "message": f"เกิดข้อผิดพลาด: {str(e)}"}), 500
    # ================== /NEW ==================

    # ================== NEW: Update Low Stock Round (ข้อ 1) ==================
    @app.route("/report/lowstock/update_round", methods=["POST"])
    @login_required
    def update_lowstock_round():
        """อัปเดต lowstock_round สำหรับออเดอร์ในรายงานสินค้าน้อย (ข้อ 1)"""
        cu = current_user()
        if not cu:
            return jsonify({"success": False, "message": "Unauthorized"}), 401

        data = request.get_json(silent=True) or {}
        order_ids = [str(s).strip() for s in (data.get("order_ids") or []) if str(s).strip()]
        round_raw = data.get("round")

        if not order_ids:
            return jsonify({"success": False, "message": "ไม่พบออเดอร์ในรายงานนี้"}), 400
        try:
            round_no = int(round_raw)
        except Exception:
            return jsonify({"success": False, "message": "รอบที่ต้องเป็นตัวเลข"}), 400

        # อัปเดตทุกบรรทัดของออเดอร์ที่เลือก (ใช้ raw SQL เพราะ lowstock_round ไม่มีในโมเดล)
        try:
            tbl = _ol_table_name()
            sql = text(f"""
                UPDATE {tbl}
                   SET lowstock_round = :r
                 WHERE order_id IN :oids
            """).bindparams(bindparam("oids", expanding=True))
            result = db.session.execute(sql, {"r": round_no, "oids": order_ids})
            db.session.commit()
            
            return jsonify({
                "success": True,
                "message": f"อัปเดตรอบเป็น {round_no} ให้ {result.rowcount} รายการ",
                "updated": result.rowcount
            })
        except Exception as e:
            db.session.rollback()
            return jsonify({
                "success": False,
                "message": f"เกิดข้อผิดพลาด: {str(e)}"
            }), 500
    # ================== /NEW ==================

    # -----------------------
    # Export dashboard
    # -----------------------
    @app.route("/export.xlsx")
    @login_required
    def export_excel():
        # รับค่าทั้งหมดเหมือน Dashboard
        platform = normalize_platform(request.args.get("platform"))
        shop_id = request.args.get("shop_id")
        
        import_from_str = request.args.get("import_from")
        import_to_str = request.args.get("import_to")
        
        date_from = request.args.get("date_from")
        date_to = request.args.get("date_to")
        status = request.args.get("status")
        
        q = (request.args.get("q") or "").strip()       # [NEW] รับคำค้นหา
        all_time = request.args.get("all_time")         # [NEW] รับ All Time
        mode = request.args.get("mode")                 # [NEW] รับ Mode (Today)

        # แปลงวันที่
        def _p(s): return parse_date_any(s)
        imp_from = _p(import_from_str)
        imp_to = _p(import_to_str)
        d_from = datetime.combine(_p(date_from), datetime.min.time(), tzinfo=TH_TZ) if date_from else None
        d_to = datetime.combine(_p(date_to) + timedelta(days=1), datetime.min.time(), tzinfo=TH_TZ) if date_to else None

        has_date_filter = bool(imp_from or imp_to or d_from or d_to)
        is_all_time = bool(all_time)

        # --- 1. Logic การดึงข้อมูล (เหมือน Dashboard เป๊ะ) ---
        rows = []
        base_filters = {
            "platform": platform if platform else None,
            "shop_id": int(shop_id) if shop_id else None,
        }

        if is_all_time:
            # All Time
            filters = base_filters.copy()
            filters["active_only"] = False 
            filters["all_time"] = True
            rows, _ = compute_allocation(db.session, filters)

        elif mode == 'today':
            # Order ปัจจุบัน (วันนี้) + Order ที่ยกเลิกวันนี้
            today = now_thai().date()
            
            # 1. ดึง Order ที่นำเข้าวันนี้
            filters = base_filters.copy()
            filters["active_only"] = False
            filters["import_from"] = today
            filters["import_to"] = today
            rows_import, _ = compute_allocation(db.session, filters)
            
            # 2. ดึง Order ที่ "ยกเลิกวันนี้" (บวก 7 ชม. เพื่อให้ตรงกับเวลาไทย)
            cancel_today_oids = [
                r[0] for r in db.session.query(CancelledOrder.order_id)
                .filter(func.date(CancelledOrder.imported_at, '+7 hours') == today).all()
            ]
            
            rows_cancel = []
            if cancel_today_oids:
                f_cancel = base_filters.copy()
                f_cancel["all_time"] = True
                f_cancel["active_only"] = False
                temp_rows, _ = compute_allocation(db.session, f_cancel)
                rows_cancel = [r for r in temp_rows if r.get("order_id") in cancel_today_oids]
            
            # 3. รวมรายการ (ตัดตัวซ้ำด้วย id)
            seen_ids = set()
            rows = []
            for r in (rows_import + rows_cancel):
                rid = r.get("id")
                if rid not in seen_ids:
                    rows.append(r)
                    seen_ids.add(rid)

        elif has_date_filter:
            # กรองตามวันที่
            filters = base_filters.copy()
            filters["active_only"] = False
            filters["import_from"] = imp_from
            filters["import_to"] = imp_to
            filters["date_from"] = d_from
            filters["date_to"] = d_to
            rows, _ = compute_allocation(db.session, filters)
            
        else:
            # Default View (Order ค้าง + จบงานวันนี้)
            f_active = base_filters.copy()
            f_active["active_only"] = True
            rows_active, _ = compute_allocation(db.session, f_active)
            
            today = now_thai().date()
            f_inactive = base_filters.copy()
            f_inactive["active_only"] = False
            f_inactive["import_from"] = today
            f_inactive["import_to"] = today
            
            rows_today_all, _ = compute_allocation(db.session, f_inactive)
            
            existing_ids = set(r["id"] for r in rows_active)
            rows = list(rows_active)
            for r in rows_today_all:
                if r["id"] not in existing_ids:
                    if r.get("is_packed") or r.get("is_cancelled"):
                         rows.append(r)

        # --- 2. Post-Processing Rows ---
        # [แก้ไข] ใช้ _cancelled_oids_map แทน set เพื่อดึงเหตุผล (note) มาด้วย
        cancelled_map = _cancelled_oids_map()
        packed_oids = _orders_packed_set(rows)
        orders_not_in_sbs = _orders_not_in_sbs_set(rows)
        orders_no_sales = _orders_no_sales_set(rows)
        
        # เตรียม Stock/AllQty
        totals = _build_allqty_map(rows)
        
        for r in rows:
            oid = (r.get("order_id") or "").strip()
            
            # Stock Logic
            if "stock_qty" not in r:
                sku = (r.get("sku") or "").strip()
                stock_qty = 0
                if sku:
                    prod = Product.query.filter_by(sku=sku).first()
                    if prod and hasattr(prod, "stock_qty"):
                        try: stock_qty = int(prod.stock_qty or 0)
                        except: stock_qty = 0
                    else:
                        st = Stock.query.filter_by(sku=sku).first()
                        stock_qty = int(st.qty) if st and st.qty is not None else 0
                r["stock_qty"] = stock_qty

            r["allqty"] = int(totals.get((r.get("sku") or "").strip(), r.get("qty", 0)) or 0)
            r["accepted"] = bool(r.get("accepted", False))
            r["sales_status"] = r.get("sales_status", None)
            r["logistic"] = r.get("logistic") or r.get("logistic_type") or "-"
            
            r["is_cancelled"] = False
            r["is_not_in_sbs"] = False
            r["packed"] = False
            r["cancel_reason"] = ""  # [NEW] เตรียมตัวแปรเก็บเหตุผล
            r["cancel_str"] = ""     # [NEW] ข้อความรวมสำหรับ Excel
            
            # [NEW] เช็คว่า Order นี้เคยแพ็คแล้วหรือยัง (ก่อนถูกยกเลิก)
            r["was_packed"] = (oid in packed_oids)

            # [แก้ไข] เช็คจาก map แทน set
            if oid in cancelled_map:
                r["allocation_status"] = "CANCELLED"
                r["is_cancelled"] = True
                
                # [NEW] แกะข้อมูล Note และ Time จาก dict ซ้อน
                c_info = cancelled_map[oid]
                note_txt = c_info.get('note', '')
                time_obj = c_info.get('at')
                
                # จัด Format เวลา (แปลงเป็น พ.ศ.)
                time_str = ""
                if time_obj:
                    try:
                        if time_obj.year < 2400:
                            time_obj_be = time_obj.replace(year=time_obj.year + 543)
                        else:
                            time_obj_be = time_obj
                        time_str = time_obj_be.strftime("%d/%m/%Y %H:%M")
                    except Exception:
                        pass
                
                r["cancel_reason"] = note_txt
                r["cancel_str"] = f"{note_txt} [เมื่อ: {time_str}]" if time_str else note_txt
            elif oid in packed_oids:
                r["allocation_status"] = "PACKED"
                r["packed"] = True
            else:
                if oid in orders_not_in_sbs:
                    r["is_not_in_sbs"] = True

        # --- 3. คำนวณ KPI Sets (ต้องใช้สำหรับการกรองสถานะแบบกลุ่ม) ---
        kpi_orders_ready = _orders_ready_set(rows)
        kpi_orders_low = _orders_lowstock_order_set(rows)
        
        kpi_orders_problem = set()
        for r in rows:
            # [แก้ไข] เพิ่มเงื่อนไข: ต้องยังไม่จ่ายงาน (is_issued) ด้วย ถึงจะนับเข้ากอง 3
            if not r.get("packed") and not r.get("is_cancelled") and not r.get("is_issued"):
                status_alloc = (r.get("allocation_status") or "").strip().upper()
                if status_alloc in ("SHORTAGE", "NOT_ENOUGH"):
                    oid = (r.get("order_id") or "").strip()
                    if oid:
                        kpi_orders_problem.add(oid)

        # --- 4. กรองข้อมูล (Filtering) ---
        
        # 4.1 กรองด้วย Search Q (ถ้ามี)
        if q:
            q_lower = q.lower()
            rows = [
                r for r in rows
                if q_lower in (
                    str(r.get("order_id") or "") + " " +
                    str(r.get("sku") or "") + " " +
                    str(r.get("brand") or "") + " " +
                    str(r.get("model") or "") + " " +
                    str(r.get("shop") or "") + " " +
                    str(r.get("sales_status") or "")
                ).lower()
            ]
        
        # 4.2 กรองด้วย Status
        status_norm = (status or "").strip().upper()
        if status_norm == "ORDER_CANCELLED":
            # [แก้ไข] กรองเฉพาะยกเลิกที่ยังไม่เคยแพ็ค (ก่อนแพ็ค)
            rows = [r for r in rows if r.get("is_cancelled") and not r.get("was_packed")]
        elif status_norm == "ORDER_CANCELLED_PACKED":
            # [NEW] กรองเฉพาะยกเลิกหลังแพ็ค (เคยแพ็คแล้ว)
            rows = [r for r in rows if r.get("is_cancelled") and r.get("was_packed")]
        elif status_norm == "ORDER_NOT_IN_SBS":
            rows = [r for r in rows if r.get("is_not_in_sbs")]
        elif status_norm == "ORDER_PROBLEM":
            rows = [r for r in rows if (r.get("order_id") or "").strip() in kpi_orders_problem]
        elif status_norm == "PACKED":
            rows = [r for r in rows if r.get("packed")]
        elif status_norm == "ORDER_READY":
            rows = [r for r in rows if (r.get("order_id") or "").strip() in kpi_orders_ready]
        elif status_norm in {"ORDER_LOW_STOCK", "ORDER_LOW"}:
            rows = [r for r in rows if (r.get("order_id") or "").strip() in kpi_orders_low]
        elif status_norm == "ORDER_NO_SALES":
            rows = [r for r in rows if (r.get("order_id") or "").strip() in orders_no_sales]
        elif status_norm:
            # กรองรายบรรทัด (Ready, Accepted, etc.)
            rows = [r for r in rows if (r.get("allocation_status") or "").strip().upper() == status_norm]
        else:
            # Default Table View (ซ่อน Packed/Cancelled) ยกเว้น All Time หรือ Mode Today
            if not q and not is_all_time and mode != 'today':
                 rows = [r for r in rows if not r.get("packed") and not r.get("is_cancelled")]

        # --- 5. จัดคอลัมน์ให้ตรงกับตาราง Dashboard ---
        rows = _annotate_order_spans(rows)

        data = []
        for r in rows:
            # แปลง Status เป็นภาษาไทย/คำที่เข้าใจง่าย
            st = r.get("allocation_status")
            if r.get("is_issued"): st_display = "จ่ายแล้ว"
            elif st == "READY_ACCEPT": st_display = "พร้อมรับ"
            elif st == "ACCEPTED": st_display = "รับแล้ว"
            elif st == "PACKED": st_display = "แพ็คแล้ว"
            elif st == "CANCELLED": st_display = "ยกเลิก"
            elif st == "LOW_STOCK": st_display = "สินค้าน้อย"
            elif st == "SHORTAGE": st_display = "ไม่มีสินค้า"
            elif st == "NOT_ENOUGH": st_display = "สินค้าไม่พอส่ง"
            else: st_display = st

            data.append({
                "แพลตฟอร์ม": r.get("platform"),
                "ร้าน": r.get("shop"),
                "เลข Order": r.get("order_id"),
                "SKU": r.get("sku"),
                "Brand": r.get("brand"),
                "ชื่อสินค้า": r.get("model"),
                "Stock": r.get("stock_qty"),
                "Qty": r.get("qty"),
                "AllQty": r.get("allqty"),
                "เวลาที่ลูกค้าสั่ง": r.get("order_time"),
                "กำหนดส่ง": r.get("due_date"),
                "SLA": r.get("sla"),
                "ประเภทขนส่ง": r.get("logistic"),
                "สั่งขาย": "Orderยังไม่นำเข้าSBS" if r.get("is_not_in_sbs") else r.get("sales_status"),
                "สถานะ": st_display,
                "ผู้กดรับ": r.get("accepted_by"),
                "หมายเหตุ": r.get("cancel_str") or r.get("cancel_reason")  # [แก้ไข] ใช้ cancel_str ที่มีเวลา
            })

        df = pd.DataFrame(data)

        out = BytesIO()
        with pd.ExcelWriter(out, engine="xlsxwriter") as w:
            df.to_excel(w, index=False, sheet_name="Dashboard")
            
            # จัดความกว้างคอลัมน์
            worksheet = w.sheets['Dashboard']
            worksheet.set_column('A:A', 12)  # แพลตฟอร์ม
            worksheet.set_column('B:B', 18)  # ร้าน
            worksheet.set_column('C:C', 22)  # เลข Order
            worksheet.set_column('D:D', 18)  # SKU
            worksheet.set_column('E:E', 15)  # Brand
            worksheet.set_column('F:F', 35)  # ชื่อสินค้า
            worksheet.set_column('G:G', 8)   # Stock
            worksheet.set_column('H:H', 8)   # Qty
            worksheet.set_column('I:I', 8)   # AllQty
            worksheet.set_column('J:J', 18)  # เวลาที่ลูกค้าสั่ง
            worksheet.set_column('K:K', 12)  # กำหนดส่ง
            worksheet.set_column('L:L', 18)  # SLA
            worksheet.set_column('M:M', 20)  # ประเภทขนส่ง
            worksheet.set_column('N:N', 25)  # สั่งขาย
            worksheet.set_column('O:O', 15)  # สถานะ
            worksheet.set_column('P:P', 12)  # ผู้กดรับ
            worksheet.set_column('Q:Q', 30)  # หมายเหตุ (กว้างหน่อย)
            
        out.seek(0)
        filename = f"Dashboard_Export_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        return send_file(out, as_attachment=True, download_name=filename)

    # -----------------------
    # ใบงานคลัง (Warehouse Job Sheet)
    # -----------------------
    @app.route("/report/warehouse", methods=["GET"])
    @login_required
    def print_warehouse():
        # รับค่า reset mode และ search query
        reset_mode = request.args.get("reset")
        q = (request.args.get("q") or "").strip()  # [NEW] รับค่าคำค้นหา
        
        if reset_mode == 'all':
            # ถ้ากดรีเฟรช: เคลียร์ตัวกรองทุกอย่างให้เป็น None
            platform = None
            shop_id = None
            logistic = None
            acc_from = None
            acc_to = None
            acc_from_str = ""
            acc_to_str = ""
            q = ""  # เคลียร์คำค้นหาด้วย
            round_sel = None
            print_count_sel = None
        else:
            # ถ้าไม่ได้กดรีเฟรช: รับค่าจากฟอร์มปกติ
            platform = normalize_platform(request.args.get("platform"))
            shop_id = request.args.get("shop_id")
            logistic = request.args.get("logistic")
            acc_from_str = request.args.get("accepted_from")
            acc_to_str = request.args.get("accepted_to")
            acc_from = parse_date_any(acc_from_str)
            acc_to = parse_date_any(acc_to_str)
            round_sel = request.args.get("round")
            print_count_sel = request.args.get("print_count")
        
        # [NEW] ถ้ามีคำค้นหา ให้ล้าง filter วันที่ (ค้นหาทั้งหมด)
        if q:
            acc_from = None
            acc_to = None
            acc_from_str = ""
            acc_to_str = ""

        filters = {
            "platform": platform, 
            "shop_id": int(shop_id) if shop_id else None, 
            "import_date": None,
            "accepted_from": datetime.combine(acc_from, datetime.min.time(), tzinfo=TH_TZ) if acc_from else None,
            "accepted_to": datetime.combine(acc_to + timedelta(days=1), datetime.min.time(), tzinfo=TH_TZ) if acc_to else None,
        }
        rows, _ = compute_allocation(db.session, filters)
        rows = _filter_out_cancelled_rows(rows)
        rows = [r for r in rows if r.get("accepted") and r.get("allocation_status") in ("ACCEPTED", "READY_ACCEPT")]

        # *** กรองออเดอร์ที่พิมพ์แล้วออก - แสดงเฉพาะที่ยังไม่ได้พิมพ์ ***
        # ดึง count จาก DB จริงแทนที่จะใช้ r.get("printed_warehouse") ที่เป็น 0 ตลอด
        oids = sorted({(r.get("order_id") or "").strip() for r in rows if r.get("order_id")})
        counts = _get_print_counts_local(oids, kind="warehouse")
        rows = [r for r in rows if int(counts.get((r.get("order_id") or "").strip(), 0)) == 0]

        if logistic:
            rows = [r for r in rows if (r.get("logistic") or "").lower().find(logistic.lower()) >= 0]

        # [NEW] กรอง rows ตามคำค้นหา q (ค้นหาใน order_id, sku, shop, logistic)
        if q:
            q_lower = q.lower()
            rows = [
                r for r in rows 
                if q_lower in (
                    str(r.get("order_id") or "") + 
                    str(r.get("sku") or "") + 
                    str(r.get("shop") or "") + 
                    str(r.get("logistic") or "")
                ).lower()
            ]

        _inject_print_counts_to_rows(rows, kind="warehouse")
        _inject_scan_status(rows)  # Inject scan data before grouping
        rows = _group_rows_for_warehouse_report(rows)  # Use warehouse-specific grouping

        # [NEW] กรอง Round และ Print Count หลังจากจัดกลุ่มแล้ว
        if round_sel and round_sel.strip():
            filtered_rows = []
            for r in rows:
                try:
                    if str(r.get("dispatch_round") or "") == str(round_sel):
                        filtered_rows.append(r)
                except:
                    pass
            rows = filtered_rows
        
        if print_count_sel and print_count_sel.strip():
            filtered_rows = []
            for r in rows:
                try:
                    p_val = int(r.get("printed_warehouse") or r.get("printed_count") or 0)
                    if p_val == int(print_count_sel):
                        filtered_rows.append(r)
                except:
                    pass
            rows = filtered_rows
        # [/NEW]

        total_orders = len(rows)  # Now 1 row = 1 order
        shops = Shop.query.all()
        logistics = sorted(set(r.get("logistic") for r in rows if r.get("logistic")))
        return render_template(
            "report.html",
            rows=rows,
            count_orders=total_orders,
            shops=shops,
            logistics=logistics,
            platform_sel=platform if reset_mode != 'all' else None,
            shop_sel=shop_id if reset_mode != 'all' else None,
            logistic_sel=logistic if reset_mode != 'all' else None,
            official_print=False,
            printed_meta=None,
            accepted_from=acc_from_str if reset_mode != 'all' else "",
            accepted_to=acc_to_str if reset_mode != 'all' else "",
            q=q,  # [NEW] ส่งค่าคำค้นหากลับไป template
            round_sel=round_sel if reset_mode != 'all' else None,
            print_count_sel=print_count_sel if reset_mode != 'all' else None,
        )

    @app.route("/report/warehouse/print", methods=["POST"])
    @login_required
    def print_warehouse_commit():
        cu = current_user()
        platform = normalize_platform(request.form.get("platform"))
        shop_id = request.form.get("shop_id")
        logistic = request.form.get("logistic")
        override = request.form.get("override") in ("1", "true", "yes")
        
        # Get selected order IDs from form
        selected_order_ids = request.form.get("order_ids", "")
        selected_order_ids = [oid.strip() for oid in selected_order_ids.split(",") if oid.strip()]

        filters = {"platform": platform, "shop_id": int(shop_id) if shop_id else None, "import_date": None}
        rows, _ = compute_allocation(db.session, filters)
        rows = _filter_out_cancelled_rows(rows)
        rows = [r for r in rows if r.get("accepted") and r.get("allocation_status") in ("ACCEPTED", "READY_ACCEPT")]

        if logistic:
            rows = [r for r in rows if (r.get("logistic") or "").lower().find(logistic.lower()) >= 0]

        # If specific order IDs were selected, filter to only those orders
        if selected_order_ids:
            rows = [r for r in rows if (r.get("order_id") or "").strip() in selected_order_ids]
            oids = sorted(selected_order_ids)
        else:
            oids = sorted({(r.get("order_id") or "").strip() for r in rows if r.get("order_id")})
        
        if not oids:
            flash("ไม่พบออเดอร์สำหรับพิมพ์", "warning")
            return redirect(url_for("print_warehouse", platform=platform, shop_id=shop_id, logistic=logistic))

        already = _detect_already_printed(oids, kind="warehouse")
        if already and not (override and cu and cu.role == "admin"):
            head = ", ".join(list(already)[:10])
            more = "" if len(already) <= 10 else f" ... (+{len(already)-10})"
            flash(f"มีบางออเดอร์เคยพิมพ์ใบงานคลังไปแล้ว: {head}{more}", "danger")
            flash("ถ้าจำเป็นต้องพิมพ์ซ้ำ โปรดให้แอดมินกดยืนยัน 'อนุญาตพิมพ์ซ้ำ' แล้วพิมพ์อีกครั้ง", "warning")
            return redirect(url_for("print_warehouse", platform=platform, shop_id=shop_id, logistic=logistic))

        now_iso = now_thai().isoformat()
        _mark_printed(oids, kind="warehouse", user_id=(cu.id if cu else None), when_iso=now_iso)
        
        # [แก้ไข] ปิดการจบงาน (Issued) ณ จุดนี้ เพื่อให้ Order ไปรอที่หน้า Picking ก่อน
        # _mark_issued(oids, user_id=(cu.id if cu else None), source="print:warehouse", when_dt=now_thai())
        
        db.session.commit()  # Ensure changes are committed
        db.session.expire_all()  # Force refresh to get updated print counts

        _inject_print_counts_to_rows(rows, kind="warehouse")
        _inject_scan_status(rows)  # Inject scan data to preserve in print view
        rows = _group_rows_for_warehouse_report(rows)  # Use warehouse-specific grouping

        total_orders = len(rows)  # Now 1 row = 1 order
        shops = Shop.query.all()
        logistics = sorted(set(r.get("logistic") for r in rows if r.get("logistic")))
        printed_meta = {"by": (cu.username if cu else "-"), "at": now_thai(), "orders": total_orders, "override": bool(already)}
        return render_template(
            "report.html",
            rows=rows,
            count_orders=total_orders,
            shops=shops,
            logistics=logistics,
            platform_sel=platform,
            shop_sel=shop_id,
            logistic_sel=logistic,
            official_print=True,
            printed_meta=printed_meta
        )

    # ================== NEW: View Printed Warehouse Jobs ==================
    @app.route("/report/warehouse/printed", methods=["GET"])
    @login_required
    def warehouse_printed_history():
        """ดูใบงานคลังที่พิมพ์แล้ว - สามารถเลือกวันที่และพิมพ์ซ้ำได้"""
        # รับค่า reset mode และ search query
        reset_mode = request.args.get("reset")
        q = (request.args.get("q") or "").strip()  # [NEW] รับค่าคำค้นหา
        
        # [NEW] ถ้ามีคำค้นหา ให้ข้ามการตั้งค่าวันที่ไปเลย (ค้นหาทั้งหมด)
        if q:
            target_date = None
            platform = None
            shop_id = None
            logistic = None
            print_date = None
            print_date_from = None
            print_date_to = None
            raw_from = None
            raw_to = None
            round_sel = None
            print_count_sel = None
        elif reset_mode == 'today':
            # ถ้ากดรีเฟรช: แสดงเฉพาะของวันนี้
            target_date = now_thai().date()
            platform = None
            shop_id = None
            logistic = None
            print_date = None
            print_date_from = None
            print_date_to = None
            raw_from = None
            raw_to = None
            round_sel = None
            print_count_sel = None
        else:
            # กรณีปกติ: รับค่าจากฟอร์ม
            platform = normalize_platform(request.args.get("platform"))
            shop_id = request.args.get("shop_id")
            logistic = request.args.get("logistic")
            print_date = request.args.get("print_date")  # วันที่พิมพ์ (YYYY-MM-DD) - เก็บไว้สำหรับ backward compatible
            
            # [NEW] รับค่า Date Range สำหรับวันที่พิมพ์
            print_date_from = request.args.get("print_date_from")
            print_date_to = request.args.get("print_date_to")
            
            raw_from = request.args.get("accepted_from")
            raw_to = request.args.get("accepted_to")
            round_sel = request.args.get("round")
            print_count_sel = request.args.get("print_count")
            
            # ============================================================
            # [แก้ไข] ถ้าเข้าหน้านี้ครั้งแรก (ไม่มี Params วันที่)
            # ให้ Default เป็น "วันนี้" ทันที
            # ============================================================
            if print_date_from is None and print_date_to is None and print_date is None:
                today_str = now_thai().date().isoformat()
                print_date_from = today_str
                print_date_to = today_str
            # ============================================================
            
            # ถ้าเลือกวันที่พิมพ์ (ระบบเก่า - single date)
            if print_date:
                try:
                    target_date = datetime.strptime(print_date, "%Y-%m-%d").date()
                except:
                    target_date = None
            else:
                target_date = None
        
        # ไม่ตั้งค่า default - ให้เป็นค่าว่าง (mm/dd/yyyy)
        acc_from = parse_date_any(raw_from)
        acc_to = parse_date_any(raw_to)
        
        # Get all orders that have been printed
        tbl = _ol_table_name()
        
        # Build query to get orders with print history
        if q:
            # [NEW] กรณีค้นหา: หาจากประวัติทั้งหมด (printed_warehouse > 0) ที่เลข Order ตรงกัน
            # ไม่สนวันที่พิมพ์ (Global Search in History)
            sql = text(f"""
                SELECT DISTINCT order_id 
                FROM {tbl} 
                WHERE printed_warehouse > 0 
                AND order_id LIKE :q
            """)
            result = db.session.execute(sql, {"q": f"%{q}%"}).fetchall()
        elif target_date:
            # Filter by specific print date (หรือวันนี้ถ้า reset)
            # หมายเหตุ: printed_warehouse_at ถูกบันทึกเป็นเวลาไทยอยู่แล้ว (ไม่ต้อง +7)
            sql = text(f"""
                SELECT DISTINCT order_id 
                FROM {tbl} 
                WHERE printed_warehouse > 0 
                AND DATE(printed_warehouse_at) = :target_date
            """)
            result = db.session.execute(sql, {"target_date": target_date.isoformat()}).fetchall()
        elif print_date_from or print_date_to:
            # [NEW] Filter by date range (เริ่ม - ถึง)
            # หมายเหตุ: printed_warehouse_at ถูกบันทึกเป็นเวลาไทยอยู่แล้ว (ไม่ต้อง +7)
            sql_where = "WHERE printed_warehouse > 0"
            params = {}
            if print_date_from:
                sql_where += " AND DATE(printed_warehouse_at) >= :pf"
                params["pf"] = print_date_from
            if print_date_to:
                sql_where += " AND DATE(printed_warehouse_at) <= :pt"
                params["pt"] = print_date_to
            sql = text(f"SELECT DISTINCT order_id FROM {tbl} {sql_where}")
            result = db.session.execute(sql, params).fetchall()
        else:
            # Get all printed orders
            sql = text(f"SELECT DISTINCT order_id FROM {tbl} WHERE printed_warehouse > 0")
            result = db.session.execute(sql).fetchall()
        
        printed_order_ids = [row[0] for row in result if row[0]]
        
        if not printed_order_ids:
            # No printed orders found
            shops = Shop.query.all()
            return render_template(
                "report.html",
                rows=[],
                count_orders=0,
                shops=shops,
                logistics=[],
                platform_sel=platform,
                shop_sel=shop_id,
                logistic_sel=logistic,
                official_print=False,
                printed_meta=None,
                is_history_view=True,
                print_date_sel=None if reset_mode == 'today' else print_date,
                print_date_from=print_date_from,
                print_date_to=print_date_to,
                accepted_from="" if reset_mode == 'today' else raw_from,
                accepted_to="" if reset_mode == 'today' else raw_to,
                q=q,  # [NEW] ส่งค่าคำค้นหากลับไป template
            )
        
        # Get full data for these orders
        filters = {
            "platform": platform if platform else None, 
            "shop_id": int(shop_id) if shop_id else None, 
            "import_date": None,
            "accepted_from": datetime.combine(acc_from, datetime.min.time(), tzinfo=TH_TZ) if acc_from else None,
            "accepted_to": datetime.combine(acc_to + timedelta(days=1), datetime.min.time(), tzinfo=TH_TZ) if acc_to else None,
        }
        rows, _ = compute_allocation(db.session, filters)
        rows = _filter_out_cancelled_rows(rows)
        
        # Filter to only printed orders
        rows = [r for r in rows if (r.get("order_id") or "").strip() in printed_order_ids]
        
        if logistic:
            rows = [r for r in rows if (r.get("logistic") or "").lower().find(logistic.lower()) >= 0]
        
        _inject_print_counts_to_rows(rows, kind="warehouse")
        _inject_scan_status(rows)  # Inject scan data before grouping
        rows = _group_rows_for_warehouse_report(rows)
        
        # [NEW] กรอง Round และ Print Count หลังจากจัดกลุ่มแล้ว
        if round_sel and round_sel.strip():
            filtered_rows = []
            for r in rows:
                try:
                    if str(r.get("dispatch_round") or "") == str(round_sel):
                        filtered_rows.append(r)
                except:
                    pass
            rows = filtered_rows
        
        if print_count_sel and print_count_sel.strip():
            filtered_rows = []
            for r in rows:
                try:
                    p_val = int(r.get("printed_warehouse") or r.get("printed_count") or 0)
                    if p_val == int(print_count_sel):
                        filtered_rows.append(r)
                except:
                    pass
            rows = filtered_rows
        # [/NEW]
        
        total_orders = len(rows)
        shops = Shop.query.all()
        logistics = sorted(set(r.get("logistic") for r in rows if r.get("logistic")))
        
        # Get available print dates for dropdown
        # หมายเหตุ: printed_warehouse_at ถูกบันทึกเป็นเวลาไทยอยู่แล้ว (ไม่ต้อง +7)
        sql_dates = text(f"""
            SELECT DISTINCT DATE(printed_warehouse_at) as print_date 
            FROM {tbl} 
            WHERE printed_warehouse > 0 AND printed_warehouse_at IS NOT NULL
            ORDER BY print_date DESC
        """)
        available_dates = [row[0] for row in db.session.execute(sql_dates).fetchall()]
        
        return render_template(
            "report.html",
            rows=rows,
            count_orders=total_orders,
            shops=shops,
            logistics=logistics,
            platform_sel=platform if reset_mode != 'today' else None,
            shop_sel=shop_id if reset_mode != 'today' else None,
            logistic_sel=logistic if reset_mode != 'today' else None,
            official_print=False,
            printed_meta=None,
            is_history_view=True,
            print_date_sel=None if reset_mode == 'today' else print_date,
            available_dates=available_dates,
            
            # [NEW] ส่งค่าวันที่พิมพ์กลับไปแสดงใน Input (Date Range)
            print_date_from=print_date_from,
            print_date_to=print_date_to,
            
            accepted_from="" if reset_mode == 'today' else raw_from,
            accepted_to="" if reset_mode == 'today' else raw_to,
            q=q,  # [NEW] ส่งค่าคำค้นหากลับไป template
            round_sel=round_sel if reset_mode != 'today' else None,
            print_count_sel=print_count_sel if reset_mode != 'today' else None,
        )

    # ================== NEW: Export Warehouse Excel ==================
    @app.route("/report/warehouse/export.xlsx")
    @login_required
    def export_warehouse_excel():
        """Export ใบงานคลัง (หน้าปัจจุบัน) - แสดงเฉพาะงานที่ยังไม่พิมพ์"""
        # รับค่า Filter เหมือนหน้า Warehouse
        reset_mode = request.args.get("reset")
        
        if reset_mode == 'all':
            platform = None
            shop_id = None
            logistic = None
            acc_from = None
            acc_to = None
        else:
            platform = normalize_platform(request.args.get("platform"))
            shop_id = request.args.get("shop_id")
            logistic = request.args.get("logistic")
            acc_from = parse_date_any(request.args.get("accepted_from"))
            acc_to = parse_date_any(request.args.get("accepted_to"))

        filters = {
            "platform": platform, 
            "shop_id": int(shop_id) if shop_id else None, 
            "import_date": None,
            "accepted_from": datetime.combine(acc_from, datetime.min.time(), tzinfo=TH_TZ) if acc_from else None,
            "accepted_to": datetime.combine(acc_to + timedelta(days=1), datetime.min.time(), tzinfo=TH_TZ) if acc_to else None,
        }
        
        rows, _ = compute_allocation(db.session, filters)
        rows = _filter_out_cancelled_rows(rows)
        rows = [r for r in rows if r.get("accepted") and r.get("allocation_status") in ("ACCEPTED", "READY_ACCEPT")]

        # กรองเฉพาะยังไม่พิมพ์
        oids = sorted({(r.get("order_id") or "").strip() for r in rows if r.get("order_id")})
        counts = _get_print_counts_local(oids, kind="warehouse")
        rows = [r for r in rows if int(counts.get((r.get("order_id") or "").strip(), 0)) == 0]

        if logistic:
            rows = [r for r in rows if (r.get("logistic") or "").lower().find(logistic.lower()) >= 0]

        _inject_print_counts_to_rows(rows, kind="warehouse")
        _inject_scan_status(rows)
        rows = _group_rows_for_warehouse_report(rows)

        # สร้าง DataFrame ให้ตรงกับคอลัมน์หน้าจอ
        data = []
        for r in rows:
            data.append({
                "แพลตฟอร์ม": r.get("platform", ""),
                "ร้าน": r.get("shop", ""),
                "เลข Order": r.get("order_id", ""),
                "ประเภทขนส่ง": r.get("logistic", ""),
                "ผู้กดรับ": r.get("accepted_by", ""),
                "Scan Order": "✓ แล้ว" if r.get("scanned_at") else "",
                "จ่ายงาน(รอบที่)": r.get("dispatch_round", ""),
                "พิมพ์แล้ว(ครั้ง)": r.get("printed_warehouse", 0),
                "วัน/เดือน/ปี/เวลา ที่พิมพ์": to_thai_be(r.get("printed_warehouse_at")) if r.get("printed_warehouse_at") else ""
            })

        df = pd.DataFrame(data)
        bio = BytesIO()
        with pd.ExcelWriter(bio, engine="xlsxwriter") as w:
            df.to_excel(w, index=False, sheet_name="Warehouse")
        bio.seek(0)
        
        filename = f"ใบงานคลัง_Warehouse_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        return send_file(bio, as_attachment=True, download_name=filename)

    @app.route("/report/warehouse/history/export.xlsx")
    @login_required
    def export_warehouse_history_excel():
        """Export ใบงานคลังประวัติ - แสดงงานที่พิมพ์แล้ว"""
        # รับค่า Filter เหมือนหน้า History
        reset_mode = request.args.get("reset")
        
        if reset_mode == 'today':
            target_date = now_thai().date()
            platform = None
            shop_id = None
            logistic = None
            print_date = None
            raw_from = None
            raw_to = None
        else:
            platform = normalize_platform(request.args.get("platform"))
            shop_id = request.args.get("shop_id")
            logistic = request.args.get("logistic")
            print_date = request.args.get("print_date")
            raw_from = request.args.get("accepted_from")
            raw_to = request.args.get("accepted_to")
            
            if print_date:
                try:
                    target_date = datetime.strptime(print_date, "%Y-%m-%d").date()
                except:
                    target_date = None
            else:
                target_date = None
        
        acc_from = parse_date_any(raw_from)
        acc_to = parse_date_any(raw_to)
        
        # Get printed orders
        tbl = _ol_table_name()
        
        if target_date:
            sql = text(f"""
                SELECT DISTINCT order_id 
                FROM {tbl} 
                WHERE printed_warehouse > 0 
                AND DATE(printed_warehouse_at) = :target_date
            """)
            result = db.session.execute(sql, {"target_date": target_date.isoformat()}).fetchall()
        else:
            sql = text(f"SELECT DISTINCT order_id FROM {tbl} WHERE printed_warehouse > 0")
            result = db.session.execute(sql).fetchall()
        
        printed_order_ids = [row[0] for row in result if row[0]]
        
        if not printed_order_ids:
            # Return empty Excel
            df = pd.DataFrame()
            bio = BytesIO()
            with pd.ExcelWriter(bio, engine="xlsxwriter") as w:
                df.to_excel(w, index=False, sheet_name="History")
            bio.seek(0)
            filename = f"ใบงานคลังประวัติ_History_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
            return send_file(bio, as_attachment=True, download_name=filename)
        
        # Get full data
        filters = {
            "platform": platform if platform else None, 
            "shop_id": int(shop_id) if shop_id else None, 
            "import_date": None,
            "accepted_from": datetime.combine(acc_from, datetime.min.time(), tzinfo=TH_TZ) if acc_from else None,
            "accepted_to": datetime.combine(acc_to + timedelta(days=1), datetime.min.time(), tzinfo=TH_TZ) if acc_to else None,
        }
        rows, _ = compute_allocation(db.session, filters)
        rows = _filter_out_cancelled_rows(rows)
        rows = [r for r in rows if (r.get("order_id") or "").strip() in printed_order_ids]
        
        if logistic:
            rows = [r for r in rows if (r.get("logistic") or "").lower().find(logistic.lower()) >= 0]
        
        _inject_print_counts_to_rows(rows, kind="warehouse")
        _inject_scan_status(rows)
        rows = _group_rows_for_warehouse_report(rows)
        
        # สร้าง DataFrame
        data = []
        for r in rows:
            data.append({
                "แพลตฟอร์ม": r.get("platform", ""),
                "ร้าน": r.get("shop", ""),
                "เลข Order": r.get("order_id", ""),
                "ประเภทขนส่ง": r.get("logistic", ""),
                "ผู้กดรับ": r.get("accepted_by", ""),
                "Scan Order": "✓ แล้ว" if r.get("scanned_at") else "",
                "จ่ายงาน(รอบที่)": r.get("dispatch_round", ""),
                "พิมพ์แล้ว(ครั้ง)": r.get("printed_warehouse", 0),
                "วัน/เดือน/ปี/เวลา ที่พิมพ์": to_thai_be(r.get("printed_warehouse_at")) if r.get("printed_warehouse_at") else ""
            })

        df = pd.DataFrame(data)
        bio = BytesIO()
        with pd.ExcelWriter(bio, engine="xlsxwriter") as w:
            df.to_excel(w, index=False, sheet_name="History")
        bio.seek(0)
        
        filename = f"ใบงานคลังประวัติ_History_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        return send_file(bio, as_attachment=True, download_name=filename)

    # ================== NEW: Low-Stock & No-Stock Reports ==================

    @app.route("/report/lowstock", methods=["GET"])
    @login_required
    def report_lowstock():
        """
        รายงานสินค้าน้อย — อ้างอิงชุด SKU/Order จาก Dashboard โดยตรง
        ข้อสำคัญตาม requirement:
          - ไม่ดึงออเดอร์ที่ PACKED แล้ว (ข้อ 1)
          - 'จ่ายงาน(รอบที่)' ใช้คอลัมน์ lowstock_round แยกจาก dispatch_round (ข้อ 2)
          - 'พิมพ์แล้ว(ครั้ง)' ใช้ printed_lowstock (ข้อ 3)
          - รองรับ filter ครบ (ข้อ 4)
          - รองรับ sort ทุกคอลัมน์ (ข้อ 5)
          - ดึงเฉพาะชุด Order สินค้าน้อยจาก Dashboard (ข้อ 6)
        """
        # ไม่ต้องใช้ services.lowstock_queue แล้ว - ใช้ compute_allocation โดยตรง

        # ---- รับตัวกรอง/เรียง ----
        platform = normalize_platform(request.args.get("platform"))
        shop_id  = request.args.get("shop_id")
        logistic = request.args.get("logistic")
        round_num = request.args.get("round")  # ข้อ 7: กรองรอบ
        q        = (request.args.get("q") or "").strip()
        sort_col = (request.args.get("sort") or "").strip().lower()
        sort_dir = (request.args.get("dir") or "asc").lower()
        
        # รับค่าวันที่กรอง
        date_from_str = request.args.get("date_from")
        date_to_str = request.args.get("date_to")
        import_from_str = request.args.get("import_from")
        import_to_str = request.args.get("import_to")

        shops = Shop.query.order_by(Shop.name.asc()).all()

        # ---- 1) ดึง allocation rows เหมือน Dashboard ----
        filters = {
            "platform": platform if platform else None,
            "shop_id": int(shop_id) if shop_id else None,
            "import_date": None
        }
        rows, _ = compute_allocation(db.session, filters)
        rows = _filter_out_cancelled_rows(rows)
        rows = _filter_out_issued_rows(rows)
        rows = _filter_out_lowstock_printed_rows(rows)  # <<<< NEW (ข้อ 2): ตัดออเดอร์ที่พิมพ์รายงานสินค้าน้อยออก

        # คำนวณออเดอร์ที่แพ็คแล้ว (เช็คจาก sales_status)
        packed_oids = _orders_packed_set(rows)

        # เติม stock_qty / logistic ให้ครบ + ไม่เอา PACKED (ข้อ 1)
        safe = []
        for r in rows:
            r = dict(r)
            # กรองออเดอร์ที่อยู่ในลิสต์แพ็คแล้วออก
            if (r.get("order_id") or "").strip() in packed_oids:
                continue
            sales_status = (str(r.get("sales_status") or "")).upper()
            if sales_status == "PACKED" or bool(r.get("packed", False)):
                continue
            if "stock_qty" not in r:
                sku = (r.get("sku") or "").strip()
                stock_qty = 0
                if sku:
                    prod = Product.query.filter_by(sku=sku).first()
                    if prod and hasattr(prod, "stock_qty"):
                        try:
                            stock_qty = int(prod.stock_qty or 0)
                        except Exception:
                            stock_qty = 0
                    else:
                        st = Stock.query.filter_by(sku=sku).first()
                        stock_qty = int(st.qty) if st and st.qty is not None else 0
                r["stock_qty"] = stock_qty
            r["logistic"] = r.get("logistic") or r.get("logistic_type") or "-"
            # ไม่ต้อง _recompute เพราะ allocation_status มาจาก compute_allocation แล้ว
            safe.append(r)

        # ---- 2) ให้ "Order สินค้าน้อย" เป็นตัวตั้ง (ข้อ 6) ----
        orders_low = _orders_lowstock_order_set(safe)
        safe = [r for r in safe if (r.get("order_id") or "").strip() in orders_low]

        # ---- 2.5) กรองตามวันที่สั่งซื้อและวันที่นำเข้า ----
        if date_from_str or date_to_str:
            from datetime import datetime
            def parse_date_str(s):
                if not s: return None
                try: return datetime.strptime(s, "%Y-%m-%d").date()
                except: return None
            date_from = parse_date_str(date_from_str)
            date_to = parse_date_str(date_to_str)
            filtered = []
            for r in safe:
                order_dt = r.get("order_time")
                if isinstance(order_dt, str):
                    try: order_dt = datetime.strptime(order_dt.split()[0], "%Y-%m-%d").date()
                    except: order_dt = None
                elif isinstance(order_dt, datetime):
                    order_dt = order_dt.date()
                if order_dt:
                    if date_from and order_dt < date_from: continue
                    if date_to and order_dt > date_to: continue
                elif date_from or date_to:
                    continue
                filtered.append(r)
            safe = filtered
        
        if import_from_str or import_to_str:
            from datetime import datetime
            def parse_date_str(s):
                if not s: return None
                try: return datetime.strptime(s, "%Y-%m-%d").date()
                except: return None
            import_from = parse_date_str(import_from_str)
            import_to = parse_date_str(import_to_str)
            filtered = []
            for r in safe:
                imp_dt = r.get("import_date")
                if isinstance(imp_dt, str):
                    try: imp_dt = datetime.strptime(imp_dt, "%Y-%m-%d").date()
                    except: imp_dt = None
                elif isinstance(imp_dt, datetime):
                    imp_dt = imp_dt.date()
                elif isinstance(imp_dt, date):
                    pass
                else:
                    imp_dt = None
                if imp_dt:
                    if import_from and imp_dt < import_from: continue
                    if import_to and imp_dt > import_to: continue
                elif import_from or import_to:
                    continue
                filtered.append(r)
            safe = filtered

        # ---- 3) กรองเฉพาะ allocation_status == "LOW_STOCK" ตาม compute_allocation ----
        # ใช้ allocation_status จาก compute_allocation โดยตรง (Single Source of Truth)
        lines = [r for r in safe if r.get("allocation_status") == "LOW_STOCK"]

        # ---- 4) กรองเพิ่มตามคำค้น/โลจิสติกส์ (ข้อ 4) ----
        if logistic:
            lines = [r for r in lines if (r.get("logistic") or "").lower().find(logistic.lower()) >= 0]
        if q:
            ql = q.lower()
            def _hit(s):
                return ql in (str(s or "").lower())
            lines = [r for r in lines if (
                _hit(r.get("order_id")) or _hit(r.get("sku")) or _hit(r.get("brand")) or
                _hit(r.get("model")) or _hit(r.get("shop")) or _hit(r.get("platform")) or _hit(r.get("logistic"))
            )]

        # ---- NEW (ข้อ 1): อ่านค่า lowstock_round จาก DB เผื่อ compute_allocation ไม่ส่งฟิลด์มา ----
        order_ids_for_round = sorted({(r.get("order_id") or "").strip() for r in lines if r.get("order_id")})
        low_round_by_oid = {}
        if order_ids_for_round:
            # ใช้ raw SQL แทน ORM เพราะ lowstock_round ไม่มีในโมเดล
            tbl = _ol_table_name()
            sql = text(f"""
                SELECT order_id, MAX(lowstock_round) AS r
                  FROM {tbl}
                 WHERE order_id IN :oids
                 GROUP BY order_id
            """).bindparams(bindparam("oids", expanding=True))
            try:
                q_round = db.session.execute(sql, {"oids": order_ids_for_round}).all()
                low_round_by_oid = {str(r[0]): (int(r[1]) if r[1] is not None else None) for r in q_round}
            except Exception:
                # ถ้าคอลัมน์ยังไม่มี ให้ใช้ค่าว่าง
                low_round_by_oid = {}

        # ---- เตรียมข้อมูล Mixed Status ----
        status_map = {
            "READY_ACCEPT": "พร้อมรับ",
            "SHORTAGE": "ไม่มีของ",
            "NOT_ENOUGH": "ไม่พอส่ง",
            "ACCEPTED": "รับแล้ว",
            "PACKED": "แพ็คแล้ว",
            "CANCELLED": "ยกเลิก",
            "ISSUED": "จ่ายงานแล้ว"
        }
        mixed_info = {}
        for r in lines:
            oid = (r.get("order_id") or "").strip()
            if oid and oid not in mixed_info:
                other_rows = [x for x in safe if (x.get("order_id") or "").strip() == oid]
                details = []
                for x in other_rows:
                    s = x.get("allocation_status")
                    if s and s != "LOW_STOCK":
                        readable_status = status_map.get(s, s)
                        product_name = x.get("model") or x.get("sku") or "?"
                        details.append(f"{readable_status} ({product_name})")
                if details:
                    mixed_info[oid] = f"มีรายการอื่น: {', '.join(details)}"
                else:
                    mixed_info[oid] = ""

        # ---- 5) แปลงเป็นคอลัมน์ของรายงาน + AllQty ----
        out = []
        for r in lines:
            oid = (r.get("order_id") or "").strip()
            out.append({
                "platform":      r.get("platform"),
                "store":         r.get("shop"),
                "order_no":      oid,
                "sku":           r.get("sku"),
                "brand":         r.get("brand"),
                "product_name":  r.get("model"),
                "stock":         int(r.get("stock_qty", 0) or 0),
                "qty":           int(r.get("qty", 0) or 0),
                "order_time":    r.get("order_time"),
                "due_date":      r.get("due_date"),
                "sla":           r.get("sla"),
                "shipping_type": r.get("logistic"),
                "assign_round":  low_round_by_oid.get(oid, r.get("lowstock_round")),  # <<<< ใช้ค่าจาก DB (ข้อ 1)
                "printed_count": 0,
                "note":          mixed_info.get(oid, ""),  # เพิ่มหมายเหตุ
            })
        from collections import defaultdict
        sum_by_sku = defaultdict(int)
        for r in out:
            sum_by_sku[(r["sku"] or "").strip()] += int(r["qty"] or 0)
        for r in out:
            r["allqty"] = sum_by_sku[(r["sku"] or "").strip()]

        # ---- 6) เรียงลำดับ (ข้อ 5) ----
        sort_col = sort_col if sort_col in {"platform","store","order_no","sku","brand","product_name","stock","qty","allqty","order_time","due_date","sla","shipping_type","assign_round","printed_count"} else "order_no"
        rev = (sort_dir == "desc")
        def _key(v):
            if sort_col in {"stock","qty","allqty","assign_round","printed_count"}:
                try: return int(v.get(sort_col) or 0)
                except: return 0
            elif sort_col in {"order_time","due_date"}:
                try: return datetime.fromisoformat(str(v.get(sort_col)))
                except: return str(v.get(sort_col) or "")
            else:
                return str(v.get(sort_col) or "")
        out.sort(key=_key, reverse=rev)

        # ---- 7) นับ "พิมพ์แล้ว(ครั้ง)" (ข้อ 3) ----
        order_ids = sorted({(r["order_no"] or "").strip() for r in out if r.get("order_no")})
        counts_low = _get_print_counts_local(order_ids, "lowstock")
        for r in out:
            oid = (r.get("order_no") or "").strip()
            r["printed_count"] = int(counts_low.get(oid, 0))

        # ---- 8) เตรียม context สำหรับ template ----
        # คำนวณจำนวน SKU ที่ไม่ซ้ำจาก out
        low_skus = {(r.get("sku") or "").strip() for r in out if r.get("sku")}
        summary = {"sku_count": len(low_skus), "orders_count": len(order_ids)}
        # ข้อ 1: ไม่ต้องแสดงเวลาพิมพ์ในหน้าปกติ (ยังไม่ได้พิมพ์จริง)
        for r in out:
            r["printed_at"] = None  # ไม่ใส่เวลา

        logistics = sorted(set([r.get("shipping_type") for r in out if r.get("shipping_type")]))
        
        # ข้อ 7: หา available rounds สำหรับ dropdown
        available_rounds = sorted({r["assign_round"] for r in out if r["assign_round"] is not None})
        if not available_rounds:
            rs = db.session.execute(text("SELECT DISTINCT lowstock_round FROM order_lines WHERE lowstock_round IS NOT NULL ORDER BY lowstock_round")).fetchall()
            available_rounds = [x[0] for x in rs]

        # [SCAN] ดึงข้อมูลการ Scan Order เพื่อส่งไปหน้าเว็บ
        if order_ids:
            tbl = _ol_table_name()
            sql_scan = text(f"SELECT order_id, MAX(scanned_at) FROM {tbl} WHERE order_id IN :oids GROUP BY order_id")
            sql_scan = sql_scan.bindparams(bindparam("oids", expanding=True))
            res_scan = db.session.execute(sql_scan, {"oids": order_ids}).fetchall()
            scan_map = {str(r[0]): r[1] for r in res_scan if r[0]}
            for r in out:
                oid = (r.get("order_no") or "").strip()
                r["scanned_at"] = scan_map.get(oid)

        return render_template(
            "report_lowstock.html",
            rows=out,
            summary=summary,
            printed_at=None,  # ข้อ 1: ไม่แสดงเวลาพิมพ์ในหน้าปกติ
            order_ids=order_ids,
            shops=shops,
            logistics=logistics,
            platform_sel=platform,
            shop_sel=shop_id,
            logistic_sel=logistic,
            round_sel=round_num,
            available_rounds=available_rounds,
            sort_col=sort_col,
            sort_dir=("desc" if rev else "asc"),
            q=q,
            date_from=date_from_str,
            date_to=date_to_str,
            import_from=import_from_str,
            import_to=import_to_str,
            mixed_status=mixed_info,
            is_history_view=False
        )

    @app.post("/report/lowstock/print")
    @login_required
    def report_lowstock_print():
        """บันทึกการพิมพ์รายงานสินค้าน้อย + ย้ายไปหน้าประวัติ (ข้อ 7)"""
        cu = current_user()
        order_ids_raw = (request.form.get("order_ids") or "").strip()
        order_ids = [s.strip() for s in order_ids_raw.split(",") if s.strip()]
        if not order_ids:
            flash("ไม่พบออเดอร์สำหรับพิมพ์", "warning")
            return redirect(url_for("report_lowstock"))

        now_iso = now_thai().isoformat()
        
        # 1. บันทึกว่าพิมพ์ Low Stock แล้ว
        _mark_lowstock_printed(order_ids, username=(cu.username if cu else None), when_iso=now_iso)
        
        # 2. ย้ายไป "Order จ่ายแล้ว" (Issued) ทันที
        _mark_issued(order_ids, user_id=(cu.id if cu else None), source="print:lowstock", when_dt=now_thai())
        
        db.session.commit()
        return redirect(url_for("report_lowstock_printed", auto_print="1"))

    @app.get("/report/lowstock/printed")
    @login_required
    def report_lowstock_printed():
        """ประวัติรายงานสินค้าน้อยที่พิมพ์แล้ว (ข้อ 7)"""
        # ไม่ต้องใช้ services.lowstock_queue แล้ว - ใช้ compute_allocation โดยตรง
        
        platform = normalize_platform(request.args.get("platform"))
        shop_id  = request.args.get("shop_id")
        logistic = request.args.get("logistic")
        q        = (request.args.get("q") or "").strip()
        round_num = request.args.get("round")
        sort_col = (request.args.get("sort") or "order_no").strip().lower()
        sort_dir = (request.args.get("dir") or "asc").lower()
        
        # รับค่าตัวกรองวันที่สั่งซื้อและนำเข้า
        date_from_str = request.args.get("date_from")
        date_to_str = request.args.get("date_to")
        import_from_str = request.args.get("import_from")
        import_to_str = request.args.get("import_to")
        
        # รับค่าช่วงวันที่พิมพ์
        print_date_from = request.args.get("print_date_from")
        print_date_to = request.args.get("print_date_to")
        reset_mode = request.args.get("reset")  # [NEW] รับค่า reset
        action = request.args.get("action")  # [NEW] รับค่า action (เพื่อแยกการกดปุ่มกรอง กับการเข้าหน้าเว็บครั้งแรก)
        
        # [SMART DEFAULT] ถ้าไม่มีวันที่ส่งมา AND ไม่มีคำค้นหา AND ไม่ได้ reset AND ไม่ใช่การกดปุ่มกรอง -> ให้กรอง "วันนี้"
        if not action and reset_mode != 'all' and not print_date_from and not print_date_to and not q:
            # เข้าหน้าเว็บครั้งแรก (ไม่มี action) = ดูงานวันนี้
            today = now_thai().date().isoformat()
            print_date_from = today
            print_date_to = today
        # ถ้ามี action (กดปุ่มกรอง) หรือ q หรือ reset='all' แต่ไม่มีวันที่ -> ค้นหาทั้งหมด

        tbl = _ol_table_name()
        
        # ========================================================
        # [FIX] ดึงข้อมูลเฉพาะเมื่อ: มีคำค้นหา หรือ มีการเลือกวันที่
        # ========================================================
        if q:
            # กรณี 1: มีคำค้นหา -> ค้นหาทั้งหมด (Global Search)
            sql = text(f"SELECT DISTINCT order_id FROM {tbl} WHERE printed_lowstock > 0")
            result = db.session.execute(sql).fetchall()
            printed_oids = [r[0] for r in result if r and r[0]]
        elif print_date_from or print_date_to:
            # กรณี 2: มีการเลือกวันที่ -> กรองตามวันที่
            sql_where = "printed_lowstock > 0"
            params = {}
            if print_date_from:
                sql_where += " AND DATE(printed_lowstock_at) >= :pf"
                params["pf"] = print_date_from
            if print_date_to:
                sql_where += " AND DATE(printed_lowstock_at) <= :pt"
                params["pt"] = print_date_to
            sql = text(f"SELECT DISTINCT order_id FROM {tbl} WHERE {sql_where}")
            result = db.session.execute(sql, params).fetchall()
            printed_oids = [r[0] for r in result if r and r[0]]
        else:
            # กรณี 3: ไม่ค้นหา และ ไม่เลือกวัน (เช่น กด reset='all') -> ไม่แสดงอะไร
            printed_oids = []

        def _available_dates():
            sql = text(f"SELECT DISTINCT DATE(printed_lowstock_at) as d FROM {tbl} WHERE printed_lowstock > 0 AND printed_lowstock_at IS NOT NULL ORDER BY d DESC")
            return [r[0] for r in db.session.execute(sql).fetchall()]

        shops = Shop.query.order_by(Shop.name.asc()).all()
        
        if not printed_oids:
            return render_template(
                "report_lowstock.html",
                rows=[],
                summary={"sku_count": 0, "orders_count": 0},
                printed_at=None,
                order_ids=[],
                shops=shops,
                logistics=[],
                platform_sel=platform,
                shop_sel=shop_id,
                logistic_sel=logistic,
                is_history_view=True,
                available_dates=_available_dates(),
                print_date_from=print_date_from,
                print_date_to=print_date_to,
                sort_col=sort_col,
                sort_dir=sort_dir,
                q=q,
                round_sel=round_num,
                date_from=date_from_str,
                date_to=date_to_str,
                import_from=import_from_str,
                import_to=import_to_str
            )

        # เตรียมตัวกรองวันที่สั่งซื้อ
        date_from_dt = None
        date_to_dt = None
        if date_from_str:
            try:
                date_from_dt = datetime.combine(parse_date_any(date_from_str), datetime.min.time(), tzinfo=TH_TZ)
            except: pass
        if date_to_str:
            try:
                date_to_dt = datetime.combine(parse_date_any(date_to_str) + timedelta(days=1), datetime.min.time(), tzinfo=TH_TZ)
            except: pass

        filters = {
            "platform": platform if platform else None,
            "shop_id": int(shop_id) if shop_id else None,
            "import_date": None,
            "date_from": date_from_dt,
            "date_to": date_to_dt
        }
        rows, _ = compute_allocation(db.session, filters)
        rows = _filter_out_cancelled_rows(rows)
        rows = [r for r in rows if (r.get("order_id") or "").strip() in printed_oids]
        
        # กรองวันที่นำเข้า (Import Date)
        if import_from_str or import_to_str:
            imp_from = parse_date_any(import_from_str) if import_from_str else None
            imp_to = parse_date_any(import_to_str) if import_to_str else None
            filtered_rows = []
            for r in rows:
                d = r.get("import_date")
                if isinstance(d, str):
                    try:
                        d = datetime.strptime(d, "%Y-%m-%d").date()
                    except:
                        d = None
                elif isinstance(d, datetime):
                    d = d.date()
                
                if d:
                    if imp_from and d < imp_from:
                        continue
                    if imp_to and d > imp_to:
                        continue
                elif imp_from or imp_to:
                    continue
                filtered_rows.append(r)
            rows = filtered_rows

        safe = []
        for r in rows:
            r = dict(r)
            r["logistic"] = r.get("logistic") or r.get("logistic_type") or "-"
            if "stock_qty" not in r:
                sku = (r.get("sku") or "").strip()
                stock_qty = 0
                if sku:
                    prod = Product.query.filter_by(sku=sku).first()
                    if prod and hasattr(prod, "stock_qty"):
                        try: stock_qty = int(prod.stock_qty or 0)
                        except Exception: stock_qty = 0
                    else:
                        st = Stock.query.filter_by(sku=sku).first()
                        stock_qty = int(st.qty) if st and st.qty is not None else 0
                r["stock_qty"] = stock_qty
            # ไม่ต้อง _recompute_allocation_row(r) เพราะ compute_allocation คำนวณให้แล้ว
            safe.append(r)

        # [CRITICAL FIX] Logic กรองสำหรับหน้าประวัติ
        # เพราะออเดอร์ถูก mark เป็น ISSUED แล้ว allocation_status อาจไม่ใช่ LOW_STOCK
        # ต้อง fallback เช็ค stock condition แทน
        def _is_low_for_history(r):
            # 1. ถ้า status เป็น LOW_STOCK อยู่แล้ว -> เอา
            if r.get("allocation_status") == "LOW_STOCK": return True
            # 2. Fallback: ถ้า stock <= 3 (เกณฑ์ Low Stock มาตรฐาน) -> เอา
            try: s = int(r.get("stock_qty") or 0)
            except: s = 0
            if s <= 3: return True
            return False

        low_skus = {(r.get("sku") or "").strip() for r in safe if _is_low_for_history(r)}
        lines = [r for r in safe if (r.get("sku") or "").strip() in low_skus]

        # เตรียมข้อมูล Mixed Status สำหรับหน้าประวัติ
        status_map = {
            "READY_ACCEPT": "พร้อมรับ",
            "SHORTAGE": "ไม่มีของ",
            "NOT_ENOUGH": "ไม่พอส่ง",
            "ACCEPTED": "รับแล้ว",
            "PACKED": "แพ็คแล้ว",
            "CANCELLED": "ยกเลิก",
            "ISSUED": "จ่ายงานแล้ว"
        }
        mixed_info = {}
        for r in lines:
            oid = (r.get("order_id") or "").strip()
            if oid and oid not in mixed_info:
                other_rows = [x for x in safe if (x.get("order_id") or "").strip() == oid]
                details = []
                for x in other_rows:
                    s = x.get("allocation_status")
                    if s and s != "LOW_STOCK":
                        readable_status = status_map.get(s, s)
                        product_name = x.get("model") or x.get("sku") or "?"
                        details.append(f"{readable_status} ({product_name})")
                if details:
                    mixed_info[oid] = f"มีรายการอื่น: {', '.join(details)}"
                else:
                    mixed_info[oid] = ""

        if logistic:
            lines = [r for r in lines if (r.get("logistic") or "").lower().find(logistic.lower()) >= 0]

        # กรองตามคำค้นหา (q)
        if q:
            q_lower = q.lower()
            lines = [
                r for r in lines
                if q_lower in (
                    str(r.get("order_id") or "") +
                    str(r.get("sku") or "") +
                    str(r.get("brand") or "") +
                    str(r.get("model") or "") +
                    str(r.get("shop") or "") +
                    str(r.get("platform") or "") +
                    str(r.get("logistic") or "")
                ).lower()
            ]

        out = []
        for r in lines:
            oid = (r.get("order_id") or "").strip()  # [FIX] เพิ่มการดึงค่า oid ในแต่ละรอบ
            out.append({
                "platform":      r.get("platform"),
                "store":         r.get("shop"),
                "order_no":      oid,
                "sku":           r.get("sku"),
                "brand":         r.get("brand"),
                "product_name":  r.get("model"),
                "stock":         int(r.get("stock_qty", 0) or 0),
                "qty":           int(r.get("qty", 0) or 0),
                "order_time":    r.get("order_time"),
                "due_date":      r.get("due_date"),
                "sla":           r.get("sla"),
                "shipping_type": r.get("logistic"),
                "assign_round":  r.get("lowstock_round"),
                "printed_count": 0,
                "note":          mixed_info.get(oid, ""),  # เพิ่มหมายเหตุ
            })
        from collections import defaultdict
        sum_by_sku = defaultdict(int)
        for r in out:
            sum_by_sku[(r["sku"] or "").strip()] += int(r["qty"] or 0)
        for r in out:
            r["allqty"] = sum_by_sku[(r["sku"] or "").strip()]

        # เรียง
        sort_col = sort_col if sort_col in {"platform","store","order_no","sku","brand","product_name","stock","qty","allqty","order_time","due_date","sla","shipping_type","assign_round","printed_count"} else "order_no"
        rev = (sort_dir == "desc")
        def _key(v):
            if sort_col in {"stock","qty","allqty","assign_round","printed_count"}:
                try: return int(v.get(sort_col) or 0)
                except: return 0
            elif sort_col in {"order_time","due_date"}:
                try: return datetime.fromisoformat(str(v.get(sort_col)))
                except: return str(v.get(sort_col) or "")
            else:
                return str(v.get(sort_col) or "")
        out.sort(key=_key, reverse=rev)

        order_ids = sorted({(r["order_no"] or "").strip() for r in out if r.get("order_no")})
        counts_low = _get_print_counts_local(order_ids, "lowstock")
        for r in out:
            oid = (r.get("order_no") or "").strip()
            r["printed_count"] = int(counts_low.get(oid, 0))

        # ข้อ 1: ดึงเวลา printed_lowstock_at ต่อ order_id จาก DB
        tbl = _ol_table_name()
        sql_ts = text(f"""
            SELECT order_id, MAX(printed_lowstock_at) AS ts
            FROM {tbl}
            WHERE order_id IN :oids AND printed_lowstock_at IS NOT NULL
            GROUP BY order_id
        """).bindparams(bindparam("oids", expanding=True))
        rows_ts = db.session.execute(sql_ts, {"oids": order_ids}).all()
        ts_map = {}
        for oid, ts in rows_ts:
            if not ts:
                continue
            try:
                dt = datetime.fromisoformat(ts)
                if dt.tzinfo is None:
                    dt = TH_TZ.localize(dt)
                ts_map[str(oid)] = dt
            except Exception:
                pass

        # ใส่ลงในแต่ละแถว
        for r in out:
            r["printed_at"] = ts_map.get((r.get("order_no") or "").strip())

        # เวลาพิมพ์บนหัวรายงาน (ล่าสุดสุดในชุด)
        meta_printed_at = max(ts_map.values()) if ts_map else None

        # ดึงค่า lowstock_round จาก DB เพื่อให้แน่ใจว่าหน้าประวัติแสดงเลขรอบ (แก้ปัญหาเลขหาย)
        if order_ids:
            tbl = _ol_table_name()
            sql = text(f"""
                SELECT order_id, MAX(lowstock_round) AS r
                  FROM {tbl}
                 WHERE order_id IN :oids
                 GROUP BY order_id
            """).bindparams(bindparam("oids", expanding=True))
            try:
                q_round = db.session.execute(sql, {"oids": order_ids}).all()
                round_map = {str(r[0]): (int(r[1]) if r[1] is not None else None) for r in q_round}
                for r in out:
                    oid = (r.get("order_no") or "").strip()
                    if oid in round_map and round_map[oid] is not None:
                        r["assign_round"] = round_map[oid]
            except Exception:
                pass  # ถ้าคอลัมน์ยังไม่มีก็ข้าม

        # กรองตามรอบ (หลังจากดึงค่าจาก DB แล้ว)
        if round_num and round_num != "all":
            try:
                r_int = int(round_num)
                out = [r for r in out if r.get("assign_round") == r_int]
                # อัปเดต order_ids หลังกรอง
                order_ids = sorted({(r["order_no"] or "").strip() for r in out if r.get("order_no")})
            except:
                pass

        logistics = sorted(set([r.get("shipping_type") for r in out if r.get("shipping_type")]))

        # [SCAN] ดึงข้อมูลการ Scan Order เพื่อส่งไปหน้าเว็บ
        if order_ids:
            tbl = _ol_table_name()
            sql_scan = text(f"SELECT order_id, MAX(scanned_at) FROM {tbl} WHERE order_id IN :oids GROUP BY order_id")
            sql_scan = sql_scan.bindparams(bindparam("oids", expanding=True))
            res_scan = db.session.execute(sql_scan, {"oids": order_ids}).fetchall()
            scan_map = {str(r[0]): r[1] for r in res_scan if r[0]}
            for r in out:
                oid = (r.get("order_no") or "").strip()
                r["scanned_at"] = scan_map.get(oid)

        return render_template(
            "report_lowstock.html",
            rows=out,
            summary={"sku_count": len(low_skus), "orders_count": len(order_ids)},
            printed_at=meta_printed_at,  # ข้อ 1: ใช้เวลาจริงที่ถูกบันทึกไว้
            order_ids=order_ids,
            shops=shops,
            logistics=logistics,
            platform_sel=platform,
            shop_sel=shop_id,
            logistic_sel=logistic,
            is_history_view=True,
            available_dates=_available_dates(),
            print_date_from=print_date_from,
            print_date_to=print_date_to,
            sort_col=sort_col,
            sort_dir=sort_dir,
            q=q,
            round_sel=round_num,
            date_from=date_from_str,
            date_to=date_to_str,
            import_from=import_from_str,
            import_to=import_to_str
        )

    @app.route("/report/lowstock.xlsx", methods=["GET"])
    @login_required
    def report_lowstock_export():
        """ส่งออกรายงานสินค้าน้อยเป็น Excel (ข้อ 2: ตรงกับตารางในหน้าเว็บ)"""
        # ไม่ต้องใช้ services.lowstock_queue แล้ว
        
        platform = normalize_platform(request.args.get("platform"))
        shop_id  = request.args.get("shop_id")
        logistic = request.args.get("logistic")
        q        = (request.args.get("q") or "").strip()
        sort_col = (request.args.get("sort") or "order_no").strip().lower()
        sort_dir = (request.args.get("dir") or "asc").lower()
        round_num = request.args.get("round")
        
        # รับค่าวันที่กรอง (เพิ่มใหม่)
        date_from_str = request.args.get("date_from")
        date_to_str = request.args.get("date_to")
        import_from_str = request.args.get("import_from")
        import_to_str = request.args.get("import_to")

        filters = {
            "platform": platform if platform else None,
            "shop_id": int(shop_id) if shop_id else None,
            "import_date": None
        }
        rows, _ = compute_allocation(db.session, filters)
        rows = _filter_out_cancelled_rows(rows)
        rows = _filter_out_issued_rows(rows)
        
        # คำนวณออเดอร์ที่แพ็คแล้ว (เช็คจาก sales_status)
        packed_oids = _orders_packed_set(rows)
        
        # ข้อ 4: กรอง PACKED
        safe = []
        for r in rows:
            r = dict(r)
            # กรองออเดอร์ที่อยู่ในลิสต์แพ็คแล้วออก
            if (r.get("order_id") or "").strip() in packed_oids:
                continue
            sales_status = (str(r.get("sales_status") or "")).upper()
            if sales_status == "PACKED" or bool(r.get("packed", False)):
                continue
            if "stock_qty" not in r:
                sku = (r.get("sku") or "").strip()
                stock_qty = 0
                if sku:
                    prod = Product.query.filter_by(sku=sku).first()
                    if prod and hasattr(prod, "stock_qty"):
                        try: stock_qty = int(prod.stock_qty or 0)
                        except: stock_qty = 0
                    else:
                        st = Stock.query.filter_by(sku=sku).first()
                        stock_qty = int(st.qty) if st and st.qty is not None else 0
                r["stock_qty"] = stock_qty
            r["logistic"] = r.get("logistic") or r.get("logistic_type") or "-"
            # ไม่ต้อง _recompute_allocation_row(r) เพราะ compute_allocation คำนวณให้แล้ว
            safe.append(r)

        orders_low = _orders_lowstock_order_set(safe)
        safe = [r for r in safe if (r.get("order_id") or "").strip() in orders_low]
        
        # กรองเฉพาะ allocation_status == "LOW_STOCK" ตาม compute_allocation
        lines = [r for r in safe if r.get("allocation_status") == "LOW_STOCK"]

        # กรองเพิ่ม
        if logistic:
            lines = [r for r in lines if (r.get("logistic") or "").lower().find(logistic.lower()) >= 0]
        if q:
            ql = q.lower()
            def _hit(s): return ql in (str(s or "").lower())
            lines = [r for r in lines if (
                _hit(r.get("order_id")) or _hit(r.get("sku")) or _hit(r.get("brand")) or
                _hit(r.get("model")) or _hit(r.get("shop")) or _hit(r.get("platform")) or _hit(r.get("logistic"))
            )]
        if round_num and round_num != "all":
            try:
                r_int = int(round_num)
                lines = [r for r in lines if r.get("lowstock_round") == r_int]
            except: pass

        # กรองวันที่ (เพิ่มใหม่)
        def _parse_date(s):
            if not s: return None
            try: return datetime.strptime(s, "%Y-%m-%d").date()
            except: return None
        
        d_from = _parse_date(date_from_str)
        d_to = _parse_date(date_to_str)
        i_from = _parse_date(import_from_str)
        i_to = _parse_date(import_to_str)
        
        if d_from or d_to or i_from or i_to:
            filtered_lines = []
            for r in lines:
                # กรองวันสั่งซื้อ
                pass_order = True
                if d_from or d_to:
                    od = r.get("order_time")
                    if isinstance(od, str):
                        try: od = datetime.strptime(od.split()[0], "%Y-%m-%d").date()
                        except: od = None
                    elif isinstance(od, datetime): od = od.date()
                    elif hasattr(od, 'date'): od = od.date()
                    else: od = None
                    
                    if od:
                        if d_from and od < d_from: pass_order = False
                        if d_to and od > d_to: pass_order = False
                    elif d_from or d_to:
                        pass_order = False
                
                # กรองวันนำเข้า
                pass_import = True
                if i_from or i_to:
                    id_ = r.get("import_date")
                    if isinstance(id_, str):
                        try: id_ = datetime.strptime(id_, "%Y-%m-%d").date()
                        except: id_ = None
                    elif isinstance(id_, datetime): id_ = id_.date()
                    elif hasattr(id_, 'date'): id_ = id_.date()
                    else: id_ = None
                    
                    if id_:
                        if i_from and id_ < i_from: pass_import = False
                        if i_to and id_ > i_to: pass_import = False
                    elif i_from or i_to:
                        pass_import = False

                if pass_order and pass_import:
                    filtered_lines.append(r)
            lines = filtered_lines

        # คำนวณ AllQty
        from collections import defaultdict
        sum_by_sku = defaultdict(int)
        for r in lines:
            sum_by_sku[(r.get("sku") or "").strip()] += int(r.get("qty") or 0)

        # อ่านค่า lowstock_round จาก DB เหมือนหน้ารายงาน (ข้อ 1)
        order_ids_for_round = sorted({(r.get("order_id") or "").strip() for r in lines if r.get("order_id")})
        low_round_by_oid = {}
        if order_ids_for_round:
            # ใช้ raw SQL แทน ORM เพราะ lowstock_round ไม่มีในโมเดล
            tbl = _ol_table_name()
            sql = text(f"""
                SELECT order_id, MAX(lowstock_round) AS r
                  FROM {tbl}
                 WHERE order_id IN :oids
                 GROUP BY order_id
            """).bindparams(bindparam("oids", expanding=True))
            try:
                q_round = db.session.execute(sql, {"oids": order_ids_for_round}).all()
                low_round_by_oid = {str(r[0]): (int(r[1]) if r[1] is not None else None) for r in q_round}
            except Exception:
                low_round_by_oid = {}

        # สร้าง output rows
        out = []
        for r in lines:
            sku = (r.get("sku") or "").strip()
            oid = (r.get("order_id") or "").strip()
            out.append({
                "platform":      r.get("platform"),
                "store":         r.get("shop"),
                "order_no":      oid,
                "sku":           sku,
                "brand":         r.get("brand"),
                "product_name":  r.get("model"),
                "stock":         int(r.get("stock_qty", 0) or 0),
                "qty":           int(r.get("qty", 0) or 0),
                "allqty":        sum_by_sku[sku],
                "order_time":    r.get("order_time"),
                "due_date":      r.get("due_date"),
                "sla":           r.get("sla"),
                "shipping_type": r.get("logistic"),
                "assign_round":  low_round_by_oid.get(oid, r.get("lowstock_round")),  # <<<< ใช้ค่าจาก DB
            })

        # เรียง
        sort_col = sort_col if sort_col in {"platform","store","order_no","sku","brand","product_name","stock","qty","allqty","order_time","due_date","sla","shipping_type","assign_round","printed_count"} else "order_no"
        rev = (sort_dir == "desc")
        def _key(v):
            if sort_col in {"stock","qty","allqty","assign_round","printed_count"}:
                try: return int(v.get(sort_col) or 0)
                except: return 0
            elif sort_col in {"order_time","due_date"}:
                try: return datetime.fromisoformat(str(v.get(sort_col)))
                except: return str(v.get(sort_col) or "")
            else:
                return str(v.get(sort_col) or "")
        out.sort(key=_key, reverse=rev)

        # เพิ่มคอลัมน์ "พิมพ์แล้ว(ครั้ง)"
        order_ids = sorted({(r["order_no"] or "").strip() for r in out if r.get("order_no")})
        counts_low = _get_print_counts_local(order_ids, "lowstock")
        for r in out:
            oid = (r.get("order_no") or "").strip()
            r["printed_count"] = int(counts_low.get(oid, 0))
        
        # สร้าง DataFrame
        df_data = []
        for r in out:
            df_data.append({
                "แพลตฟอร์ม": r["platform"],
                "ร้าน": r["store"],
                "เลข Order": r["order_no"],
                "SKU": r["sku"],
                "Brand": r["brand"],
                "ชื่อสินค้า": r["product_name"],
                "Stock": r["stock"],
                "Qty": r["qty"],
                "AllQty": r["allqty"],
                "เวลาที่ลูกค้าสั่ง": r["order_time"],
                "กำหนดส่ง": r["due_date"],
                "SLA (ชม.)": r["sla"],
                "ประเภทขนส่ง": r["shipping_type"],
                "จ่ายงาน(รอบที่)": r["assign_round"] if r["assign_round"] is not None else "",
                "พิมพ์แล้ว(ครั้ง)": r["printed_count"],
            })

        df = pd.DataFrame(df_data)
        bio = BytesIO()
        with pd.ExcelWriter(bio, engine="xlsxwriter") as w:
            df.to_excel(w, index=False, sheet_name="LowStock")
        bio.seek(0)
        
        filename = f"lowstock_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(
            bio,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )


    @app.route("/report/nostock", methods=["GET"])
    @login_required
    def report_nostock():
        """
        รายงานไม่มีสินค้า — กรองเฉพาะ SHORTAGE (stock = 0) เท่านั้น
        """
        platform = normalize_platform(request.args.get("platform"))
        shop_id  = request.args.get("shop_id")
        logistic = request.args.get("logistic")
        round_num = request.args.get("round")
        q        = (request.args.get("q") or "").strip()
        sort_col = (request.args.get("sort") or "").strip().lower()
        sort_dir = (request.args.get("dir") or "asc").lower()
        
        # รับค่าวันที่กรอง
        date_from_str = request.args.get("date_from")
        date_to_str = request.args.get("date_to")
        import_from_str = request.args.get("import_from")
        import_to_str = request.args.get("import_to")

        shops = Shop.query.order_by(Shop.name.asc()).all()

        # 1) ดึง allocation rows
        filters = {"platform": platform or None, "shop_id": int(shop_id) if shop_id else None, "import_date": None}
        rows, _ = compute_allocation(db.session, filters)
        rows = _filter_out_cancelled_rows(rows)
        rows = _filter_out_issued_rows(rows)

        # คำนวณออเดอร์ที่แพ็คแล้ว (เช็คจาก sales_status)
        packed_oids = _orders_packed_set(rows)

        # เติม stock_qty/logistic
        safe = []
        for r in rows:
            r = dict(r)
            # กรองออเดอร์ที่อยู่ในลิสต์แพ็คแล้วออก
            if (r.get("order_id") or "").strip() in packed_oids:
                continue
            if (str(r.get("sales_status") or "")).upper() == "PACKED" or bool(r.get("packed", False)):
                continue
            if "stock_qty" not in r:
                sku = (r.get("sku") or "").strip()
                stock_qty = 0
                if sku:
                    prod = Product.query.filter_by(sku=sku).first()
                    if prod and hasattr(prod, "stock_qty"):
                        try:
                            stock_qty = int(prod.stock_qty or 0)
                        except Exception:
                            stock_qty = 0
                    else:
                        st = Stock.query.filter_by(sku=sku).first()
                        stock_qty = int(st.qty) if st and st.qty is not None else 0
                r["stock_qty"] = stock_qty
            r["logistic"] = r.get("logistic") or r.get("logistic_type") or "-"
            # ไม่ต้อง _recompute_allocation_row(r) เพราะ compute_allocation คำนวณให้แล้ว
            safe.append(r)

        # กรองตามวันที่สั่งซื้อและวันที่นำเข้า
        if date_from_str or date_to_str:
            from datetime import datetime
            def parse_date_str(s):
                if not s: return None
                try: return datetime.strptime(s, "%Y-%m-%d").date()
                except: return None
            date_from = parse_date_str(date_from_str)
            date_to = parse_date_str(date_to_str)
            filtered = []
            for r in safe:
                order_dt = r.get("order_time")
                if isinstance(order_dt, str):
                    try: order_dt = datetime.strptime(order_dt.split()[0], "%Y-%m-%d").date()
                    except: order_dt = None
                elif isinstance(order_dt, datetime):
                    order_dt = order_dt.date()
                if order_dt:
                    if date_from and order_dt < date_from: continue
                    if date_to and order_dt > date_to: continue
                elif date_from or date_to:
                    continue
                filtered.append(r)
            safe = filtered
        
        if import_from_str or import_to_str:
            from datetime import datetime
            def parse_date_str(s):
                if not s: return None
                try: return datetime.strptime(s, "%Y-%m-%d").date()
                except: return None
            import_from = parse_date_str(import_from_str)
            import_to = parse_date_str(import_to_str)
            filtered = []
            for r in safe:
                imp_dt = r.get("import_date")
                if isinstance(imp_dt, str):
                    try: imp_dt = datetime.strptime(imp_dt, "%Y-%m-%d").date()
                    except: imp_dt = None
                elif isinstance(imp_dt, datetime):
                    imp_dt = imp_dt.date()
                elif isinstance(imp_dt, date):
                    pass
                else:
                    imp_dt = None
                if imp_dt:
                    if import_from and imp_dt < import_from: continue
                    if import_to and imp_dt > import_to: continue
                elif import_from or import_to:
                    continue
                filtered.append(r)
            safe = filtered

        # 2) กรองเฉพาะ allocation_status == "SHORTAGE" ตาม compute_allocation
        lines = [r for r in safe if r.get("allocation_status") == "SHORTAGE"]

        # 3) ฟิลเตอร์
        if logistic:
            lines = [r for r in lines if (r.get("logistic") or "").lower().find(logistic.lower()) >= 0]
        if q:
            ql = q.lower()
            lines = [r for r in lines if ql in (str(r.get("order_id","")) + str(r.get("sku","")) + 
                    str(r.get("brand","")) + str(r.get("model","")) + str(r.get("shop",""))).lower()]

        # 4) ดึงค่า nostock_round จาก DB
        order_ids_for_round = sorted({(r.get("order_id") or "").strip() for r in lines if r.get("order_id")})
        nostock_round_by_oid = {}
        if order_ids_for_round:
            tbl = _ol_table_name()
            sql = text(f"SELECT order_id, MAX(nostock_round) AS r FROM {tbl} WHERE order_id IN :oids GROUP BY order_id")
            sql = sql.bindparams(bindparam("oids", expanding=True))
            try:
                q_round = db.session.execute(sql, {"oids": order_ids_for_round}).all()
                nostock_round_by_oid = {str(r[0]): (int(r[1]) if r[1] is not None else None) for r in q_round}
            except Exception:
                nostock_round_by_oid = {}

        # กรองตาม round ถ้ามีเลือก
        if round_num not in (None, "", "all"):
            try:
                round_filter = int(round_num)
                lines = [r for r in lines if nostock_round_by_oid.get((r.get("order_id") or "").strip()) == round_filter]
            except:
                pass

        # เตรียมข้อมูล Mixed Status
        status_map = {
            "READY_ACCEPT": "พร้อมรับ",
            "LOW_STOCK": "สินค้าน้อย",
            "NOT_ENOUGH": "ไม่พอส่ง",
            "ACCEPTED": "รับแล้ว",
            "PACKED": "แพ็คแล้ว",
            "CANCELLED": "ยกเลิก",
            "ISSUED": "จ่ายงานแล้ว"
        }
        mixed_info = {}
        for r in lines:
            oid = (r.get("order_id") or "").strip()
            if oid and oid not in mixed_info:
                other_rows = [x for x in safe if (x.get("order_id") or "").strip() == oid]
                details = []
                for x in other_rows:
                    s = x.get("allocation_status")
                    if s and s != "SHORTAGE":
                        readable_status = status_map.get(s, s)
                        product_name = x.get("model") or x.get("sku") or "?"
                        details.append(f"{readable_status} ({product_name})")
                if details:
                    mixed_info[oid] = f"มีรายการอื่น: {', '.join(details)}"
                else:
                    mixed_info[oid] = ""

        # 5) แปลงเป็นคอลัมน์รายงาน
        out = []
        for r in lines:
            oid = (r.get("order_id") or "").strip()
            out.append({
                "platform":      r.get("platform"),
                "store":         r.get("shop"),
                "order_no":      oid,
                "sku":           r.get("sku"),
                "brand":         r.get("brand"),
                "product_name":  r.get("model"),
                "stock":         int(r.get("stock_qty", 0) or 0),
                "qty":           int(r.get("qty", 0) or 0),
                "order_time":    r.get("order_time"),
                "due_date":      r.get("due_date"),
                "sla":           r.get("sla"),
                "shipping_type": r.get("logistic"),
                "assign_round":  nostock_round_by_oid.get(oid, r.get("nostock_round")),
                "printed_count": 0,
                "note":          mixed_info.get(oid, ""),  # เพิ่มหมายเหตุ
            })
        
        from collections import defaultdict
        sum_by_sku = defaultdict(int)
        for r in out:
            sum_by_sku[(r["sku"] or "").strip()] += int(r["qty"] or 0)
        for r in out:
            r["allqty"] = sum_by_sku[(r["sku"] or "").strip()]

        # 6) เรียงลำดับ
        sort_col = sort_col if sort_col in {"platform","store","order_no","sku","brand","product_name","stock","qty","allqty","order_time","due_date","sla","shipping_type","assign_round","printed_count"} else "order_no"
        rev = (sort_dir == "desc")
        def _key(v):
            if sort_col in {"stock","qty","allqty","assign_round","printed_count"}:
                try: return int(v.get(sort_col) or 0)
                except: return 0
            elif sort_col in {"order_time","due_date"}:
                try: return datetime.fromisoformat(str(v.get(sort_col)))
                except: return str(v.get(sort_col) or "")
            else:
                return str(v.get(sort_col) or "")
        out.sort(key=_key, reverse=rev)

        # 7) นับ "พิมพ์แล้ว(ครั้ง)"
        order_ids = sorted({(r["order_no"] or "").strip() for r in out if r.get("order_no")})
        counts_nostock = _get_print_counts_local(order_ids, "nostock")
        for r in out:
            oid = (r.get("order_no") or "").strip()
            r["printed_count"] = int(counts_nostock.get(oid, 0))
            r["printed_at"] = None  # ไม่แสดงเวลาในหน้าปกติ

        # 8) กรองเฉพาะออเดอร์ที่ยังไม่พิมพ์
        out = [r for r in out if (r.get("printed_count") or 0) == 0]

        # 9) คำนวณสรุป + order_ids ใหม่หลังกรอง
        order_ids = sorted({(r.get("order_no") or "").strip() for r in out if r.get("order_no")})
        nostock_skus = {(r["sku"] or "").strip() for r in out if r.get("sku")}
        summary = {"sku_count": len(nostock_skus), "orders_count": len(order_ids)}

        logistics = sorted(set([r.get("shipping_type") for r in out if r.get("shipping_type")]))
        available_rounds = sorted({r["assign_round"] for r in out if r["assign_round"] is not None})

        # [SCAN] ดึงข้อมูลการ Scan Order เพื่อส่งไปหน้าเว็บ
        if order_ids:
            tbl = _ol_table_name()
            sql_scan = text(f"SELECT order_id, MAX(scanned_at) FROM {tbl} WHERE order_id IN :oids GROUP BY order_id")
            sql_scan = sql_scan.bindparams(bindparam("oids", expanding=True))
            res_scan = db.session.execute(sql_scan, {"oids": order_ids}).fetchall()
            scan_map = {str(r[0]): r[1] for r in res_scan if r[0]}
            for r in out:
                oid = (r.get("order_no") or "").strip()
                r["scanned_at"] = scan_map.get(oid)

        return render_template(
            "report_nostock_READY.html",
            rows=out,
            summary=summary,
            printed_at=None,
            order_ids=order_ids,
            shops=shops,
            logistics=logistics,
            platform_sel=platform,
            shop_sel=shop_id,
            logistic_sel=logistic,
            round_sel=round_num,
            available_rounds=available_rounds,
            sort_col=sort_col,
            sort_dir=("desc" if rev else "asc"),
            q=q,
            date_from=date_from_str,
            date_to=date_to_str,
            import_from=import_from_str,
            import_to=import_to_str,
            mixed_status=mixed_info,
            is_history_view=False
        )

    @app.post("/report/nostock/print")
    @login_required
    def report_nostock_print():
        """บันทึกการพิมพ์รายงานไม่มีสินค้า + ย้ายไปหน้าประวัติ"""
        cu = current_user()
        order_ids_raw = (request.form.get("order_ids") or "").strip()
        order_ids = [s.strip() for s in order_ids_raw.split(",") if s.strip()]
        if not order_ids:
            flash("ไม่พบออเดอร์สำหรับพิมพ์", "warning")
            return redirect(url_for("report_nostock"))

        now_iso = now_thai().isoformat()
        
        # 1. บันทึกว่าพิมพ์ No Stock แล้ว
        _mark_nostock_printed(order_ids, username=(cu.username if cu else None), when_iso=now_iso)
        
        # 2. ย้ายไป "Order จ่ายแล้ว" (Issued) ทันที
        _mark_issued(order_ids, user_id=(cu.id if cu else None), source="print:nostock", when_dt=now_thai())
        
        db.session.commit()
        return redirect(url_for("report_nostock_printed", auto_print="1"))

    @app.get("/report/nostock/printed")
    @login_required
    def report_nostock_printed():
        """ประวัติรายงานไม่มีสินค้าที่พิมพ์แล้ว"""
        platform = normalize_platform(request.args.get("platform"))
        shop_id  = request.args.get("shop_id")
        logistic = request.args.get("logistic")
        q        = (request.args.get("q") or "").strip()
        round_num = request.args.get("round")
        sort_col = (request.args.get("sort") or "order_no").strip().lower()
        sort_dir = (request.args.get("dir") or "asc").lower()
        
        # รับค่าตัวกรองวันที่สั่งซื้อและนำเข้า
        date_from_str = request.args.get("date_from")
        date_to_str = request.args.get("date_to")
        import_from_str = request.args.get("import_from")
        import_to_str = request.args.get("import_to")
        
        # รับค่าช่วงวันที่พิมพ์
        print_date_from = request.args.get("print_date_from")
        print_date_to = request.args.get("print_date_to")
        reset_mode = request.args.get("reset")  # [NEW] รับค่า reset
        action = request.args.get("action")  # [NEW] รับค่า action (เพื่อแยกการกดปุ่มกรอง กับการเข้าหน้าเว็บครั้งแรก)
        
        # [SMART DEFAULT] ถ้าไม่มีวันที่ส่งมา AND ไม่มีคำค้นหา AND ไม่ได้ reset AND ไม่ใช่การกดปุ่มกรอง -> ให้กรอง "วันนี้"
        if not action and reset_mode != 'all' and not print_date_from and not print_date_to and not q:
            # เข้าหน้าเว็บครั้งแรก (ไม่มี action) = ดูงานวันนี้
            today = now_thai().date().isoformat()
            print_date_from = today
            print_date_to = today
        # ถ้ามี action (กดปุ่มกรอง) หรือ q หรือ reset='all' แต่ไม่มีวันที่ -> ค้นหาทั้งหมด

        tbl = _ol_table_name()
        
        # ========================================================
        # [FIX] ดึงข้อมูลเฉพาะเมื่อ: มีคำค้นหา หรือ มีการเลือกวันที่
        # ========================================================
        if q:
            # กรณี 1: มีคำค้นหา -> ค้นหาทั้งหมด (Global Search)
            sql = text(f"SELECT DISTINCT order_id FROM {tbl} WHERE printed_nostock > 0")
            result = db.session.execute(sql).fetchall()
            printed_oids = [r[0] for r in result if r and r[0]]
        elif print_date_from or print_date_to:
            # กรณี 2: มีการเลือกวันที่ -> กรองตามวันที่
            sql_where = "printed_nostock > 0"
            params = {}
            if print_date_from:
                sql_where += " AND DATE(printed_nostock_at) >= :pf"
                params["pf"] = print_date_from
            if print_date_to:
                sql_where += " AND DATE(printed_nostock_at) <= :pt"
                params["pt"] = print_date_to
            sql = text(f"SELECT DISTINCT order_id FROM {tbl} WHERE {sql_where}")
            result = db.session.execute(sql, params).fetchall()
            printed_oids = [r[0] for r in result if r and r[0]]
        else:
            # กรณี 3: ไม่ค้นหา และ ไม่เลือกวัน (เช่น กด reset='all') -> ไม่แสดงอะไร
            printed_oids = []

        def _available_dates():
            sql = text(f"SELECT DISTINCT DATE(printed_nostock_at) as d FROM {tbl} WHERE printed_nostock > 0 AND printed_nostock_at IS NOT NULL ORDER BY d DESC")
            return [r[0] for r in db.session.execute(sql).fetchall()]

        shops = Shop.query.order_by(Shop.name.asc()).all()
        
        if not printed_oids:
            return render_template(
                "report_nostock_READY.html",
                rows=[],
                summary={"sku_count": 0, "orders_count": 0},
                printed_at=None,
                order_ids=[],
                shops=shops,
                logistics=[],
                platform_sel=platform,
                shop_sel=shop_id,
                logistic_sel=logistic,
                is_history_view=True,
                available_dates=_available_dates(),
                print_date_from=print_date_from,
                print_date_to=print_date_to,
                sort_col=sort_col,
                sort_dir=sort_dir,
                q=q,
                round_sel=round_num,
                date_from=date_from_str,
                date_to=date_to_str,
                import_from=import_from_str,
                import_to=import_to_str
            )

        # เตรียมตัวกรองวันที่สั่งซื้อ
        date_from_dt = None
        date_to_dt = None
        if date_from_str:
            try:
                date_from_dt = datetime.combine(parse_date_any(date_from_str), datetime.min.time(), tzinfo=TH_TZ)
            except: pass
        if date_to_str:
            try:
                date_to_dt = datetime.combine(parse_date_any(date_to_str) + timedelta(days=1), datetime.min.time(), tzinfo=TH_TZ)
            except: pass

        filters = {
            "platform": platform if platform else None,
            "shop_id": int(shop_id) if shop_id else None,
            "import_date": None,
            "date_from": date_from_dt,
            "date_to": date_to_dt
        }
        rows, _ = compute_allocation(db.session, filters)
        rows = _filter_out_cancelled_rows(rows)
        rows = [r for r in rows if (r.get("order_id") or "").strip() in printed_oids]
        
        # กรองวันที่นำเข้า (Import Date) - [แก้ไข] ถ้าไม่มีวันที่ก็เอามาด้วย
        if import_from_str or import_to_str:
            from datetime import date as date_type
            imp_from = parse_date_any(import_from_str) if import_from_str else None
            imp_to = parse_date_any(import_to_str) if import_to_str else None
            filtered_rows = []
            for r in rows:
                raw_d = r.get("import_date")
                d_obj = None
                if isinstance(raw_d, str):
                    try: d_obj = datetime.strptime(raw_d, "%Y-%m-%d").date()
                    except: pass
                elif isinstance(raw_d, datetime):
                    d_obj = raw_d.date()
                elif isinstance(raw_d, date_type):
                    d_obj = raw_d
                
                if d_obj:
                    if imp_from and d_obj < imp_from: continue
                    if imp_to and d_obj > imp_to: continue
                    filtered_rows.append(r)
                else:
                    # ข้อมูลไม่มีวันที่นำเข้า -> เอามาด้วย
                    filtered_rows.append(r)
            rows = filtered_rows

        safe = []
        for r in rows:
            r = dict(r)
            r["logistic"] = r.get("logistic") or r.get("logistic_type") or "-"
            if "stock_qty" not in r:
                sku = (r.get("sku") or "").strip()
                stock_qty = 0
                if sku:
                    prod = Product.query.filter_by(sku=sku).first()
                    if prod and hasattr(prod, "stock_qty"):
                        try: stock_qty = int(prod.stock_qty or 0)
                        except Exception: stock_qty = 0
                    else:
                        st = Stock.query.filter_by(sku=sku).first()
                        stock_qty = int(st.qty) if st and st.qty is not None else 0
                r["stock_qty"] = stock_qty
            # ไม่ต้อง _recompute เพราะ allocation_status มาจาก compute_allocation แล้ว
            safe.append(r)

        # กรองเฉพาะ SHORTAGE (stock = 0)
        def is_nostock(r):
            try:
                stk = int(r.get("stock_qty") or 0)
            except:
                stk = 0
            return (r.get("allocation_status") == "SHORTAGE") or (stk <= 0)
        
        lines = [r for r in safe if is_nostock(r)]

        # เตรียมข้อมูล Mixed Status สำหรับหน้าประวัติ
        status_map = {
            "READY_ACCEPT": "พร้อมรับ",
            "LOW_STOCK": "สินค้าน้อย",
            "NOT_ENOUGH": "ไม่พอส่ง",
            "ACCEPTED": "รับแล้ว",
            "PACKED": "แพ็คแล้ว",
            "CANCELLED": "ยกเลิก",
            "ISSUED": "จ่ายงานแล้ว"
        }
        mixed_info = {}
        for r in lines:
            oid = (r.get("order_id") or "").strip()
            if oid and oid not in mixed_info:
                other_rows = [x for x in safe if (x.get("order_id") or "").strip() == oid]
                details = []
                for x in other_rows:
                    s = x.get("allocation_status")
                    if s and s != "SHORTAGE":
                        readable_status = status_map.get(s, s)
                        product_name = x.get("model") or x.get("sku") or "?"
                        details.append(f"{readable_status} ({product_name})")
                if details:
                    mixed_info[oid] = f"มีรายการอื่น: {', '.join(details)}"
                else:
                    mixed_info[oid] = ""

        if logistic:
            lines = [r for r in lines if (r.get("logistic") or "").lower().find(logistic.lower()) >= 0]

        # กรองตามคำค้นหา (q)
        if q:
            q_lower = q.lower()
            lines = [
                r for r in lines
                if q_lower in (
                    str(r.get("order_id") or "") +
                    str(r.get("sku") or "") +
                    str(r.get("brand") or "") +
                    str(r.get("model") or "") +
                    str(r.get("shop") or "") +
                    str(r.get("platform") or "") +
                    str(r.get("logistic") or "")
                ).lower()
            ]

        # ดึงค่า nostock_round จาก DB
        order_ids_for_round = sorted({(r.get("order_id") or "").strip() for r in lines if r.get("order_id")})
        nostock_round_by_oid = {}
        if order_ids_for_round:
            sql = text(f"SELECT order_id, MAX(nostock_round) AS r FROM {tbl} WHERE order_id IN :oids GROUP BY order_id")
            sql = sql.bindparams(bindparam("oids", expanding=True))
            try:
                q_round = db.session.execute(sql, {"oids": order_ids_for_round}).all()
                nostock_round_by_oid = {str(r[0]): (int(r[1]) if r[1] is not None else None) for r in q_round}
            except Exception:
                nostock_round_by_oid = {}

        # กรองตาม round ถ้ามี
        if round_num not in (None, "", "all"):
            try:
                round_filter = int(round_num)
                lines = [r for r in lines if nostock_round_by_oid.get((r.get("order_id") or "").strip()) == round_filter]
            except:
                pass

        out = []
        for r in lines:
            oid = (r.get("order_id") or "").strip()
            out.append({
                "platform":      r.get("platform"),
                "store":         r.get("shop"),
                "order_no":      oid,
                "sku":           r.get("sku"),
                "brand":         r.get("brand"),
                "product_name":  r.get("model"),
                "stock":         int(r.get("stock_qty", 0) or 0),
                "qty":           int(r.get("qty", 0) or 0),
                "order_time":    r.get("order_time"),
                "due_date":      r.get("due_date"),
                "sla":           r.get("sla"),
                "shipping_type": r.get("logistic"),
                "assign_round":  nostock_round_by_oid.get(oid, r.get("nostock_round")),
                "printed_count": 0,
                "note":          mixed_info.get(oid, ""),  # เพิ่มหมายเหตุ
            })
        
        from collections import defaultdict
        sum_by_sku = defaultdict(int)
        for r in out:
            sum_by_sku[(r["sku"] or "").strip()] += int(r["qty"] or 0)
        for r in out:
            r["allqty"] = sum_by_sku[(r["sku"] or "").strip()]

        # เรียง
        sort_col = sort_col if sort_col in {"platform","store","order_no","sku","brand","product_name","stock","qty","allqty","order_time","due_date","sla","shipping_type","assign_round","printed_count"} else "order_no"
        rev = (sort_dir == "desc")
        def _key(v):
            if sort_col in {"stock","qty","allqty","assign_round","printed_count"}:
                try: return int(v.get(sort_col) or 0)
                except: return 0
            elif sort_col in {"order_time","due_date"}:
                try: return datetime.fromisoformat(str(v.get(sort_col)))
                except: return str(v.get(sort_col) or "")
            else:
                return str(v.get(sort_col) or "")
        out.sort(key=_key, reverse=rev)

        order_ids = sorted({(r["order_no"] or "").strip() for r in out if r.get("order_no")})
        counts_nostock = _get_print_counts_local(order_ids, "nostock")
        for r in out:
            oid = (r.get("order_no") or "").strip()
            r["printed_count"] = int(counts_nostock.get(oid, 0))

        # ดึงเวลาพิมพ์จาก DB
        sql_ts = text(f"""
            SELECT order_id, MAX(printed_nostock_at) AS ts
            FROM {tbl}
            WHERE printed_nostock > 0
              AND order_id IN :oids
            GROUP BY order_id
        """).bindparams(bindparam("oids", expanding=True))
        rows_ts = db.session.execute(sql_ts, {"oids": order_ids}).all() if order_ids else []
        ts_map = {}
        for row_ts in rows_ts:
            if not row_ts or not row_ts[0] or not row_ts[1]:
                continue
            oid_str = str(row_ts[0]).strip()
            ts_str = row_ts[1]
            try:
                dt = datetime.fromisoformat(ts_str)
                if dt.tzinfo is None:
                    dt = TH_TZ.localize(dt)
                ts_map[oid_str] = dt
            except Exception:
                pass

        for r in out:
            r["printed_at"] = ts_map.get((r.get("order_no") or "").strip())

        meta_printed_at = max(ts_map.values()) if ts_map else None

        # ดึงค่า nostock_round จาก DB
        if order_ids:
            sql = text(f"SELECT order_id, MAX(nostock_round) AS r FROM {tbl} WHERE order_id IN :oids GROUP BY order_id")
            sql = sql.bindparams(bindparam("oids", expanding=True))
            try:
                q_round = db.session.execute(sql, {"oids": order_ids}).all()
                round_map = {str(r[0]): (int(r[1]) if r[1] is not None else None) for r in q_round}
                for r in out:
                    oid = (r.get("order_no") or "").strip()
                    if oid in round_map and round_map[oid] is not None:
                        r["assign_round"] = round_map[oid]
            except Exception:
                pass

        if round_num and round_num != "all":
            try:
                r_int = int(round_num)
                out = [r for r in out if r.get("assign_round") == r_int]
                order_ids = sorted({(r["order_no"] or "").strip() for r in out if r.get("order_no")})
            except:
                pass

        logistics = sorted(set([r.get("shipping_type") for r in out if r.get("shipping_type")]))
        nostock_skus = {(r["sku"] or "").strip() for r in out if r.get("sku")}

        # [SCAN] ดึงข้อมูลการ Scan Order เพื่อส่งไปหน้าเว็บ
        if order_ids:
            tbl = _ol_table_name()
            sql_scan = text(f"SELECT order_id, MAX(scanned_at) FROM {tbl} WHERE order_id IN :oids GROUP BY order_id")
            sql_scan = sql_scan.bindparams(bindparam("oids", expanding=True))
            res_scan = db.session.execute(sql_scan, {"oids": order_ids}).fetchall()
            scan_map = {str(r[0]): r[1] for r in res_scan if r[0]}
            for r in out:
                oid = (r.get("order_no") or "").strip()
                r["scanned_at"] = scan_map.get(oid)

        return render_template(
            "report_nostock_READY.html",
            rows=out,
            summary={"sku_count": len(nostock_skus), "orders_count": len(order_ids)},
            printed_at=meta_printed_at,
            order_ids=order_ids,
            shops=shops,
            logistics=logistics,
            platform_sel=platform,
            shop_sel=shop_id,
            logistic_sel=logistic,
            is_history_view=True,
            available_dates=_available_dates(),
            print_date_from=print_date_from,
            print_date_to=print_date_to,
            sort_col=sort_col,
            sort_dir=sort_dir,
            q=q,
            round_sel=round_num,
            date_from=date_from_str,
            date_to=date_to_str,
            import_from=import_from_str,
            import_to=import_to_str
        )

    @app.route("/report/nostock.xlsx", methods=["GET"])
    @login_required
    def report_nostock_export():
        """Export Excel รายงานไม่มีสินค้า"""
        # ไม่ต้องใช้ services.lowstock แล้ว
        import pandas as pd
        
        platform = normalize_platform(request.args.get("platform"))
        shop_id = request.args.get("shop_id")
        logistic = request.args.get("logistic")
        q = (request.args.get("q") or "").strip()
        round_num = request.args.get("round")
        
        # รับค่าวันที่กรอง (เพิ่มใหม่)
        date_from_str = request.args.get("date_from")
        date_to_str = request.args.get("date_to")
        import_from_str = request.args.get("import_from")
        import_to_str = request.args.get("import_to")
        
        filters = {"platform": platform or None, "shop_id": int(shop_id) if shop_id else None, "import_date": None}
        rows, _ = compute_allocation(db.session, filters)
        rows = _filter_out_cancelled_rows(rows)
        rows = _filter_out_issued_rows(rows)
        
        # คำนวณออเดอร์ที่แพ็คแล้ว (เช็คจาก sales_status)
        packed_oids = _orders_packed_set(rows)
        
        safe = []
        for r in rows:
            r = dict(r)
            # กรองออเดอร์ที่อยู่ในลิสต์แพ็คแล้วออก
            if (r.get("order_id") or "").strip() in packed_oids:
                continue
            if (str(r.get("sales_status") or "")).upper() == "PACKED":
                continue
            if "stock_qty" not in r:
                sku = (r.get("sku") or "").strip()
                stock_qty = 0
                if sku:
                    prod = Product.query.filter_by(sku=sku).first()
                    if prod and hasattr(prod, "stock_qty"):
                        try: stock_qty = int(prod.stock_qty or 0)
                        except: stock_qty = 0
                r["stock_qty"] = stock_qty
            safe.append(r)
        
        # กรองเฉพาะ allocation_status == "SHORTAGE"
        lines = [r for r in safe if r.get("allocation_status") == "SHORTAGE"]
        
        # ---------- กรอง logistic ----------
        if logistic:
            lines = [r for r in lines if (r.get("logistic") or "").strip().upper() == logistic.strip().upper()]
        
        # ---------- กรอง round ----------
        if round_num:
            try:
                round_int = int(round_num)
                lines = [r for r in lines if r.get("nostock_round") == round_int]
            except:
                pass
        
        # ---------- กรองด้วย date filters (เพิ่มใหม่) ----------
        def _parse_date(d_str):
            if not d_str:
                return None
            from utils import parse_datetime_guess
            dt = parse_datetime_guess(d_str)
            if dt:
                return dt.date() if hasattr(dt, 'date') else dt
            return None
        
        date_from = _parse_date(date_from_str)
        date_to = _parse_date(date_to_str)
        import_from = _parse_date(import_from_str)
        import_to = _parse_date(import_to_str)
        
        # กรองด้วยวันที่สั่ง (order_time / due_date)
        if date_from or date_to:
            filtered = []
            for r in lines:
                order_time = r.get("order_time") or r.get("due_date")
                if not order_time:
                    continue
                try:
                    from utils import parse_datetime_guess
                    dt = parse_datetime_guess(order_time)
                    if dt:
                        dt_date = dt.date() if hasattr(dt, 'date') else dt
                        if date_from and dt_date < date_from:
                            continue
                        if date_to and dt_date > date_to:
                            continue
                        filtered.append(r)
                except:
                    continue
            lines = filtered
        
        # กรองด้วยวันที่นำเข้า (import_date)
        if import_from or import_to:
            filtered = []
            for r in lines:
                imp_date = r.get("import_date")
                if not imp_date:
                    continue
                try:
                    from utils import parse_datetime_guess
                    dt = parse_datetime_guess(imp_date)
                    if dt:
                        dt_date = dt.date() if hasattr(dt, 'date') else dt
                        if import_from and dt_date < import_from:
                            continue
                        if import_to and dt_date > import_to:
                            continue
                        filtered.append(r)
                except:
                    continue
            lines = filtered
        
        # ---------- กรองด้วยคำค้นหา q ----------
        if q:
            q_lower = q.lower()
            lines = [r for r in lines if q_lower in (r.get("sku") or "").lower() 
                     or q_lower in (r.get("model") or "").lower() 
                     or q_lower in (r.get("order_id") or "").lower()]
        
        df = pd.DataFrame([{
            "แพลตฟอร์ม": r.get("platform"),
            "ร้าน": r.get("shop"),
            "เลข Order": r.get("order_id"),
            "SKU": r.get("sku"),
            "Brand": r.get("brand"),
            "ชื่อสินค้า": r.get("model"),
            "Stock": int(r.get("stock_qty", 0) or 0),
            "Qty": int(r.get("qty", 0) or 0),
            "เวลาที่ลูกค้าสั่ง": r.get("order_time"),
            "กำหนดส่ง": r.get("due_date"),
            "ประเภทขนส่ง": r.get("logistic"),
        } for r in lines])
        
        out = BytesIO()
        with pd.ExcelWriter(out, engine="xlsxwriter") as w:
            df.to_excel(w, index=False, sheet_name="NoStock")
        out.seek(0)
        return send_file(out, as_attachment=True, download_name="report_nostock.xlsx")

    # ================== NEW: Update No Stock Round ==================
    @app.route("/report/nostock/update_round", methods=["POST"])
    @login_required
    def update_nostock_round():
        """อัปเดตรอบสำหรับรายงานไม่มีสินค้า"""
        data = request.get_json() or {}
        order_ids = data.get("order_ids", [])
        round_num = data.get("round")
        
        if not order_ids or round_num is None:
            return jsonify({"success": False, "message": "ข้อมูลไม่ครบ"})
        
        try:
            round_int = int(round_num)
        except:
            return jsonify({"success": False, "message": "รอบต้องเป็นตัวเลข"})
        
        tbl = _ol_table_name()
        sql = text(f"UPDATE {tbl} SET nostock_round = :r WHERE order_id IN :oids")
        sql = sql.bindparams(bindparam("oids", expanding=True))
        db.session.execute(sql, {"r": round_int, "oids": order_ids})
        db.session.commit()
        
        return jsonify({"success": True, "message": f"อัปเดตรอบเป็น {round_int} สำเร็จ ({len(order_ids)} ออเดอร์)"})
    # ================== /NEW ==================

    # ================== NEW: Report Not Enough (สินค้าไม่พอส่ง) ==================
    @app.route("/report/notenough", methods=["GET"])
    @login_required
    def report_notenough():
        """รายงานสินค้าไม่พอส่ง (NOT_ENOUGH) — กรองเฉพาะสินค้าไม่พอส่ง"""
        platform = normalize_platform(request.args.get("platform"))
        shop_id  = request.args.get("shop_id")
        logistic = request.args.get("logistic")
        round_num = request.args.get("round")
        q        = (request.args.get("q") or "").strip()
        sort_col = (request.args.get("sort") or "").strip().lower()
        sort_dir = (request.args.get("dir") or "asc").lower()
        
        # รับค่าวันที่กรอง
        date_from_str = request.args.get("date_from")
        date_to_str = request.args.get("date_to")
        import_from_str = request.args.get("import_from")
        import_to_str = request.args.get("import_to")

        shops = Shop.query.order_by(Shop.name.asc()).all()

        # 1) ดึง allocation rows
        filters = {"platform": platform or None, "shop_id": int(shop_id) if shop_id else None, "import_date": None}
        rows, _ = compute_allocation(db.session, filters)
        rows = _filter_out_cancelled_rows(rows)
        rows = _filter_out_issued_rows(rows)

        # คำนวณออเดอร์ที่แพ็คแล้ว (เช็คจาก sales_status)
        packed_oids = _orders_packed_set(rows)
        
        safe = []
        for r in rows:
            r = dict(r)
            # กรองออเดอร์ที่อยู่ในลิสต์แพ็คแล้วออก
            if (r.get("order_id") or "").strip() in packed_oids:
                continue
            # หรือถ้า sales_status เป็น 'PACKED' ก็ข้ามไป
            if (str(r.get("sales_status") or "")).upper() == "PACKED":
                continue
            if bool(r.get("packed", False)):
                continue
            
            # ตรวจ stock_qty (ถ้า compute_allocation ไม่ได้เติมให้)
            if "stock_qty" not in r:
                sku = (r.get("sku") or "").strip()
                stock_qty = 0
                if sku:
                    prod = Product.query.filter_by(sku=sku).first()
                    if prod and hasattr(prod, "stock_qty"):
                        try:
                            stock_qty = int(prod.stock_qty or 0)
                        except:
                            stock_qty = 0
                    if not prod:
                        st = Stock.query.filter_by(sku=sku).first()
                        if st and st.qty is not None:
                            stock_qty = int(st.qty)
                r["stock_qty"] = stock_qty
            
            r["logistic"] = r.get("logistic") or r.get("logistic_type") or "-"
            # ไม่ต้อง _recompute_allocation_row(r) เพราะ compute_allocation คำนวณให้แล้ว
            safe.append(r)

        # กรองตามวันที่สั่งซื้อและวันที่นำเข้า
        if date_from_str or date_to_str:
            from datetime import datetime
            def parse_date_str(s):
                if not s: return None
                try: return datetime.strptime(s, "%Y-%m-%d").date()
                except: return None
            date_from = parse_date_str(date_from_str)
            date_to = parse_date_str(date_to_str)
            filtered = []
            for r in safe:
                order_dt = r.get("order_time")
                if isinstance(order_dt, str):
                    try: order_dt = datetime.strptime(order_dt.split()[0], "%Y-%m-%d").date()
                    except: order_dt = None
                elif isinstance(order_dt, datetime):
                    order_dt = order_dt.date()
                if order_dt:
                    if date_from and order_dt < date_from: continue
                    if date_to and order_dt > date_to: continue
                elif date_from or date_to:
                    continue
                filtered.append(r)
            safe = filtered
        
        if import_from_str or import_to_str:
            from datetime import datetime
            def parse_date_str(s):
                if not s: return None
                try: return datetime.strptime(s, "%Y-%m-%d").date()
                except: return None
            import_from = parse_date_str(import_from_str)
            import_to = parse_date_str(import_to_str)
            filtered = []
            for r in safe:
                imp_dt = r.get("import_date")
                if isinstance(imp_dt, str):
                    try: imp_dt = datetime.strptime(imp_dt, "%Y-%m-%d").date()
                    except: imp_dt = None
                elif isinstance(imp_dt, datetime):
                    imp_dt = imp_dt.date()
                elif isinstance(imp_dt, date):
                    pass
                else:
                    imp_dt = None
                if imp_dt:
                    if import_from and imp_dt < import_from: continue
                    if import_to and imp_dt > import_to: continue
                elif import_from or import_to:
                    continue
                filtered.append(r)
            safe = filtered

        # กรองเฉพาะ allocation_status == "NOT_ENOUGH" ตาม compute_allocation
        lines = [r for r in safe if r.get("allocation_status") == "NOT_ENOUGH"]

        # Filter ตามขนส่ง
        if logistic:
            lines = [r for r in lines if (r.get("logistic") or "").lower().find(logistic.lower()) >= 0]

        # Search
        if q:
            ql = q.lower()
            lines = [
                r for r in lines 
                if ql in (
                    str(r.get("order_id","")) + 
                    str(r.get("sku","")) + 
                    str(r.get("model","")) +
                    str(r.get("brand","")) +
                    str(r.get("shop","")) +
                    str(r.get("logistic",""))
                ).lower()
            ]

        # ดึง Round
        order_ids_for_round = sorted({(r.get("order_id") or "").strip() for r in lines if r.get("order_id")})
        round_by_oid = {}
        if order_ids_for_round:
            tbl = _ol_table_name()
            sql = text(f"SELECT order_id, MAX(notenough_round) AS r FROM {tbl} WHERE order_id IN :oids GROUP BY order_id")
            sql = sql.bindparams(bindparam("oids", expanding=True))
            try:
                q_round = db.session.execute(sql, {"oids": order_ids_for_round}).all()
                round_by_oid = {str(r[0]): (int(r[1]) if r[1] is not None else None) for r in q_round}
            except:
                pass

        # Filter by round
        if round_num not in (None, "", "all"):
            try:
                r_int = int(round_num)
                lines = [r for r in lines if round_by_oid.get((r.get("order_id") or "").strip()) == r_int]
            except:
                pass

        # เตรียมข้อมูล Mixed Status
        status_map = {
            "READY_ACCEPT": "พร้อมรับ",
            "LOW_STOCK": "สินค้าน้อย",
            "SHORTAGE": "ไม่มีของ",
            "ACCEPTED": "รับแล้ว",
            "PACKED": "แพ็คแล้ว",
            "CANCELLED": "ยกเลิก",
            "ISSUED": "จ่ายงานแล้ว"
        }
        mixed_info = {}
        for r in lines:
            oid = (r.get("order_id") or "").strip()
            if oid and oid not in mixed_info:
                other_rows = [x for x in safe if (x.get("order_id") or "").strip() == oid]
                details = []
                for x in other_rows:
                    s = x.get("allocation_status")
                    if s and s != "NOT_ENOUGH":
                        readable_status = status_map.get(s, s)
                        product_name = x.get("model") or x.get("sku") or "?"
                        details.append(f"{readable_status} ({product_name})")
                if details:
                    mixed_info[oid] = f"มีรายการอื่น: {', '.join(details)}"
                else:
                    mixed_info[oid] = ""

        # Map output
        out = []
        for r in lines:
            oid = (r.get("order_id") or "").strip()
            out.append({
                "platform": r.get("platform"),
                "store": r.get("shop"),
                "order_no": oid,
                "sku": r.get("sku"),
                "brand": r.get("brand"),
                "product_name": r.get("model"),
                "stock": int(r.get("stock_qty", 0) or 0),
                "qty": int(r.get("qty", 0) or 0),
                "order_time": r.get("order_time"),
                "due_date": r.get("due_date"),
                "sla": r.get("sla"),
                "shipping_type": r.get("logistic"),
                "assign_round": round_by_oid.get(oid),
                "printed_count": 0,
                "printed_at": None,
                "note": mixed_info.get(oid, ""),  # เพิ่มหมายเหตุ
            })
        
        # AllQty
        from collections import defaultdict
        sum_by_sku = defaultdict(int)
        for r in out:
            sum_by_sku[(r["sku"] or "").strip()] += int(r["qty"] or 0)
        for r in out:
            r["allqty"] = sum_by_sku[(r["sku"] or "").strip()]

        # Sort
        sort_col = sort_col if sort_col else "order_no"
        rev = (sort_dir == "desc")
        def _key(v):
            return str(v.get(sort_col) or "")
        out.sort(key=_key, reverse=rev)

        # Print Count
        oids = sorted({(r["order_no"] or "").strip() for r in out if r["order_no"]})
        counts = _get_print_counts_local(oids, "notenough")
        
        # [เพิ่ม] ดึงเวลาพิมพ์ล่าสุด (printed_notenough_at) จาก DB
        ts_map = {}
        if oids:
            tbl = _ol_table_name()
            sql_ts = text(f"""
                SELECT order_id, MAX(printed_notenough_at) 
                FROM {tbl} 
                WHERE order_id IN :oids 
                GROUP BY order_id
            """).bindparams(bindparam("oids", expanding=True))
            try:
                res_ts = db.session.execute(sql_ts, {"oids": oids}).fetchall()
                for row in res_ts:
                    if row[1]:
                        dt = datetime.fromisoformat(row[1])
                        if dt.tzinfo is None: dt = TH_TZ.localize(dt)
                        ts_map[str(row[0])] = dt
            except: pass

        for r in out:
            oid = (r.get("order_no") or "").strip()
            r["printed_count"] = int(counts.get(oid, 0))
            r["printed_at"] = ts_map.get(oid)  # ใส่เวลาจริงแทน None

        # กรองที่พิมพ์แล้วออก (ไม่แสดงในรายงานหลัก)
        out = [r for r in out if r["printed_count"] == 0]
        
        # Summary
        final_oids = sorted({(r["order_no"] or "").strip() for r in out if r["order_no"]})
        skus = {(r["sku"] or "").strip() for r in out if r["sku"]}
        summary = {
            "sku_count": len(skus),
            "orders_count": len(final_oids),
        }
        
        # ดึงรายการขนส่ง
        logistics = sorted(set([r.get("shipping_type") for r in out if r.get("shipping_type")]))
        
        # ดึงรอบที่มี
        available_rounds = sorted({r["assign_round"] for r in out if r["assign_round"] is not None})

        # [SCAN] ดึงข้อมูลการ Scan Order เพื่อส่งไปหน้าเว็บ
        if final_oids:
            tbl = _ol_table_name()
            sql_scan = text(f"SELECT order_id, MAX(scanned_at) FROM {tbl} WHERE order_id IN :oids GROUP BY order_id")
            sql_scan = sql_scan.bindparams(bindparam("oids", expanding=True))
            res_scan = db.session.execute(sql_scan, {"oids": final_oids}).fetchall()
            scan_map = {str(r[0]): r[1] for r in res_scan if r[0]}
            for r in out:
                oid = (r.get("order_no") or "").strip()
                r["scanned_at"] = scan_map.get(oid)

        return render_template(
            "report_notenough.html",
            rows=out,
            summary=summary,
            printed_at=None,
            order_ids=final_oids,
            shops=shops,
            logistics=logistics,
            platform_sel=platform,
            shop_sel=shop_id,
            logistic_sel=logistic,
            round_sel=round_num,
            available_rounds=available_rounds,
            sort_col=sort_col,
            sort_dir=sort_dir,
            q=q,
            date_from=date_from_str,
            date_to=date_to_str,
            import_from=import_from_str,
            import_to=import_to_str,
            mixed_status=mixed_info,
            is_history_view=False
        )

    @app.post("/report/notenough/print")
    @login_required
    def report_notenough_print():
        """บันทึกการพิมพ์รายงานสินค้าไม่พอส่ง + ย้ายไปหน้าประวัติ"""
        cu = current_user()
        order_ids_raw = (request.form.get("order_ids") or "").strip()
        order_ids = [s.strip() for s in order_ids_raw.split(",") if s.strip()]
        if not order_ids:
            flash("ไม่พบออเดอร์สำหรับพิมพ์", "warning")
            return redirect(url_for("report_notenough"))
        
        now_iso = now_thai().isoformat()
        
        # 1. บันทึกว่าพิมพ์ Not Enough แล้ว
        _mark_notenough_printed(order_ids, username=(cu.username if cu else None), when_iso=now_iso)
        
        # 2. ย้ายไป "Order จ่ายแล้ว" (Issued) ทันที
        _mark_issued(order_ids, user_id=(cu.id if cu else None), source="print:notenough", when_dt=now_thai())
        
        db.session.commit()
        return redirect(url_for("report_notenough_printed", auto_print="1"))

    @app.get("/report/notenough/printed")
    @login_required
    def report_notenough_printed():
        """ประวัติรายงานสินค้าไม่พอส่งที่พิมพ์แล้ว"""
        platform = normalize_platform(request.args.get("platform"))
        shop_id  = request.args.get("shop_id")
        logistic = request.args.get("logistic")
        q        = (request.args.get("q") or "").strip()
        round_num = request.args.get("round")
        sort_col = (request.args.get("sort") or "order_no").strip().lower()
        sort_dir = (request.args.get("dir") or "asc").lower()
        
        # รับค่าตัวกรองวันที่สั่งซื้อและนำเข้า
        date_from_str = request.args.get("date_from")
        date_to_str = request.args.get("date_to")
        import_from_str = request.args.get("import_from")
        import_to_str = request.args.get("import_to")
        
        # รับค่าช่วงวันที่พิมพ์
        print_date_from = request.args.get("print_date_from")
        print_date_to = request.args.get("print_date_to")
        reset_mode = request.args.get("reset")  # [NEW] รับค่า reset
        action = request.args.get("action")  # [NEW] รับค่า action (เพื่อแยกการกดปุ่มกรอง กับการเข้าหน้าเว็บครั้งแรก)
        
        # [SMART DEFAULT] ถ้าไม่มีวันที่ส่งมา AND ไม่มีคำค้นหา AND ไม่ได้ reset AND ไม่ใช่การกดปุ่มกรอง -> ให้กรอง "วันนี้"
        if not action and reset_mode != 'all' and not print_date_from and not print_date_to and not q:
            # เข้าหน้าเว็บครั้งแรก (ไม่มี action) = ดูงานวันนี้
            today = now_thai().date().isoformat()
            print_date_from = today
            print_date_to = today
        # ถ้ามี action (กดปุ่มกรอง) หรือ q หรือ reset='all' แต่ไม่มีวันที่ -> ค้นหาทั้งหมด

        tbl = _ol_table_name()
        
        # ========================================================
        # [FIX] ดึงข้อมูลเฉพาะเมื่อ: มีคำค้นหา หรือ มีการเลือกวันที่
        # ========================================================
        if q:
            # กรณี 1: มีคำค้นหา -> ค้นหาทั้งหมด (Global Search)
            sql = text(f"SELECT DISTINCT order_id FROM {tbl} WHERE printed_notenough > 0")
            result = db.session.execute(sql).fetchall()
            printed_oids = [r[0] for r in result if r and r[0]]
        elif print_date_from or print_date_to:
            # กรณี 2: มีการเลือกวันที่ -> กรองตามวันที่
            sql_where = "printed_notenough > 0"
            params = {}
            if print_date_from:
                sql_where += " AND DATE(printed_notenough_at) >= :pf"
                params["pf"] = print_date_from
            if print_date_to:
                sql_where += " AND DATE(printed_notenough_at) <= :pt"
                params["pt"] = print_date_to
            sql = text(f"SELECT DISTINCT order_id FROM {tbl} WHERE {sql_where}")
            result = db.session.execute(sql, params).fetchall()
            printed_oids = [r[0] for r in result if r and r[0]]
        else:
            # กรณี 3: ไม่ค้นหา และ ไม่เลือกวัน (เช่น กด reset='all') -> ไม่แสดงอะไร
            printed_oids = []

        def _available_dates():
            sql = text(f"SELECT DISTINCT DATE(printed_notenough_at) as d FROM {tbl} WHERE printed_notenough > 0 AND printed_notenough_at IS NOT NULL ORDER BY d DESC")
            return [r[0] for r in db.session.execute(sql).fetchall()]

        shops = Shop.query.order_by(Shop.name.asc()).all()
        
        if not printed_oids:
            return render_template(
                "report_notenough.html",
                rows=[],
                summary={"sku_count": 0, "orders_count": 0},
                printed_at=None,
                order_ids=[],
                shops=shops,
                logistics=[],
                platform_sel=platform,
                shop_sel=shop_id,
                logistic_sel=logistic,
                is_history_view=True,
                available_dates=_available_dates(),
                print_date_from=print_date_from,
                print_date_to=print_date_to,
                sort_col=sort_col,
                sort_dir=sort_dir,
                q=q,
                round_sel=round_num,
                date_from=date_from_str,
                date_to=date_to_str,
                import_from=import_from_str,
                import_to=import_to_str
            )

        # เตรียมตัวกรองวันที่สั่งซื้อ
        date_from_dt = None
        date_to_dt = None
        if date_from_str:
            try:
                date_from_dt = datetime.combine(parse_date_any(date_from_str), datetime.min.time(), tzinfo=TH_TZ)
            except: pass
        if date_to_str:
            try:
                date_to_dt = datetime.combine(parse_date_any(date_to_str) + timedelta(days=1), datetime.min.time(), tzinfo=TH_TZ)
            except: pass

        # ดึงข้อมูลจริง
        filters = {
            "platform": platform or None,
            "shop_id": int(shop_id) if shop_id else None,
            "import_date": None,
            "date_from": date_from_dt,
            "date_to": date_to_dt
        }
        rows, _ = compute_allocation(db.session, filters)
        rows = _filter_out_cancelled_rows(rows)
        # [FIX] ในหน้าประวัติ (printed) ไม่กรอง Issued ออก เพราะเราเพิ่ง mark Issued ไป
        rows = [r for r in rows if (r.get("order_id") or "").strip() in printed_oids]
        
        # กรองวันที่นำเข้า (Import Date) - [แก้ไข] ถ้าไม่มีวันที่ก็เอามาด้วย
        if import_from_str or import_to_str:
            from datetime import date as date_type
            imp_from = parse_date_any(import_from_str) if import_from_str else None
            imp_to = parse_date_any(import_to_str) if import_to_str else None
            filtered_rows = []
            for r in rows:
                raw_d = r.get("import_date")
                d_obj = None
                if isinstance(raw_d, str):
                    try: d_obj = datetime.strptime(raw_d, "%Y-%m-%d").date()
                    except: pass
                elif isinstance(raw_d, datetime):
                    d_obj = raw_d.date()
                elif isinstance(raw_d, date_type):
                    d_obj = raw_d
                
                if d_obj:
                    if imp_from and d_obj < imp_from: continue
                    if imp_to and d_obj > imp_to: continue
                    filtered_rows.append(r)
                else:
                    # ข้อมูลไม่มีวันที่นำเข้า -> เอามาด้วย
                    filtered_rows.append(r)
            rows = filtered_rows

        packed_oids = _orders_packed_set(rows)
        
        safe = []
        for r in rows:
            r = dict(r)
            oid = (r.get("order_id") or "").strip()
            if oid not in printed_oids:
                continue
            if oid in packed_oids:
                continue
            if (str(r.get("sales_status") or "")).upper() == "PACKED":
                continue
            if bool(r.get("packed", False)):
                continue
            
            if "stock_qty" not in r:
                sku = (r.get("sku") or "").strip()
                stock_qty = 0
                if sku:
                    prod = Product.query.filter_by(sku=sku).first()
                    if prod and hasattr(prod, "stock_qty"):
                        try:
                            stock_qty = int(prod.stock_qty or 0)
                        except:
                            stock_qty = 0
                r["stock_qty"] = stock_qty
            
            r["logistic"] = r.get("logistic") or r.get("logistic_type") or "-"
            # ไม่ต้อง _recompute_allocation_row(r) เพราะ compute_allocation คำนวณให้แล้ว
            safe.append(r)

        # [CRITICAL FIX] Fallback Logic สำหรับ Not Enough
        # เพราะออเดอร์ถูก mark เป็น ISSUED แล้ว allocation_status อาจไม่ใช่ NOT_ENOUGH
        def _is_not_enough_for_history(r):
            # 1. ถ้า status เป็น NOT_ENOUGH อยู่แล้ว -> เอา
            if r.get("allocation_status") == "NOT_ENOUGH": return True
            # 2. Fallback: ถ้า stock < qty (เกณฑ์ Not Enough) -> เอา
            try:
                s = int(r.get("stock_qty") or 0)
                q = int(r.get("qty") or 0)
                return s < q and s > 0  # stock มีแต่ไม่พอ
            except: return False

        lines = [r for r in safe if _is_not_enough_for_history(r)]

        # เตรียมข้อมูล Mixed Status สำหรับหน้าประวัติ
        status_map = {
            "READY_ACCEPT": "พร้อมรับ",
            "LOW_STOCK": "สินค้าน้อย",
            "SHORTAGE": "ไม่มีของ",
            "ACCEPTED": "รับแล้ว",
            "PACKED": "แพ็คแล้ว",
            "CANCELLED": "ยกเลิก",
            "ISSUED": "จ่ายงานแล้ว"
        }
        mixed_info = {}
        for r in lines:
            oid = (r.get("order_id") or "").strip()
            if oid and oid not in mixed_info:
                other_rows = [x for x in safe if (x.get("order_id") or "").strip() == oid]
                details = []
                for x in other_rows:
                    s = x.get("allocation_status")
                    if s and s != "NOT_ENOUGH":
                        readable_status = status_map.get(s, s)
                        product_name = x.get("model") or x.get("sku") or "?"
                        details.append(f"{readable_status} ({product_name})")
                if details:
                    mixed_info[oid] = f"มีรายการอื่น: {', '.join(details)}"
                else:
                    mixed_info[oid] = ""

        if logistic:
            lines = [r for r in lines if (r.get("logistic") or "").lower().find(logistic.lower()) >= 0]

        # กรองตามคำค้นหา (q)
        if q:
            q_lower = q.lower()
            lines = [
                r for r in lines
                if q_lower in (
                    str(r.get("order_id") or "") +
                    str(r.get("sku") or "") +
                    str(r.get("brand") or "") +
                    str(r.get("model") or "") +
                    str(r.get("shop") or "") +
                    str(r.get("platform") or "") +
                    str(r.get("logistic") or "")
                ).lower()
            ]

        # ดึง Round
        order_ids_for_round = sorted({(r.get("order_id") or "").strip() for r in lines if r.get("order_id")})
        round_by_oid = {}
        if order_ids_for_round:
            sql = text(f"SELECT order_id, MAX(notenough_round) AS r FROM {tbl} WHERE order_id IN :oids GROUP BY order_id")
            sql = sql.bindparams(bindparam("oids", expanding=True))
            try:
                q_round = db.session.execute(sql, {"oids": order_ids_for_round}).all()
                round_by_oid = {str(r[0]): (int(r[1]) if r[1] is not None else None) for r in q_round}
            except:
                pass

        if round_num not in (None, "", "all"):
            try:
                r_int = int(round_num)
                lines = [r for r in lines if round_by_oid.get((r.get("order_id") or "").strip()) == r_int]
            except:
                pass

        out = []
        for r in lines:
            oid = (r.get("order_id") or "").strip()
            out.append({
                "platform": r.get("platform"),
                "store": r.get("shop"),
                "order_no": oid,
                "sku": r.get("sku"),
                "brand": r.get("brand"),
                "product_name": r.get("model"),
                "stock": int(r.get("stock_qty", 0) or 0),
                "qty": int(r.get("qty", 0) or 0),
                "order_time": r.get("order_time"),
                "due_date": r.get("due_date"),
                "sla": r.get("sla"),
                "shipping_type": r.get("logistic"),
                "assign_round": round_by_oid.get(oid),
                "printed_count": 0,
                "printed_at": None,
                "note": mixed_info.get(oid, ""),  # เพิ่มหมายเหตุ
            })
        
        from collections import defaultdict
        sum_by_sku = defaultdict(int)
        for r in out:
            sum_by_sku[(r["sku"] or "").strip()] += int(r["qty"] or 0)
        for r in out:
            r["allqty"] = sum_by_sku[(r["sku"] or "").strip()]

        sort_col = sort_col if sort_col else "order_no"
        rev = (sort_dir == "desc")
        def _key(v):
            return str(v.get(sort_col) or "")
        out.sort(key=_key, reverse=rev)

        oids = sorted({(r["order_no"] or "").strip() for r in out if r["order_no"]})
        counts = _get_print_counts_local(oids, "notenough")
        for r in out:
            r["printed_count"] = int(counts.get(r["order_no"], 0))

        # [เพิ่ม] ดึงเวลาพิมพ์จาก DB (ใช้ printed_notenough_at ที่ถูกต้อง)
        ts_map = {}
        if oids:
            tbl_ts = _ol_table_name()
            sql_ts = text(f"""
                SELECT order_id, MAX(printed_notenough_at) AS ts 
                FROM {tbl_ts}
                WHERE printed_notenough > 0
                  AND order_id IN :oids
                GROUP BY order_id
            """).bindparams(bindparam("oids", expanding=True))
            try:
                rows_ts = db.session.execute(sql_ts, {"oids": oids}).all()
                for row_ts in rows_ts:
                    if not row_ts or not row_ts[0] or not row_ts[1]:
                        continue
                    oid_str = str(row_ts[0]).strip()
                    ts_str = row_ts[1]
                    try:
                        dt = datetime.fromisoformat(ts_str)
                        if dt.tzinfo is None:
                            dt = TH_TZ.localize(dt)
                        ts_map[oid_str] = dt
                    except:
                        pass
            except:
                pass

        for r in out:
            oid = (r.get("order_no") or "").strip()
            r["printed_at"] = ts_map.get(oid)

        final_oids = sorted({(r["order_no"] or "").strip() for r in out if r["order_no"]})
        skus = {(r["sku"] or "").strip() for r in out if r["sku"]}
        summary = {
            "sku_count": len(skus),
            "orders_count": len(final_oids),
        }
        
        logistics = sorted(set([r.get("shipping_type") for r in out if r.get("shipping_type")]))
        available_rounds = sorted({r["assign_round"] for r in out if r["assign_round"] is not None})

        # [SCAN] ดึงข้อมูลการ Scan Order เพื่อส่งไปหน้าเว็บ
        if final_oids:
            tbl = _ol_table_name()
            sql_scan = text(f"SELECT order_id, MAX(scanned_at) FROM {tbl} WHERE order_id IN :oids GROUP BY order_id")
            sql_scan = sql_scan.bindparams(bindparam("oids", expanding=True))
            res_scan = db.session.execute(sql_scan, {"oids": final_oids}).fetchall()
            scan_map = {str(r[0]): r[1] for r in res_scan if r[0]}
            for r in out:
                oid = (r.get("order_no") or "").strip()
                r["scanned_at"] = scan_map.get(oid)

        return render_template(
            "report_notenough.html",
            rows=out,
            summary=summary,
            printed_at=None,
            order_ids=final_oids,
            shops=shops,
            logistics=logistics,
            platform_sel=platform,
            shop_sel=shop_id,
            logistic_sel=logistic,
            is_history_view=True,
            available_dates=_available_dates(),
            print_date_from=print_date_from,
            print_date_to=print_date_to,
            sort_col=sort_col,
            sort_dir=sort_dir,
            q=q,
            round_sel=round_num,
            date_from=date_from_str,
            date_to=date_to_str,
            import_from=import_from_str,
            import_to=import_to_str
        )

    @app.route("/report/notenough/update_round", methods=["POST"])
    @login_required
    def update_notenough_round():
        """อัปเดตรอบสำหรับรายงานสินค้าไม่พอส่ง"""
        data = request.get_json() or {}
        order_ids = data.get("order_ids", [])
        round_num = data.get("round")
        
        if not order_ids or round_num is None:
            return jsonify({"success": False, "message": "ข้อมูลไม่ครบ"})
        
        try:
            round_int = int(round_num)
        except:
            return jsonify({"success": False, "message": "รอบต้องเป็นตัวเลข"})
        
        tbl = _ol_table_name()
        sql = text(f"UPDATE {tbl} SET notenough_round = :r WHERE order_id IN :oids")
        sql = sql.bindparams(bindparam("oids", expanding=True))
        db.session.execute(sql, {"r": round_int, "oids": order_ids})
        db.session.commit()
        
        return jsonify({"success": True, "message": f"อัปเดตรอบเป็น {round_int} สำเร็จ ({len(order_ids)} ออเดอร์)"})

    @app.route("/report/notenough.xlsx", methods=["GET"])
    @login_required
    def report_notenough_export():
        """Export Excel รายงานสินค้าไม่พอส่ง"""
        # ไม่ต้องใช้ services.lowstock แล้ว
        import pandas as pd
        
        platform = normalize_platform(request.args.get("platform"))
        shop_id = request.args.get("shop_id")
        logistic = request.args.get("logistic")
        q = (request.args.get("q") or "").strip()
        round_num = request.args.get("round")
        
        # รับค่าวันที่กรอง (เพิ่มใหม่)
        date_from_str = request.args.get("date_from")
        date_to_str = request.args.get("date_to")
        import_from_str = request.args.get("import_from")
        import_to_str = request.args.get("import_to")
        
        filters = {"platform": platform or None, "shop_id": int(shop_id) if shop_id else None, "import_date": None}
        rows, _ = compute_allocation(db.session, filters)
        rows = _filter_out_cancelled_rows(rows)
        rows = _filter_out_issued_rows(rows)
        
        packed_oids = _orders_packed_set(rows)
        
        safe = []
        for r in rows:
            r = dict(r)
            if (r.get("order_id") or "").strip() in packed_oids:
                continue
            if (str(r.get("sales_status") or "")).upper() == "PACKED":
                continue
            if "stock_qty" not in r:
                sku = (r.get("sku") or "").strip()
                stock_qty = 0
                if sku:
                    prod = Product.query.filter_by(sku=sku).first()
                    if prod and hasattr(prod, "stock_qty"):
                        try:
                            stock_qty = int(prod.stock_qty or 0)
                        except:
                            stock_qty = 0
                r["stock_qty"] = stock_qty
            # ไม่ต้อง _recompute_allocation_row(r) เพราะ compute_allocation คำนวณให้แล้ว
            safe.append(r)
        
        # กรองเฉพาะ allocation_status == "NOT_ENOUGH"
        lines = [r for r in safe if r.get("allocation_status") == "NOT_ENOUGH"]
        
        # ---------- กรอง logistic ----------
        if logistic:
            lines = [r for r in lines if (r.get("logistic") or "").strip().upper() == logistic.strip().upper()]
        
        # ---------- กรอง round ----------
        if round_num:
            try:
                round_int = int(round_num)
                lines = [r for r in lines if r.get("notenough_round") == round_int]
            except:
                pass
        
        # ---------- กรองด้วย date filters (เพิ่มใหม่) ----------
        def _parse_date(d_str):
            if not d_str:
                return None
            from utils import parse_datetime_guess
            dt = parse_datetime_guess(d_str)
            if dt:
                return dt.date() if hasattr(dt, 'date') else dt
            return None
        
        date_from = _parse_date(date_from_str)
        date_to = _parse_date(date_to_str)
        import_from = _parse_date(import_from_str)
        import_to = _parse_date(import_to_str)
        
        # กรองด้วยวันที่สั่ง (order_time / due_date)
        if date_from or date_to:
            filtered = []
            for r in lines:
                order_time = r.get("order_time") or r.get("due_date")
                if not order_time:
                    continue
                try:
                    from utils import parse_datetime_guess
                    dt = parse_datetime_guess(order_time)
                    if dt:
                        dt_date = dt.date() if hasattr(dt, 'date') else dt
                        if date_from and dt_date < date_from:
                            continue
                        if date_to and dt_date > date_to:
                            continue
                        filtered.append(r)
                except:
                    continue
            lines = filtered
        
        # กรองด้วยวันที่นำเข้า (import_date)
        if import_from or import_to:
            filtered = []
            for r in lines:
                imp_date = r.get("import_date")
                if not imp_date:
                    continue
                try:
                    from utils import parse_datetime_guess
                    dt = parse_datetime_guess(imp_date)
                    if dt:
                        dt_date = dt.date() if hasattr(dt, 'date') else dt
                        if import_from and dt_date < import_from:
                            continue
                        if import_to and dt_date > import_to:
                            continue
                        filtered.append(r)
                except:
                    continue
            lines = filtered
        
        # ---------- กรองด้วยคำค้นหา q ----------
        if q:
            q_lower = q.lower()
            lines = [r for r in lines if q_lower in (r.get("sku") or "").lower() 
                     or q_lower in (r.get("model") or "").lower() 
                     or q_lower in (r.get("order_id") or "").lower()]
        
        df = pd.DataFrame([{
            "แพลตฟอร์ม": r.get("platform"),
            "ร้าน": r.get("shop"),
            "เลข Order": r.get("order_id"),
            "SKU": r.get("sku"),
            "Brand": r.get("brand"),
            "ชื่อสินค้า": r.get("model"),
            "Stock": int(r.get("stock_qty", 0) or 0),
            "Qty": int(r.get("qty", 0) or 0),
            "เวลาที่ลูกค้าสั่ง": r.get("order_time"),
            "กำหนดส่ง": r.get("due_date"),
            "ประเภทขนส่ง": r.get("logistic"),
        } for r in lines])
        
        out = BytesIO()
        with pd.ExcelWriter(out, engine="xlsxwriter") as w:
            df.to_excel(w, index=False, sheet_name="NotEnough")
        out.seek(0)
        return send_file(out, as_attachment=True, download_name="report_notenough.xlsx")
    # ================== /NEW: Report Not Enough ==================

    # -----------------------
    # Picking (รวมยอดหยิบ)
    # -----------------------
    def _aggregate_picking(rows: list[dict], group_by_round: bool = False) -> list[dict]:
        """
        รวมยอดหยิบตาม SKU
        - group_by_round=False: รวมทุกรอบเข้าด้วยกัน (default เดิม)
        - group_by_round=True: แยก key ตาม SKU+dispatch_round (ไม่รวมข้ามรอบ)
        """
        rows = rows or []
        agg: dict[str, dict] = {}
        for r in rows:
            if not bool(r.get("accepted")):
                continue
            # [แก้ไข] เพิ่ม "ISSUED" เพื่อให้หน้าประวัติ (ที่จ่ายงานแล้ว) แสดงข้อมูลได้
            if (r.get("allocation_status") or "") not in ("ACCEPTED", "READY_ACCEPT", "ISSUED"):
                continue
            sku = str(r.get("sku") or "").strip()
            if not sku:
                continue
            brand = str(r.get("brand") or "").strip()
            model = str(r.get("model") or "").strip()
            qty = int(r.get("qty", 0) or 0)
            stock_qty = int(r.get("stock_qty", 0) or 0)
            dispatch_round = r.get("dispatch_round")
            
            # [NEW] สร้าง key ที่รวม dispatch_round ด้วย (ถ้า group_by_round=True)
            if group_by_round and dispatch_round is not None:
                agg_key = f"{sku}__round_{dispatch_round}"
            else:
                agg_key = sku
            
            a = agg.setdefault(agg_key, {
                "sku": sku, 
                "brand": brand, 
                "model": model, 
                "need_qty": 0, 
                "stock_qty": 0,
                "dispatch_rounds": set(),
                "dispatch_round_single": dispatch_round  # เก็บค่าเดี่ยวไว้
            })
            a["need_qty"] += qty
            if stock_qty > a["stock_qty"]:
                a["stock_qty"] = stock_qty
            if dispatch_round is not None:
                a["dispatch_rounds"].add(dispatch_round)

        items = []
        for _, a in agg.items():
            need = int(a["need_qty"])
            stock = int(a["stock_qty"])
            shortage = max(0, need - stock)
            remain = stock - need
            
            # Handle dispatch_round display
            dispatch_rounds = sorted(a["dispatch_rounds"])
            if len(dispatch_rounds) == 0:
                dispatch_round_display = None
            elif len(dispatch_rounds) == 1:
                dispatch_round_display = dispatch_rounds[0]
            else:
                # ถ้า group_by_round=True ไม่ควรมีกรณีนี้ แต่ fallback ไว้
                dispatch_round_display = f"{dispatch_rounds[0]}-{dispatch_rounds[-1]}"
            
            items.append({
                "sku": a["sku"], 
                "brand": a["brand"], 
                "model": a["model"],
                "need_qty": need, 
                "stock_qty": stock, 
                "shortage": shortage, 
                "remain_after_pick": remain,
                "dispatch_round": dispatch_round_display,
            })
        items.sort(key=lambda x: (x["brand"].lower(), x["model"].lower(), x["sku"].lower()))
        return items

    @app.route("/report/picking", methods=["GET"])
    @login_required
    def picking_list():
        # Check for reset mode
        reset_mode = request.args.get("reset")
        
        if reset_mode == 'all':
            # Clear all filters and show all pending orders
            platform = None
            shop_id = None
            logistic = None
            acc_from = None
            acc_to = None
            acc_from_str = ""
            acc_to_str = ""
            round_sel = None
            print_count_sel = None
        else:
            platform = normalize_platform(request.args.get("platform"))
            shop_id = request.args.get("shop_id")
            logistic = request.args.get("logistic")
            
            # รับค่าวันที่กดพร้อมรับ (accepted_at)
            acc_from_str = request.args.get("accepted_from")
            acc_to_str = request.args.get("accepted_to")
            acc_from = parse_date_any(acc_from_str)
            acc_to = parse_date_any(acc_to_str)
            
            # [NEW] รับค่ารอบจ่ายงาน และจำนวนครั้งที่พิมพ์
            round_sel = request.args.get("round")
            print_count_sel = request.args.get("print_count")

        # [แก้ไข] ไม่กรอง accepted date ใน compute_allocation
        # เราจะกรองด้วย printed_warehouse_at เองทีหลัง
        filters = {
            "platform": platform if platform else None, 
            "shop_id": int(shop_id) if shop_id else None, 
            "import_date": None,
            "accepted_from": None,  # ไม่กรองตรงนี้
            "accepted_to": None,    # ไม่กรองตรงนี้
        }
        rows, _ = compute_allocation(db.session, filters)
        rows = _filter_out_cancelled_rows(rows)

        # ========================================================
        # [FIXED LOGIC] กรอง Order ที่ "พิมพ์คลังแล้ว" แต่ "ยังไม่พิมพ์หยิบ"
        # + กรองด้วยวันที่พิมพ์ใบงานคลัง (printed_warehouse_at)
        # ========================================================
        
        # 1. รวบรวม Order ID ทั้งหมดในหน้านี้
        all_oids = sorted({(r.get("order_id") or "").strip() for r in rows if r.get("order_id")})
        
        valid_rows = []
        
        if all_oids:
            tbl = _ol_table_name()
            # 2. Query เช็คสถานะการพิมพ์จาก DB โดยตรง (แม่นยำกว่า)
            # ดึงจำนวนครั้งที่พิมพ์ Warehouse และ Picking + เวลาที่พิมพ์ Warehouse
            sql = text(f"""
                SELECT order_id, 
                       MAX(COALESCE(printed_warehouse, 0)) as wh_count, 
                       MAX(COALESCE(printed_picking, 0)) as pk_count,
                       MAX(printed_warehouse_at) as wh_at
                FROM {tbl} 
                WHERE order_id IN :oids 
                GROUP BY order_id
            """)
            sql = sql.bindparams(bindparam("oids", expanding=True))
            
            print_status = db.session.execute(sql, {"oids": all_oids}).fetchall()
            
            # สร้าง Map {order_id: (wh_count, pk_count, wh_at_str)}
            status_map = {}
            for row in print_status:
                status_map[row[0]] = (int(row[1] or 0), int(row[2] or 0), row[3])
            
            # แปลงวันที่กรองเป็น datetime เพื่อเปรียบเทียบ
            f_start = datetime.combine(acc_from, datetime.min.time(), tzinfo=TH_TZ) if acc_from else None
            f_end = datetime.combine(acc_to + timedelta(days=1), datetime.min.time(), tzinfo=TH_TZ) if acc_to else None
            
            for r in rows:
                oid = (r.get("order_id") or "").strip()
                wh, pk, wh_at_str = status_map.get(oid, (0, 0, None))
                
                # เงื่อนไขสำคัญ: 
                # 1. ต้องพิมพ์คลังแล้ว (wh > 0)
                # 2. ต้องยังไม่พิมพ์หยิบ (pk == 0)
                if wh > 0 and pk == 0:
                    # เงื่อนไข 3: กรองวันที่พิมพ์ใบงานคลัง (ถ้ามีการกรอง)
                    pass_date = True
                    if f_start or f_end:
                        if not wh_at_str:
                            pass_date = False  # ไม่มีวันที่พิมพ์ = ไม่ผ่าน
                        else:
                            try:
                                dt_print = datetime.fromisoformat(wh_at_str)
                                if dt_print.tzinfo is None:
                                    dt_print = TH_TZ.localize(dt_print)
                                if f_start and dt_print < f_start:
                                    pass_date = False
                                if f_end and dt_print >= f_end:
                                    pass_date = False
                            except Exception:
                                pass_date = False
                    
                    if pass_date:
                        valid_rows.append(r)
            
        rows = valid_rows

        # เตรียมข้อมูลปลอดภัย + ใส่ stock_qty ให้ครบ
        safe_rows = []
        for r in rows:
            r = dict(r)
            if "stock_qty" not in r:
                sku = (r.get("sku") or "").strip()
                stock_qty = 0
                if sku:
                    prod = Product.query.filter_by(sku=sku).first()
                    if prod and hasattr(prod, "stock_qty"):
                        try:
                            stock_qty = int(prod.stock_qty or 0)
                        except Exception:
                            stock_qty = 0
                    else:
                        st = Stock.query.filter_by(sku=sku).first()
                        stock_qty = int(st.qty) if st and st.qty is not None else 0
                r["stock_qty"] = stock_qty
            r["accepted"] = bool(r.get("accepted", False))
            r["sales_status"] = r.get("sales_status", None)
            r["logistic"] = r.get("logistic") or r.get("logistic_type") or "-"
            safe_rows.append(r)

        if logistic:
            safe_rows = [r for r in safe_rows if (r.get("logistic") or "").lower().find(logistic.lower()) >= 0]
        
        # [NEW] กรองตามรอบจ่ายงาน (dispatch_round)
        if round_sel:
            try:
                round_int = int(round_sel)
                safe_rows = [r for r in safe_rows if r.get("dispatch_round") == round_int]
            except (ValueError, TypeError):
                pass
        
        # [NEW] กรองตามจำนวนครั้งที่พิมพ์ - ต้องดึงข้อมูล print count ของแต่ละ order ก่อน
        if print_count_sel:
            try:
                target_pc = int(print_count_sel)
                # ดึง order_ids ที่มี print count ตามเงื่อนไข
                temp_oids = sorted({(r.get("order_id") or "").strip() for r in safe_rows if r.get("order_id")})
                if temp_oids:
                    pc_map = _get_print_counts_local(temp_oids, "picking")
                    # กรองเฉพาะ order ที่มี print count ตรงกับที่ระบุ
                    valid_oids = {oid for oid, cnt in pc_map.items() if cnt == target_pc}
                    safe_rows = [r for r in safe_rows if (r.get("order_id") or "").strip() in valid_oids]
            except (ValueError, TypeError):
                pass

        # รวมต่อ SKU
        items = _aggregate_picking(safe_rows)

        # ===== นับจำนวนครั้งที่พิมพ์ Picking (รวมทั้งชุดงาน) — ใช้ MAX ไม่ใช่ SUM =====
        valid_rows = [r for r in safe_rows if r.get("accepted") and r.get("allocation_status") in ("ACCEPTED", "READY_ACCEPT")]
        order_ids = sorted({(r.get("order_id") or "").strip() for r in valid_rows if r.get("order_id")})
        print_counts_pick = _get_print_counts_local(order_ids, "picking")
        print_count_overall = max(print_counts_pick.values()) if print_counts_pick else 0
        
        # Get the latest print timestamp and user
        print_timestamp_overall = None
        print_user_overall = None
        if order_ids:
            tbl = _ol_table_name()
            sql = text(f"SELECT printed_picking_at, printed_picking_by FROM {tbl} WHERE order_id IN :oids AND printed_picking_at IS NOT NULL ORDER BY printed_picking_at DESC LIMIT 1")
            sql = sql.bindparams(bindparam("oids", expanding=True))
            result = db.session.execute(sql, {"oids": order_ids}).first()
            if result:
                try:
                    dt = datetime.fromisoformat(result[0])
                    if dt.tzinfo is None:
                        dt = TH_TZ.localize(dt)
                    print_timestamp_overall = dt
                    print_user_overall = result[1]  # username
                except Exception:
                    pass

        # ชื่อร้านสำหรับแสดงในคอลัมน์ใหม่
        shop_sel_name = None
        if shop_id:
            s = Shop.query.get(int(shop_id))
            if s:
                shop_sel_name = f"{s.platform} • {s.name}"

        # เติมแพลตฟอร์ม/ร้าน/ประเภทขนส่งให้แต่ละ item เพื่อไม่ให้ขึ้น '-'
        for it in items:
            it["platform"] = platform or "-"
            it["shop"] = shop_sel_name or "-"
            it["logistic"] = logistic or "-"
        
        # ===== ดึงข้อมูลการเชื่อมโยงกับใบงานคลัง (Warehouse) =====
        # เพื่อแสดงว่า Picking ใบนี้ผูกกับใบงานคลังที่พิมพ์เมื่อไหร่
        warehouse_print_info = None
        if order_ids:
            tbl = _ol_table_name()
            sql = text(f"""
                SELECT printed_warehouse_at, printed_warehouse_by, printed_warehouse
                FROM {tbl} 
                WHERE order_id IN :oids 
                AND printed_warehouse > 0 
                ORDER BY printed_warehouse_at DESC 
                LIMIT 1
            """)
            sql = sql.bindparams(bindparam("oids", expanding=True))
            result = db.session.execute(sql, {"oids": order_ids}).first()
            if result and result[0]:
                try:
                    dt = datetime.fromisoformat(result[0])
                    if dt.tzinfo is None:
                        dt = TH_TZ.localize(dt)
                    warehouse_print_info = {
                        "printed_at": dt,
                        "printed_by": result[1],
                        "print_count": result[2]
                    }
                except Exception:
                    pass

        totals = {
            "total_skus": len(items),
            "total_need_qty": sum(i["need_qty"] for i in items),
            "total_shortage": sum(i["shortage"] for i in items),
        }
        shops = Shop.query.order_by(Shop.name.asc()).all()
        logistics = sorted(set(r.get("logistic") for r in safe_rows if r.get("logistic")))

        return render_template(
            "picking.html",
            items=items,
            totals=totals,
            shops=shops,
            logistics=logistics,
            platform_sel=platform if reset_mode != 'all' else None,
            shop_sel=shop_id if reset_mode != 'all' else None,
            shop_sel_name=shop_sel_name if reset_mode != 'all' else None,
            logistic_sel=logistic if reset_mode != 'all' else None,
            official_print=False,
            printed_meta=None,
            print_count_overall=print_count_overall,
            print_timestamp_overall=print_timestamp_overall,
            print_user_overall=print_user_overall,
            order_ids=order_ids,  # Pass order IDs for dispatch round update
            accepted_from=acc_from_str if reset_mode != 'all' else "",
            accepted_to=acc_to_str if reset_mode != 'all' else "",
            is_history_view=False,
            warehouse_print_info=warehouse_print_info,  # เชื่อมโยงกับใบงานคลัง
            round_sel=round_sel if reset_mode != 'all' else None,  # [NEW] ส่งค่ารอบจ่ายงาน
            print_count_sel=print_count_sel if reset_mode != 'all' else None,  # [NEW] ส่งค่าจำนวนครั้งที่พิมพ์
            available_rounds=[],  # [NEW] สำหรับหน้าปัจจุบันไม่มี dropdown รอบ
        )

    @app.route("/report/picking/print", methods=["POST"])
    @login_required
    def picking_list_commit():
        cu = current_user()
        platform = normalize_platform(request.form.get("platform"))
        shop_id = request.form.get("shop_id")
        logistic = request.form.get("logistic")
        override = request.form.get("override") in ("1", "true", "yes")
        
        # Get selected order IDs from form (comma-separated)
        # ถ้าเป็น '', 'all', 'ALL' ให้ถือว่า "ไม่ระบุ"
        order_ids_raw = (request.form.get("order_ids") or "").strip()
        selected_order_ids = [] if order_ids_raw.lower() in ("", "all") else \
            [oid.strip() for oid in order_ids_raw.split(",") if oid.strip()]

        filters = {"platform": platform if platform else None, "shop_id": int(shop_id) if shop_id else None, "import_date": None}
        rows, _ = compute_allocation(db.session, filters)
        rows = _filter_out_cancelled_rows(rows)

        safe_rows = []
        for r in rows:
            r = dict(r)
            if "stock_qty" not in r:
                sku = (r.get("sku") or "").strip()
                stock_qty = 0
                if sku:
                    prod = Product.query.filter_by(sku=sku).first()
                    if prod and hasattr(prod, "stock_qty"):
                        try:
                            stock_qty = int(prod.stock_qty or 0)
                        except Exception:
                            stock_qty = 0
                    else:
                        st = Stock.query.filter_by(sku=sku).first()
                        stock_qty = int(st.qty) if st and st.qty is not None else 0
                r["stock_qty"] = stock_qty
            r["accepted"] = bool(r.get("accepted", False))
            r["logistic"] = r.get("logistic") or r.get("logistic_type") or "-"
            safe_rows.append(r)

        if logistic:
            safe_rows = [r for r in safe_rows if (r.get("logistic") or "").lower().find(logistic.lower()) >= 0]

        valid_rows = [r for r in safe_rows if r.get("accepted") and r.get("allocation_status") in ("ACCEPTED", "READY_ACCEPT")]
        
        # If specific order IDs were selected, filter to only those
        if selected_order_ids:
            valid_rows = [r for r in valid_rows if (r.get("order_id") or "").strip() in selected_order_ids]
            oids = sorted(selected_order_ids)
        else:
            oids = sorted({(r.get("order_id") or "").strip() for r in valid_rows if r.get("order_id")})

        if not oids:
            flash("ไม่พบออเดอร์สำหรับพิมพ์ Picking", "warning")
            return redirect(url_for("picking_list", platform=platform, shop_id=shop_id, logistic=logistic))

        already = _detect_already_printed(oids, kind="picking")
        if already and not (override and cu and cu.role == "admin"):
            head = ", ".join(list(already)[:10])
            more = "" if len(already) <= 10 else f" ... (+{len(already)-10})"
            flash(f"มีบางออเดอร์เคยพิมพ์ Picking ไปแล้ว: {head}{more}", "danger")
            flash("ถ้าจำเป็นต้องพิมพ์ซ้ำ โปรดให้แอดมินติ๊ก 'อนุญาตพิมพ์ซ้ำ' แล้วพิมพ์อีกครั้ง", "warning")
            return redirect(url_for("picking_list", platform=platform, shop_id=shop_id, logistic=logistic))

        now_iso = now_thai().isoformat()
        _mark_printed(oids, kind="picking", user_id=(cu.id if cu else None), when_iso=now_iso)
        
        # >>> NEW: ย้ายไป Orderจ่ายแล้ว (บันทึกเวลาตอนพิมพ์)
        _mark_issued(oids, user_id=(cu.id if cu else None), source="print:picking", when_dt=now_thai())
        
        db.session.commit()  # Ensure changes are committed
        db.session.expire_all()  # Force refresh to get updated print counts

        items = _aggregate_picking(safe_rows)
        for it in items:
            it["platform"] = platform or "-"
            if shop_id:
                s = Shop.query.get(int(shop_id))
                it["shop"] = (f"{s.platform} • {s.name}") if s else "-"
            else:
                it["shop"] = "-"
            it["logistic"] = logistic or "-"

        totals = {
            "total_skus": len(items),
            "total_need_qty": sum(i["need_qty"] for i in items),
            "total_shortage": sum(i["shortage"] for i in items),
        }
        shops = Shop.query.order_by(Shop.name.asc()).all()
        logistics = sorted(set(r.get("logistic") for r in safe_rows if r.get("logistic")))
        printed_meta = {"by": (cu.username if cu else "-"), "at": now_thai(), "orders": len(oids), "override": bool(already)}

        print_counts_pick = _get_print_counts_local(oids, "picking")
        print_count_overall = max(print_counts_pick.values()) if print_counts_pick else 0
        
        # Use current timestamp and user
        print_timestamp_overall = now_thai()
        print_user_overall = cu.username if cu else None

        shop_sel_name = None
        if shop_id:
            s = Shop.query.get(int(shop_id))
            if s:
                shop_sel_name = f"{s.platform} • {s.name}"

        return render_template(
            "picking.html",
            items=items,
            totals=totals,
            shops=shops,
            logistics=logistics,
            platform_sel=platform,
            shop_sel=shop_id,
            shop_sel_name=shop_sel_name,
            logistic_sel=logistic,
            official_print=True,
            printed_meta=printed_meta,
            print_count_overall=print_count_overall,
            print_timestamp_overall=print_timestamp_overall,
            print_user_overall=print_user_overall,
            order_ids=oids,  # Pass order IDs for dispatch round update
        )

    # ================== NEW: Update Dispatch Round from Picking ==================
    @app.route("/picking/update_dispatch", methods=["POST"])
    @login_required
    def picking_update_dispatch():
        """อัปเดตเลขรอบจ่ายงานจากหน้า Picking List และทำเครื่องหมายจ่ายงานแล้ว"""
        cu = current_user()
        if not cu:
            flash("กรุณาเข้าสู่ระบบก่อน", "warning")
            return redirect(url_for("login"))
        
        try:
            # รับค่าจาก Form
            order_ids = request.form.getlist("order_ids[]")
            dispatch_round = request.form.get("dispatch_round", type=int) or 1
            
            if not order_ids:
                flash("ไม่พบรายการที่เลือก", "warning")
                return redirect(request.referrer or url_for("picking_list"))
            
            # 1. บันทึกสถานะ Issued (จ่ายงานแล้ว)
            source = f"picking:round_{dispatch_round}"
            _mark_issued(order_ids, cu.id, source=source)
            
            # 2. อัปเดตเลขรอบ (dispatch_round) ลง DB
            tbl = _ol_table_name()
            sql = text(f"UPDATE {tbl} SET dispatch_round = :r WHERE order_id IN :oids")
            sql = sql.bindparams(bindparam("oids", expanding=True))
            db.session.execute(sql, {"r": dispatch_round, "oids": order_ids})
            db.session.commit()
            
            flash(f"✅ จ่ายงาน {len(order_ids)} ออเดอร์ เป็นรอบที่ {dispatch_round} เรียบร้อยแล้ว", "success")
            
        except Exception as e:
            db.session.rollback()
            flash(f"เกิดข้อผิดพลาด: {e}", "danger")
        
        # Redirect กลับหน้าเดิมพร้อมฟิลเตอร์
        platform = request.form.get("platform") or ""
        shop_id = request.form.get("shop_id") or ""
        logistic = request.form.get("logistic") or ""
        return redirect(url_for("picking_list", platform=platform, shop_id=shop_id, logistic=logistic))
    # ================== /NEW ==================

    # ================== NEW: View Printed Picking Lists ==================
    @app.route("/report/picking/printed", methods=["GET"])
    @login_required
    def picking_printed_history():
        """ดู Picking List ที่พิมพ์แล้ว - สามารถเลือกวันที่และพิมพ์ซ้ำได้"""
        # Check for reset mode
        reset_mode = request.args.get("reset")
        target_date = None
        
        # ตรวจสอบว่ามีการส่งพารามิเตอร์มาบ้างไหม (เพื่อดูว่าเป็น First Load หรือไม่)
        has_params = any([
            request.args.get("platform"),
            request.args.get("shop_id"),
            request.args.get("logistic"),
            request.args.get("print_date"),
            request.args.get("accepted_from"),
            request.args.get("accepted_to"),
            request.args.get("round"),
            request.args.get("print_count"),  # [NEW]
            request.args.get("reset")
        ])
        
        if reset_mode == 'today' or not has_params:
            # Reset หรือ เข้ามาครั้งแรก (ไม่ส่ง param) -> เอาของ "วันนี้"
            target_date = now_thai().date()
            platform = None
            shop_id = None
            logistic = None
            print_date = None
            raw_from = None
            raw_to = None
            round_sel = None  # [NEW]
            print_count_sel = None  # [NEW]
        else:
            platform = normalize_platform(request.args.get("platform"))
            shop_id = request.args.get("shop_id")
            logistic = request.args.get("logistic")
            print_date = request.args.get("print_date")  # วันที่พิมพ์ (YYYY-MM-DD)
            
            # รับค่าวันที่กดพร้อมรับ - ไม่ตั้งค่า default
            raw_from = request.args.get("accepted_from")
            raw_to = request.args.get("accepted_to")
            
            # [NEW] รับค่ารอบจ่ายงาน และจำนวนครั้งที่พิมพ์
            round_sel = request.args.get("round")
            print_count_sel = request.args.get("print_count")
            
            if print_date:
                try:
                    target_date = datetime.strptime(print_date, "%Y-%m-%d").date()
                except:
                    target_date = None
        
        # ไม่ตั้งค่า default - ให้เป็นค่าว่าง (mm/dd/yyyy)
        acc_from = parse_date_any(raw_from)
        acc_to = parse_date_any(raw_to)
        
        # Get all orders that have been printed for picking
        tbl = _ol_table_name()
        
        # Build query to get orders with print history
        if target_date:
            # Filter by specific print date (or today if reset)
            # หมายเหตุ: printed_picking_at ถูกบันทึกเป็นเวลาไทยอยู่แล้ว (ไม่ต้อง +7)
            sql = text(f"""
                SELECT DISTINCT order_id 
                FROM {tbl} 
                WHERE printed_picking > 0 
                AND DATE(printed_picking_at) = :target_date
            """)
            result = db.session.execute(sql, {"target_date": target_date.isoformat()}).fetchall()
        else:
            # Get all printed orders
            sql = text(f"SELECT DISTINCT order_id FROM {tbl} WHERE printed_picking > 0")
            result = db.session.execute(sql).fetchall()
        
        printed_order_ids = [row[0] for row in result if row[0]]
        
        if not printed_order_ids:
            # No printed orders found
            shops = Shop.query.order_by(Shop.name.asc()).all()
            return render_template(
                "picking.html",
                items=[],
                totals={"total_skus": 0, "total_need_qty": 0, "total_shortage": 0},
                shops=shops,
                logistics=[],
                platform_sel=platform,
                shop_sel=shop_id,
                shop_sel_name=None,
                logistic_sel=logistic,
                official_print=False,
                printed_meta=None,
                print_count_overall=0,
                print_timestamp_overall=None,
                order_ids=[],
                is_history_view=True,
                print_date_sel=print_date,
                available_dates=[],
                available_rounds=[],
                round_sel=round_sel,
                print_count_sel=print_count_sel,  # [NEW]
                accepted_from=raw_from,
                accepted_to=raw_to,
            )
        
        # Get full data for these orders
        # [แก้ไข] ไม่กรอง accepted date ใน compute_allocation - เราจะกรองด้วย printed_warehouse_at เอง
        filters = {
            "platform": platform if platform else None, 
            "shop_id": int(shop_id) if shop_id else None, 
            "import_date": None,
            "accepted_from": None,  # ไม่กรองตรงนี้
            "accepted_to": None,    # ไม่กรองตรงนี้
        }
        rows, _ = compute_allocation(db.session, filters)
        rows = _filter_out_cancelled_rows(rows)
        
        # [แก้ไข] ดึงเวลาพิมพ์ Warehouse มาเพื่อกรองด้วย printed_warehouse_at
        all_oids = sorted({(r.get("order_id") or "").strip() for r in rows if r.get("order_id")})
        wh_print_map = {}
        dispatch_round_map = {}  # [NEW] เก็บ dispatch_round ของแต่ละ order+sku
        if all_oids:
            sql = text(f"SELECT order_id, MAX(printed_warehouse_at) FROM {tbl} WHERE order_id IN :oids GROUP BY order_id")
            sql = sql.bindparams(bindparam("oids", expanding=True))
            res = db.session.execute(sql, {"oids": all_oids}).fetchall()
            for row in res:
                wh_print_map[row[0]] = row[1]
            
            # [NEW] ดึง dispatch_round แยกตาม order_id + sku (ระดับบรรทัด)
            sql_dr = text(f"SELECT order_id, sku, dispatch_round FROM {tbl} WHERE order_id IN :oids AND dispatch_round IS NOT NULL")
            sql_dr = sql_dr.bindparams(bindparam("oids", expanding=True))
            res_dr = db.session.execute(sql_dr, {"oids": all_oids}).fetchall()
            for row_dr in res_dr:
                key = (row_dr[0], row_dr[1])  # (order_id, sku)
                dispatch_round_map[key] = row_dr[2]
        
        # แปลงวันที่กรองเป็น datetime เพื่อเปรียบเทียบ
        f_start = datetime.combine(acc_from, datetime.min.time(), tzinfo=TH_TZ) if acc_from else None
        f_end = datetime.combine(acc_to + timedelta(days=1), datetime.min.time(), tzinfo=TH_TZ) if acc_to else None
        
        # Filter to only printed orders + กรองด้วยวันที่พิมพ์ Warehouse
        safe_rows = []
        for r in rows:
            oid = (r.get("order_id") or "").strip()
            if oid not in printed_order_ids:
                continue
            
            # [แก้ไข] กรองด้วยวันที่พิมพ์ Warehouse (ถ้ามีการกรอง)
            if f_start or f_end:
                wh_at_str = wh_print_map.get(oid)
                if not wh_at_str:
                    continue  # ไม่มีวันที่พิมพ์ Warehouse -> ข้าม
                try:
                    dt_print = datetime.fromisoformat(wh_at_str)
                    if dt_print.tzinfo is None:
                        dt_print = TH_TZ.localize(dt_print)
                    if f_start and dt_print < f_start:
                        continue
                    if f_end and dt_print >= f_end:
                        continue
                except Exception:
                    continue
            
            r = dict(r)
            if "stock_qty" not in r:
                sku = (r.get("sku") or "").strip()
                stock_qty = 0
                if sku:
                    prod = Product.query.filter_by(sku=sku).first()
                    if prod and hasattr(prod, "stock_qty"):
                        try:
                            stock_qty = int(prod.stock_qty or 0)
                        except Exception:
                            stock_qty = 0
                    else:
                        st = Stock.query.filter_by(sku=sku).first()
                        stock_qty = int(st.qty) if st and st.qty is not None else 0
                r["stock_qty"] = stock_qty
            r["accepted"] = bool(r.get("accepted", False))
            r["sales_status"] = r.get("sales_status", None)
            r["logistic"] = r.get("logistic") or r.get("logistic_type") or "-"
            
            # [NEW] ใส่ dispatch_round จาก DB (ระดับบรรทัด order_id + sku)
            sku_key = (oid, (r.get("sku") or "").strip())
            if sku_key in dispatch_round_map:
                r["dispatch_round"] = dispatch_round_map[sku_key]
            
            safe_rows.append(r)
        
        if logistic:
            safe_rows = [r for r in safe_rows if (r.get("logistic") or "").lower().find(logistic.lower()) >= 0]
        
        # [NEW] กรองตามรอบจ่ายงาน (dispatch_round) ถ้ามีการเลือก
        if round_sel:
            try:
                round_int = int(round_sel)
                safe_rows = [r for r in safe_rows if r.get("dispatch_round") == round_int]
            except (ValueError, TypeError):
                pass
        
        # [NEW] กรองตามจำนวนครั้งที่พิมพ์ - ต้องดึงข้อมูล print count ของแต่ละ order ก่อน
        if print_count_sel:
            try:
                target_pc = int(print_count_sel)
                # ดึง order_ids ที่มี print count ตามเงื่อนไข
                temp_oids = sorted({(r.get("order_id") or "").strip() for r in safe_rows if r.get("order_id")})
                if temp_oids:
                    pc_map = _get_print_counts_local(temp_oids, "picking")
                    # กรองเฉพาะ order ที่มี print count ตรงกับที่ระบุ
                    valid_oids = {oid for oid, cnt in pc_map.items() if cnt == target_pc}
                    safe_rows = [r for r in safe_rows if (r.get("order_id") or "").strip() in valid_oids]
            except (ValueError, TypeError):
                pass
        
        # [NEW] ดึงรายการรอบที่มีทั้งหมด (สำหรับ Dropdown)
        available_rounds = []
        try:
            rounds_sql = text(f"""
                SELECT DISTINCT dispatch_round 
                FROM {tbl} 
                WHERE printed_picking > 0 
                  AND dispatch_round IS NOT NULL
                ORDER BY dispatch_round ASC
            """)
            rounds_result = db.session.execute(rounds_sql).fetchall()
            available_rounds = [r[0] for r in rounds_result if r[0] is not None]
        except Exception:
            pass
        
        # Aggregate by SKU
        # [แก้ไข] ถ้ามีการเลือกรอบ ให้แยก aggregate ตามรอบด้วย (ไม่รวมข้ามรอบ)
        items = _aggregate_picking(safe_rows, group_by_round=bool(round_sel))
        
        # Get print counts
        # [แก้ไข] เพิ่ม "ISSUED" เพื่อให้หน้าประวัติ (ที่จ่ายงานแล้ว) นับ order ได้ถูกต้อง
        valid_rows = [r for r in safe_rows if r.get("accepted") and r.get("allocation_status") in ("ACCEPTED", "READY_ACCEPT", "ISSUED")]
        order_ids = sorted({(r.get("order_id") or "").strip() for r in valid_rows if r.get("order_id")})
        print_counts_pick = _get_print_counts_local(order_ids, "picking")
        print_count_overall = max(print_counts_pick.values()) if print_counts_pick else 0
        
        # Get the latest print timestamp and user
        print_timestamp_overall = None
        print_user_overall = None
        if order_ids:
            sql = text(f"SELECT printed_picking_at, printed_picking_by FROM {tbl} WHERE order_id IN :oids AND printed_picking_at IS NOT NULL ORDER BY printed_picking_at DESC LIMIT 1")
            sql = sql.bindparams(bindparam("oids", expanding=True))
            result = db.session.execute(sql, {"oids": order_ids}).first()
            if result:
                try:
                    dt = datetime.fromisoformat(result[0])
                    if dt.tzinfo is None:
                        dt = TH_TZ.localize(dt)
                    print_timestamp_overall = dt
                    print_user_overall = result[1]
                except Exception:
                    pass
        
        # Shop name
        shop_sel_name = None
        if shop_id:
            s = Shop.query.get(int(shop_id))
            if s:
                shop_sel_name = f"{s.platform} • {s.name}"
        
        # Fill in platform/shop/logistic for each item
        for it in items:
            it["platform"] = platform or "-"
            it["shop"] = shop_sel_name or "-"
            it["logistic"] = logistic or "-"
        
        totals = {
            "total_skus": len(items),
            "total_need_qty": sum(i["need_qty"] for i in items),
            "total_shortage": sum(i["shortage"] for i in items),
        }
        shops = Shop.query.order_by(Shop.name.asc()).all()
        logistics = sorted(set(r.get("logistic") for r in safe_rows if r.get("logistic")))
        
        # Get available print dates for dropdown
        # หมายเหตุ: printed_picking_at ถูกบันทึกเป็นเวลาไทยอยู่แล้ว (ไม่ต้อง +7)
        sql_dates = text(f"""
            SELECT DISTINCT DATE(printed_picking_at) as print_date 
            FROM {tbl} 
            WHERE printed_picking > 0 AND printed_picking_at IS NOT NULL
            ORDER BY print_date DESC
        """)
        available_dates = [row[0] for row in db.session.execute(sql_dates).fetchall()]
        
        return render_template(
            "picking.html",
            items=items,
            totals=totals,
            shops=shops,
            logistics=logistics,
            platform_sel=platform if reset_mode != 'today' else None,
            shop_sel=shop_id if reset_mode != 'today' else None,
            shop_sel_name=shop_sel_name if reset_mode != 'today' else None,
            logistic_sel=logistic if reset_mode != 'today' else None,
            official_print=False,
            printed_meta=None,
            print_count_overall=print_count_overall,
            print_timestamp_overall=print_timestamp_overall,
            print_user_overall=print_user_overall,
            order_ids=order_ids,
            is_history_view=True,
            # ถ้าเป็น Default/Today ให้ส่งค่า print_date_sel เป็นวันนี้ เพื่อให้ Dropdown เลือกถูก
            print_date_sel=print_date if print_date else (target_date.isoformat() if target_date else None),
            available_dates=available_dates,
            available_rounds=available_rounds,
            round_sel=round_sel if reset_mode != 'today' else None,
            print_count_sel=print_count_sel if reset_mode != 'today' else None,  # [NEW] ส่งค่าจำนวนครั้งที่พิมพ์
            accepted_from=raw_from if reset_mode != 'today' else "",
            accepted_to=raw_to if reset_mode != 'today' else "",
        )

    @app.route("/export_picking.xlsx")
    @login_required
    def export_picking_excel():
        """Export ใบงานหยิบสินค้าปัจจุบัน - แสดงงานที่ยังไม่ได้พิมพ์"""
        # Check for reset mode
        reset_mode = request.args.get("reset")
        
        if reset_mode == 'all':
            platform = None
            shop_id = None
            logistic = None
            acc_from = None
            acc_to = None
        else:
            platform = normalize_platform(request.args.get("platform"))
            shop_id = request.args.get("shop_id")
            logistic = request.args.get("logistic")
            acc_from = parse_date_any(request.args.get("accepted_from"))
            acc_to = parse_date_any(request.args.get("accepted_to"))

        filters = {
            "platform": platform if platform else None, 
            "shop_id": int(shop_id) if shop_id else None, 
            "import_date": None,
            "accepted_from": datetime.combine(acc_from, datetime.min.time(), tzinfo=TH_TZ) if acc_from else None,
            "accepted_to": datetime.combine(acc_to + timedelta(days=1), datetime.min.time(), tzinfo=TH_TZ) if acc_to else None,
        }
        rows, _ = compute_allocation(db.session, filters)
        rows = _filter_out_cancelled_rows(rows)
        
        # *** [NEW LOGIC] กรองเฉพาะ Order ที่พิมพ์คลังแล้ว แต่ยังไม่พิมพ์หยิบ ***
        oids = sorted({(r.get("order_id") or "").strip() for r in rows if r.get("order_id")})
        warehouse_counts = _get_print_counts_local(oids, kind="warehouse")
        picking_counts = _get_print_counts_local(oids, kind="picking")
        
        valid_rows = []
        for r in rows:
            oid = (r.get("order_id") or "").strip()
            wh_count = int(warehouse_counts.get(oid, 0))
            pk_count = int(picking_counts.get(oid, 0))
            if wh_count > 0 and pk_count == 0:
                valid_rows.append(r)
        
        rows = valid_rows

        safe_rows = []
        for r in rows:
            r = dict(r)
            if "stock_qty" not in r:
                sku = (r.get("sku") or "").strip()
                stock_qty = 0
                if sku:
                    prod = Product.query.filter_by(sku=sku).first()
                    if prod and hasattr(prod, "stock_qty"):
                        try:
                            stock_qty = int(prod.stock_qty or 0)
                        except Exception:
                            stock_qty = 0
                    else:
                        st = Stock.query.filter_by(sku=sku).first()
                        stock_qty = int(st.qty) if st and st.qty is not None else 0
                r["stock_qty"] = stock_qty
            r["accepted"] = bool(r.get("accepted", False))
            r["logistic"] = r.get("logistic") or r.get("logistic_type") or "-"
            safe_rows.append(r)

        if logistic:
            safe_rows = [r for r in safe_rows if (r.get("logistic") or "").lower().find(logistic.lower()) >= 0]

        items = _aggregate_picking(safe_rows)

        valid_rows = [r for r in safe_rows if r.get("accepted") and r.get("allocation_status") in ("ACCEPTED", "READY_ACCEPT")]
        order_ids = sorted({(r.get("order_id") or "").strip() for r in valid_rows if r.get("order_id")})
        print_counts_pick = _get_print_counts_local(order_ids, "picking")
        print_count_overall = max(print_counts_pick.values()) if print_counts_pick else 0

        shop_name = ""
        if shop_id:
            s = Shop.query.get(int(shop_id))
            if s:
                shop_name = f"{s.platform} • {s.name}"

        for it in items:
            it["platform"] = platform or ""
            it["shop_name"] = shop_name or ""
            it["logistic"] = logistic or ""

        # Get dispatch_round data for items
        dispatch_rounds = {}
        if order_ids:
            tbl = _ol_table_name()
            sql = text(f"SELECT DISTINCT order_id, dispatch_round FROM {tbl} WHERE order_id IN :oids")
            sql = sql.bindparams(bindparam("oids", expanding=True))
            for row in db.session.execute(sql, {"oids": order_ids}).fetchall():
                if row[1] is not None:
                    dispatch_rounds[row[0]] = row[1]
        
        df = pd.DataFrame([{
            "แพลตฟอร์ม": it["platform"],
            "ร้าน": it["shop_name"],
            "SKU": it["sku"],
            "Brand": it["brand"],
            "สินค้า": it["model"],
            "ต้องหยิบ": it["need_qty"],
            "สต็อก": it["stock_qty"],
            "ขาด": it["shortage"],
            "คงเหลือหลังหยิบ": it["remain_after_pick"],
            "ประเภทขนส่ง": it["logistic"],
            "จ่ายงาน(รอบที่)": it.get("dispatch_round", ""),
            "พิมพ์แล้ว (ครั้ง)": 0,  # Current page: not printed yet
        } for it in items])

        out = BytesIO()
        with pd.ExcelWriter(out, engine="xlsxwriter") as w:
            df.to_excel(w, index=False, sheet_name="Picking List")
        out.seek(0)
        
        filename = f"ใบงานหยิบสินค้า_Picking_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        return send_file(out, as_attachment=True, download_name=filename)

    @app.route("/report/picking/history/export.xlsx")
    @login_required
    def export_picking_history_excel():
        """Export ใบงานหยิบสินค้าประวัติ - แสดงงานที่พิมพ์แล้ว"""
        # Check for reset mode
        reset_mode = request.args.get("reset")
        
        if reset_mode == 'today':
            target_date = now_thai().date()
            platform = None
            shop_id = None
            logistic = None
            print_date = None
            raw_from = None
            raw_to = None
        else:
            platform = normalize_platform(request.args.get("platform"))
            shop_id = request.args.get("shop_id")
            logistic = request.args.get("logistic")
            print_date = request.args.get("print_date")
            raw_from = request.args.get("accepted_from")
            raw_to = request.args.get("accepted_to")
            
            if print_date:
                try:
                    target_date = datetime.strptime(print_date, "%Y-%m-%d").date()
                except:
                    target_date = None
            else:
                target_date = None
        
        acc_from = parse_date_any(raw_from)
        acc_to = parse_date_any(raw_to)
        
        # Get printed orders
        tbl = _ol_table_name()
        
        if target_date:
            sql = text(f"""
                SELECT DISTINCT order_id 
                FROM {tbl} 
                WHERE printed_picking > 0 
                AND DATE(printed_picking_at) = :target_date
            """)
            result = db.session.execute(sql, {"target_date": target_date.isoformat()}).fetchall()
        else:
            sql = text(f"SELECT DISTINCT order_id FROM {tbl} WHERE printed_picking > 0")
            result = db.session.execute(sql).fetchall()
        
        printed_order_ids = [row[0] for row in result if row[0]]
        
        if not printed_order_ids:
            # Return empty file if no data
            df = pd.DataFrame(columns=["แพลตฟอร์ม", "ร้าน", "SKU", "Brand", "สินค้า", "ต้องหยิบ", "สต็อก", "ขาด", "คงเหลือหลังหยิบ", "ประเภทขนส่ง", "จ่ายงาน(รอบที่)", "พิมพ์แล้ว (ครั้ง)"])
            bio = BytesIO()
            with pd.ExcelWriter(bio, engine="xlsxwriter") as w:
                df.to_excel(w, index=False, sheet_name="Picking History")
            bio.seek(0)
            return send_file(bio, as_attachment=True, download_name="ใบงานหยิบสินค้าประวัติ_Empty.xlsx")
        
        # Get full data for printed orders
        filters = {
            "platform": platform if platform else None, 
            "shop_id": int(shop_id) if shop_id else None, 
            "import_date": None,
            "accepted_from": datetime.combine(acc_from, datetime.min.time(), tzinfo=TH_TZ) if acc_from else None,
            "accepted_to": datetime.combine(acc_to + timedelta(days=1), datetime.min.time(), tzinfo=TH_TZ) if acc_to else None,
        }
        rows, _ = compute_allocation(db.session, filters)
        rows = _filter_out_cancelled_rows(rows)
        
        # Filter to only printed orders
        safe_rows = []
        for r in rows:
            if (r.get("order_id") or "").strip() not in printed_order_ids:
                continue
            r = dict(r)
            if "stock_qty" not in r:
                sku = (r.get("sku") or "").strip()
                stock_qty = 0
                if sku:
                    prod = Product.query.filter_by(sku=sku).first()
                    if prod and hasattr(prod, "stock_qty"):
                        try:
                            stock_qty = int(prod.stock_qty or 0)
                        except Exception:
                            stock_qty = 0
                    else:
                        st = Stock.query.filter_by(sku=sku).first()
                        stock_qty = int(st.qty) if st and st.qty is not None else 0
                r["stock_qty"] = stock_qty
            r["accepted"] = bool(r.get("accepted", False))
            r["logistic"] = r.get("logistic") or r.get("logistic_type") or "-"
            safe_rows.append(r)
        
        if logistic:
            safe_rows = [r for r in safe_rows if (r.get("logistic") or "").lower().find(logistic.lower()) >= 0]
        
        # Aggregate by SKU
        items = _aggregate_picking(safe_rows)
        
        # Get print counts
        valid_rows = [r for r in safe_rows if r.get("accepted") and r.get("allocation_status") in ("ACCEPTED", "READY_ACCEPT")]
        order_ids = sorted({(r.get("order_id") or "").strip() for r in valid_rows if r.get("order_id")})
        print_counts_pick = _get_print_counts_local(order_ids, "picking")
        print_count_overall = max(print_counts_pick.values()) if print_counts_pick else 0
        
        # Shop name
        shop_name = ""
        if shop_id and reset_mode != 'today':
            s = Shop.query.get(int(shop_id))
            if s:
                shop_name = f"{s.platform} • {s.name}"
        
        # Fill in platform/shop/logistic for each item
        for it in items:
            it["platform"] = platform or "-"
            it["shop_name"] = shop_name or "-"
            it["logistic"] = logistic or "-"
        
        # Get dispatch_round data
        dispatch_rounds = {}
        if order_ids:
            sql = text(f"SELECT DISTINCT order_id, dispatch_round FROM {tbl} WHERE order_id IN :oids")
            sql = sql.bindparams(bindparam("oids", expanding=True))
            for row in db.session.execute(sql, {"oids": order_ids}).fetchall():
                if row[1] is not None:
                    dispatch_rounds[row[0]] = row[1]
        
        # Create DataFrame
        df = pd.DataFrame([{
            "แพลตฟอร์ม": it["platform"],
            "ร้าน": it["shop_name"],
            "SKU": it["sku"],
            "Brand": it["brand"],
            "สินค้า": it["model"],
            "ต้องหยิบ": it["need_qty"],
            "สต็อก": it["stock_qty"],
            "ขาด": it["shortage"],
            "คงเหลือหลังหยิบ": it["remain_after_pick"],
            "ประเภทขนส่ง": it["logistic"],
            "จ่ายงาน(รอบที่)": it.get("dispatch_round", ""),
            "พิมพ์แล้ว (ครั้ง)": print_count_overall,
        } for it in items])
        
        bio = BytesIO()
        with pd.ExcelWriter(bio, engine="xlsxwriter") as w:
            df.to_excel(w, index=False, sheet_name="Picking History")
        bio.seek(0)
        
        filename = f"ใบงานหยิบสินค้าประวัติ_History_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        return send_file(bio, as_attachment=True, download_name=filename)

    # -----------------------
    # ดาวน์โหลด Orders Excel Template (เดิม)
    # -----------------------
    @app.route("/download/orders-template")
    @login_required
    def download_orders_template():
        platform = normalize_platform(request.args.get("platform") or "Shopee")
        cols = ["ชื่อร้าน", "Order ID", "SKU", "Item Name", "Qty", "Order Time", "Logistics"]

        sample = pd.DataFrame(columns=cols)
        sample.loc[0] = ["Your Shop", "ORDER123", "SKU-001", "สินค้าทดลอง", 1, "2025-01-01 12:00", "J&T"]

        out = BytesIO()
        with pd.ExcelWriter(out, engine="xlsxwriter") as writer:
            sample.to_excel(writer, index=False, sheet_name=f"{platform} Orders")
        out.seek(0)
        return send_file(out, as_attachment=True, download_name=f"{platform}_Orders_Template.xlsx")

    # -----------------------
    # Admin clear
    # -----------------------
    @app.route("/admin/clear", methods=["GET","POST"])
    @login_required
    def admin_clear():
        cu = current_user()
        if not cu or cu.role != "admin":
            flash("เฉพาะแอดมินเท่านั้นที่สามารถล้างข้อมูลได้", "danger")
            return redirect(url_for("dashboard"))
        
        if request.method == "POST":
            scope = request.form.get("scope")
            
            if scope == "today":
                today = now_thai().date()
                
                # [แก้ไข] หา order_id ที่จะถูกลบก่อน เพื่อเอาไปลบในตาราง DeletedOrder ด้วย
                today_lines = db.session.query(OrderLine.order_id).filter(OrderLine.import_date == today).all()
                today_oids = list(set(r[0] for r in today_lines if r[0]))
                
                # ลบข้อมูลจริง
                deleted = OrderLine.query.filter(OrderLine.import_date == today).delete(synchronize_session=False)
                
                # [เพิ่ม] ลบข้อมูลในถังขยะที่เกี่ยวข้องกับ ID พวกนี้
                del_bin = 0
                if today_oids:
                    del_bin = db.session.query(DeletedOrder).filter(DeletedOrder.order_id.in_(today_oids)).delete(synchronize_session=False)

                db.session.commit()
                flash(f"ลบข้อมูลของวันนี้แล้ว ({deleted} รายการ, ถังขยะ {del_bin} รายการ)", "warning")
            
            elif scope == "date_range":
                d_from_str = request.form.get("date_from")
                d_to_str = request.form.get("date_to")
                
                # รับค่า Checkbox เป็น list (เช่น ['orders', 'sales'])
                targets = request.form.getlist("targets")
                
                if not d_from_str or not d_to_str:
                    flash("กรุณาระบุวันที่เริ่มต้นและสิ้นสุด", "danger")
                elif not targets:
                    flash("กรุณาติ๊กเลือกประเภทข้อมูลที่ต้องการลบอย่างน้อย 1 อย่าง", "warning")
                else:
                    try:
                        # 1. แปลงวันที่จาก String เป็น Date Object
                        d_from = datetime.strptime(d_from_str, "%Y-%m-%d").date()
                        d_to = datetime.strptime(d_to_str, "%Y-%m-%d").date()
                        
                        # สร้างตัวแปร DateTime สำหรับฟิลด์ที่เป็น timestamp (เริ่ม 00:00:00 ถึง 23:59:59)
                        dt_start = datetime.combine(d_from, datetime.min.time())
                        dt_end = datetime.combine(d_to, datetime.max.time())
                        
                        msg_parts = []

                        # 2. เช็คว่าติ๊ก "ออเดอร์" ไหม
                        if "orders" in targets:
                            # [แก้ไข] หา order_id ก่อนลบ เพื่อตามไปลบในถังขยะด้วย
                            lines_q = db.session.query(OrderLine.order_id).filter(
                                OrderLine.import_date >= d_from,
                                OrderLine.import_date <= d_to
                            )
                            target_oids = list(set(r[0] for r in lines_q.all() if r[0]))
                            
                            del_orders = OrderLine.query.filter(
                                OrderLine.import_date >= d_from,
                                OrderLine.import_date <= d_to
                            ).delete(synchronize_session=False)
                            
                            # ลบในถังขยะด้วย (Cascading delete logic)
                            del_bin = 0
                            if target_oids:
                                del_bin = db.session.query(DeletedOrder).filter(DeletedOrder.order_id.in_(target_oids)).delete(synchronize_session=False)

                            msg_parts.append(f"ออเดอร์ {del_orders} รายการ (ถังขยะ {del_bin})")

                        # 3. เช็คว่าติ๊ก "ใบสั่งขาย" ไหม
                        if "sales" in targets:
                            try:
                                if hasattr(Sales, 'import_date'):
                                    del_sales = Sales.query.filter(
                                        Sales.import_date >= d_from,
                                        Sales.import_date <= d_to
                                    ).delete(synchronize_session=False)
                                    msg_parts.append(f"ใบขาย {del_sales} รายการ")
                            except Exception:
                                pass

                        # 4. เช็คว่าติ๊ก "จ่ายงานแล้ว" ไหม
                        if "issued" in targets:
                            del_issued = IssuedOrder.query.filter(
                                IssuedOrder.issued_at >= dt_start,
                                IssuedOrder.issued_at <= dt_end
                            ).delete(synchronize_session=False)
                            msg_parts.append(f"จ่ายแล้ว {del_issued} รายการ")

                        # 5. เช็คว่าติ๊ก "ยกเลิก" ไหม
                        if "cancelled" in targets:
                            # [แก้ไข] แปลงเวลาไทย (Input) -> UTC (Database) เพื่อให้ลบได้ตรงช่วง
                            dt_start_utc = dt_start - timedelta(hours=7)
                            dt_end_utc = dt_end - timedelta(hours=7)
                            
                            del_cancelled = CancelledOrder.query.filter(
                                CancelledOrder.imported_at >= dt_start_utc,
                                CancelledOrder.imported_at <= dt_end_utc
                            ).delete(synchronize_session=False)
                            msg_parts.append(f"ยกเลิก {del_cancelled} รายการ")
                        
                        # [เพิ่ม] 6. เช็คว่าติ๊ก "ประวัติการลบ" ไหม
                        if "deleted" in targets:
                            del_deleted_log = db.session.query(DeletedOrder).filter(
                                DeletedOrder.deleted_at >= dt_start,
                                DeletedOrder.deleted_at <= dt_end
                            ).delete(synchronize_session=False)
                            msg_parts.append(f"ประวัติลบ {del_deleted_log} รายการ")

                        db.session.commit()
                        
                        if msg_parts:
                            flash(f"ลบข้อมูลช่วง {d_from_str} - {d_to_str} เรียบร้อย: " + ", ".join(msg_parts), "success")
                        else:
                            flash("ไม่ได้ลบข้อมูลใดๆ", "info")
                              
                    except Exception as e:
                        db.session.rollback()
                        app.logger.exception("Clear date range failed")
                        flash(f"เกิดข้อผิดพลาดในการลบ: {e}", "danger")
                
            elif scope == "all":
                deleted = OrderLine.query.delete()
                # [เพิ่ม] ล้างถังขยะด้วยเลย
                del_bin = db.session.query(DeletedOrder).delete()
                db.session.commit()
                flash(f"ลบข้อมูลออเดอร์ทั้งหมดแล้ว ({deleted} รายการ, ถังขยะ {del_bin} รายการ)", "danger")
            
            # --- [เพิ่ม] CASE: ล้างถังขยะอย่างเดียว ---
            elif scope == "deleted_bin":
                n = db.session.query(DeletedOrder).delete()
                db.session.commit()
                flash(f"ล้างถังขยะเรียบร้อย ({n} รายการ)", "success")
                
            elif scope == "cancelled":
                # Get all cancelled order IDs
                cancelled_orders = CancelledOrder.query.all()
                cancelled_order_ids = [co.order_id for co in cancelled_orders]
                
                if cancelled_order_ids:
                    # Delete OrderLine records
                    deleted_lines = OrderLine.query.filter(
                        OrderLine.order_id.in_(cancelled_order_ids)
                    ).delete(synchronize_session=False)
                    
                    # Delete CancelledOrder records
                    deleted_cancelled = CancelledOrder.query.delete()
                    
                    db.session.commit()
                    flash(f"ลบ Order ยกเลิกทั้งหมดแล้ว ({len(cancelled_order_ids)} ออเดอร์, {deleted_lines} รายการ)", "warning")
                else:
                    flash("ไม่พบ Order ยกเลิก", "info")
                    
            elif scope == "issued":
                # Get all issued order IDs
                issued_orders = IssuedOrder.query.all()
                issued_order_ids = [io.order_id for io in issued_orders]
                
                if issued_order_ids:
                    # Delete OrderLine records
                    deleted_lines = OrderLine.query.filter(
                        OrderLine.order_id.in_(issued_order_ids)
                    ).delete(synchronize_session=False)
                    
                    # Delete IssuedOrder records
                    deleted_issued = IssuedOrder.query.delete()
                    
                    db.session.commit()
                    flash(f"ลบ Order จ่ายแล้วทั้งหมดแล้ว ({len(issued_order_ids)} ออเดอร์, {deleted_lines} รายการ)", "warning")
                else:
                    flash("ไม่พบ Order จ่ายแล้ว", "info")
                    
            elif scope == "sales":
                # ลบข้อมูลในตาราง Sales ทั้งหมด
                deleted = db.session.query(Sales).delete()
                db.session.commit()
                flash(f"ลบข้อมูลใบสั่งขาย (Sales) ทั้งหมดแล้ว ({deleted} รายการ)", "danger")
            
            return redirect(url_for("admin_clear"))
        
        # GET request - show stats
        today = now_thai().date()
        stats = {
            "total_orders": db.session.query(func.count(func.distinct(OrderLine.order_id))).scalar() or 0,
            "cancelled_orders": CancelledOrder.query.count(),
            "issued_orders": IssuedOrder.query.count(),
            "deleted_orders": DeletedOrder.query.count(),  # [เพิ่ม] นับถังขยะ
            "today_orders": db.session.query(func.count(func.distinct(OrderLine.order_id))).filter(
                OrderLine.import_date == today
            ).scalar() or 0,
            "total_sales": Sales.query.count(),
        }
        
        return render_template("clear_confirm.html", stats=stats)

    return app


app = create_app()

if __name__ == "__main__":
    from waitress import serve
    port = int(os.environ.get("PORT", "8000"))
    serve(app, host="0.0.0.0", port=port)